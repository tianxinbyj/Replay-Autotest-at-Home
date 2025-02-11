"""
@Author: BU YUJUN
@Date: 2024/7/8 上午10:00  
"""
import glob
import json
import os.path
import shutil
import subprocess
import time
import uuid

import matplotlib.lines as mlines
import numpy as np
import openpyxl
import pandas as pd
import yaml
from PIL import Image
from matplotlib import patches as pc
from matplotlib import pyplot as plt, image as mpimg
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas
from matplotlib.colors import LinearSegmentedColormap
from openpyxl.styles import PatternFill, Border, Side, Alignment
from scipy.interpolate import interp1d
from spire.xls import *

from Utils.VideoProcess import image2video

plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from Libs import get_project_path, copy_to_destination

sys.path.append(get_project_path())

from Utils.Libs import test_encyclopaedia, create_folder, contains_chinese, get_string_display_length, project_path
from Utils.Libs import generate_unique_id, bench_config
from Utils.Libs import font_size, title_font, axis_font, legend_font
from Utils.Logger import send_log
from Utils.SSHClient import SSHClient
from Envs.Master.Modules.PDFReportTemplate import PDFReportTemplate

# 导入评测api
from Envs.Master.Modules.PerceptMetrics.PerceptMetrics import PreProcess, MatchTool, MetricEvaluator, MetricStatistics

scenario_test_record_path = os.path.join(project_path, 'Docs', 'Resources', 'scenario_info', 'scenario_test_record.csv')
kpi_target_file_path = os.path.join(project_path, 'Docs', 'Resources', 'ObstaclesKpi.xlsx')
kpi_target_threshold = pd.read_excel(kpi_target_file_path, sheet_name=0, header=[0, 1, 2], index_col=[0, 1])
kpi_target_ratio = pd.read_excel(kpi_target_file_path, sheet_name=1, header=[0, 1, 2], index_col=[0, 1])


def sync_test_result(method):
    def wrapper(self, *args, **kwargs):
        method_start_time = time.time()
        self.load_test_result()
        result = method(self, *args, **kwargs)
        self.save_test_result()
        method_end_time = time.time()
        send_log(self, f'{self.__class__.__name__}.{method.__name__} '
              f'-> {method_end_time - method_start_time:.2f} sec')
        return result

    return wrapper


def get_obstacles_kpi_threshold(col_1, col_2, target_type, x=0, y=0, region=None):

    def get_region_text(x, y):
        x_text = None
        if -100 < x <= -50:
            x_text = 'x(-100~-50)'
        elif -50 < x <= 0:
            x_text = 'x(-50~0)'
        elif 0 < x <= 50:
            x_text = 'x(0~50)'
        elif 50 < x <= 100:
            x_text = 'x(50~100)'
        elif 100 < x <= 150:
            x_text = 'x(100~150)'

        y_text = None
        if -8 <= y <= 8:
            y_text = 'y(-8~8)'

        if x_text is None or y_text is None:
            return None

        return f'{x_text},{y_text}'

    df = kpi_target_threshold
    if region is not None:
        region_text = region
    else:
        region_text = get_region_text(x, y)
        if region_text is None:
            return None

    index = (target_type, region_text)
    col = (col_1, col_2, '/VA/Obstacles')
    threshold = df.at[index, col]
    if np.isnan(threshold):
        return None
    else:
        return threshold


def get_obstacles_kpi_ratio(col_1, col_2, target_type, kpi_date_label):
    df = kpi_target_ratio
    index = (target_type, int(kpi_date_label))
    col = (col_1, col_2, '/VA/Obstacles')
    ratio = df.at[index, col]
    if np.isnan(ratio):
        return 1
    else:
        return ratio


class DataGrinderOneCase:

    def __init__(self, scenario_unit_folder):
        send_log(self, '=' * 25 + self.__class__.__name__ + '=' * 25)
        # 变量初始化
        self.gt_ego_flag = False
        self.pred_ego_flag = False
        self.ego_velocity_generator = None
        self.cut_frame_offset = - 1.05

        scenario_config_yaml = os.path.join(scenario_unit_folder, 'TestConfig.yaml')
        with open(scenario_config_yaml, 'r', encoding='utf-8') as file:
            self.test_config = yaml.safe_load(file)
        self.scenario_id = self.test_config['scenario_id']
        send_log(self, '=' * 25 + f'{self.scenario_id}开始数据处理' + '=' * 25)

        # 加载测试相关的参数
        self.scenario_unit_folder = scenario_unit_folder
        self.product = self.test_config['product']
        self.version = self.test_config['version']
        self.test_date = str(self.test_config['test_date'])
        self.kpi_date_label = self.test_config['kpi_date_label']
        self.truth_source = self.test_config['truth_source']

        self.scenario_tag = self.test_config['scenario_tag']
        self.test_action = self.test_config['test_action']
        self.test_encyclopaedia = test_encyclopaedia[self.product]
        self.test_item = self.test_config['test_item']
        self.test_topic = self.test_config['test_topic']
        self.test_information = test_encyclopaedia['Information'][self.test_topic]
        self.topics_for_evaluation = [
            topic for topic in self.test_item.keys() if topic in self.test_information['topics']]

        # 不同产品用不同的相机
        if self.product == 'ES37':
            # self.camera_list = [
            #     'CAM_FRONT_LEFT', 'CAM_FRONT_120', 'CAM_FRONT_RIGHT',
            #     'CAM_BACK_LEFT', 'CAM_BACK', 'CAM_BACK_RIGHT',
            # ]
            self.camera_list = [
                'CAM_FISHEYE_LEFT', 'CAM_FRONT_120', 'CAM_FISHEYE_RIGHT',
                'CAM_FISHEYE_FRONT', 'CAM_BACK', 'CAM_FISHEYE_BACK',
            ]
        else:
            self.camera_list = [
                'CAM_FISHEYE_LEFT', 'CAM_FRONT_120', 'CAM_FISHEYE_RIGHT',
                'CAM_FISHEYE_FRONT', 'CAM_BACK', 'CAM_FISHEYE_BACK',
            ]

        # 初始化文件夹
        self.pred_raw_folder = self.test_config['pred_folder']
        self.gt_raw_folder = self.test_config['gt_folder']
        self.DataFolder = os.path.join(self.scenario_unit_folder, '01_Data')
        self.BugFolder = os.path.join(self.scenario_unit_folder, '02_Bug')
        self.RenderFolder = os.path.join(self.scenario_unit_folder, '03_Render')

        # 初始化测试配置
        self.test_result_yaml = os.path.join(self.scenario_unit_folder, 'TestResult.yaml')
        if not os.path.exists(self.test_result_yaml):
            self.test_result = {'General': {}, self.test_topic: {'GroundTruth': {}}}
            for topic in self.topics_for_evaluation:
                self.test_result[self.test_topic][topic] = {}

            self.save_test_result()

        test_encyclopaedia_yaml = os.path.join(self.scenario_unit_folder, 'TestEncyclopaedia.yaml')
        with open(test_encyclopaedia_yaml, 'w', encoding='utf-8') as f:
            yaml.dump(self.test_encyclopaedia,
                      f, encoding='utf-8', allow_unicode=True, sort_keys=False)

        test_information_yaml = os.path.join(self.scenario_unit_folder, 'TestInformation.yaml')
        with open(test_information_yaml, 'w', encoding='utf-8') as f:
            yaml.dump(self.test_information,
                      f, encoding='utf-8', allow_unicode=True, sort_keys=False)

    @sync_test_result
    def load_pred_data(self):
        test_topic_info_path = os.path.join(self.pred_raw_folder, 'RawData', 'TestTopicInfo.yaml')
        if not os.path.exists(test_topic_info_path):
            send_log(self, 'No TestTopicInfo.yaml is found. Please check')
            return

        with open(test_topic_info_path) as f:
            topics_for_parser = yaml.load(f, Loader=yaml.FullLoader)['topics_for_parser']

        # 初始化感知数据
        for topic in topics_for_parser:
            topic_tag = topic.replace('/', '')

            if topic in self.topics_for_evaluation:
                send_log(self, f'Prediction 正在读取{topic}')

                raw_column = self.test_information['raw_column']

                # 读取原始数据
                csv_list = glob.glob(os.path.join(self.pred_raw_folder, 'RawData', f'{topic_tag}*.csv'))
                csv_data = []
                for csv_file in sorted(csv_list, reverse=False):
                    if 'hz' not in csv_file:
                        csv_data.append(pd.read_csv(csv_file, index_col=False))
                pred_data = pd.concat(csv_data).sort_values(by=['time_stamp'])[raw_column]

                # 读取时间辍, 用于时间辍匹配
                pred_timestamp_csv = glob.glob(os.path.join(self.pred_raw_folder, 'RawData', f'{topic_tag}*hz.csv'))[0]
                pred_timestamp = pd.read_csv(pred_timestamp_csv, index_col=False).sort_values(
                    by=['time_stamp']).drop_duplicates(subset=['time_stamp'], keep='first')
                pred_hz = float(os.path.basename(pred_timestamp_csv).split('_')[1])

                raw_folder = os.path.join(self.DataFolder, self.test_topic, topic_tag, 'raw')
                create_folder(raw_folder, False)
                self.test_result[self.test_topic][topic] = {
                    'frequency': pred_hz, 'raw': {},
                }

                path = os.path.join(raw_folder, 'pred_data.csv')
                pred_data.to_csv(path, index=False, encoding='utf_8_sig')
                self.test_result[self.test_topic][topic]['raw']['pred_data'] = self.get_relpath(path)

                path = os.path.join(raw_folder, 'pred_timestamp.csv')
                pred_timestamp.to_csv(path, index=False, encoding='utf_8_sig')
                self.test_result[self.test_topic][topic]['raw']['pred_timestamp'] = self.get_relpath(path)

            elif topic == '/PI/EG/EgoMotionInfo':
                send_log(self, f'Prediction 正在读取{topic}, 用于时间同步')

                # 读取原始数据
                csv_list = glob.glob(os.path.join(self.pred_raw_folder, 'RawData', f'{topic_tag}*.csv'))
                csv_data = []
                for csv_file in sorted(csv_list, reverse=False):
                    if 'hz' not in csv_file:
                        csv_data.append(pd.read_csv(csv_file, index_col=False))
                pred_data = pd.concat(csv_data).sort_values(by=['time_stamp']).iloc[50:]
                create_folder(os.path.join(self.DataFolder, 'General'), update=False)
                path = os.path.join(self.DataFolder, 'General', 'pred_ego.csv')
                self.pred_ego_flag = True
                pred_data[['time_stamp', 'ego_vx']].to_csv(path, index=False, encoding='utf_8_sig')
                self.test_result['General']['pred_ego'] = self.get_relpath(path)

            # elif topic == '/VA/VehicleMotionIpd':
            #     send_log(self, f'Prediction 正在读取{topic}, 用于时间同步')
            #
            #     # 读取原始数据
            #     csv_list = glob.glob(os.path.join(self.pred_raw_folder, 'RawData', f'{topic_tag}*.csv'))
            #     csv_data = []
            #     for csv_file in sorted(csv_list, reverse=False):
            #         if 'hz' not in csv_file:
            #             csv_data.append(pd.read_csv(csv_file, index_col=False))
            #     pred_data = pd.concat(csv_data).sort_values(by=['time_stamp']).iloc[50:]
            #     pred_data['ego_vx'] = pred_data['vehicle_speed']
            #     create_folder(os.path.join(self.DataFolder, 'General'), update=False)
            #     path = os.path.join(self.DataFolder, 'General', 'pred_ego.csv')
            #     pred_data[['time_stamp', 'ego_vx']].to_csv(path, index=False, encoding='utf_8_sig')
            #     self.test_result['General']['pred_ego'] = self.get_relpath(path)

            # 拿到/SA/INSPVA得到的时间差
            elif topic == '/SA/INSPVA':
                send_log(self, f'Prediction 正在读取{topic}, 用于时间同步的备选')

                # 读取原始数据
                csv_list = glob.glob(os.path.join(self.pred_raw_folder, 'RawData', f'{topic_tag}*.csv'))
                csv_data = []
                for csv_file in sorted(csv_list, reverse=False):
                    if 'hz' not in csv_file:
                        csv_data.append(pd.read_csv(csv_file, index_col=False))
                pred_data = pd.concat(csv_data).sort_values(by=['time_stamp']).iloc[50:]
                if len(pred_data):
                    self.test_result['General']['sa_time_gap'] \
                        = float(pred_data['time_stamp'].mean() - pred_data['header_stamp'].mean())

        # 复制场景相关的信息, 解析相机参数
        scenario_info_folder = os.path.join(self.pred_raw_folder, 'scenario_info')
        if copy_to_destination(scenario_info_folder, self.scenario_unit_folder):

            self.test_result['General']['camera_position'] = {}
            yaml_folder = os.path.join(scenario_info_folder, 'yaml_calib')
            cam_description_path = os.path.join(yaml_folder, 'cam_description.yaml')

            if os.path.exists(cam_description_path):
                # 说明是老样式
                with open(cam_description_path, 'r', encoding='utf-8') as f:
                    cam_description = yaml.safe_load(f)

                for i, cam_name in cam_description.items():
                    cam_par_path = os.path.join(yaml_folder, f'camera_{i}.yaml')
                    with open(cam_par_path, 'r', encoding='utf-8') as f:
                        cam_par = yaml.safe_load(f)
                        x, y, z = [
                            float(cam_par[f'cam_{i}_pos_x']),
                            float(cam_par[f'cam_{i}_pos_y']),
                            float(cam_par[f'cam_{i}_pos_z']),
                        ]
                        image_width = float(cam_par[f'cam_{i}_image_width'])
                        focal_x = float(cam_par[f'cam_{i}_focal_x'])
                        yaw = np.rad2deg(float(cam_par[f'cam_{i}_yaw']))
                        h_fov = np.rad2deg(2 * np.arctan(image_width / 2 / focal_x))

                        fov_left = float(yaw + h_fov / 2)
                        fov_right = float(yaw - h_fov / 2)

                        if fov_left < 0:
                            fov_left += 360
                        if fov_left > 360:
                            fov_left -= 360

                        if fov_right < 0:
                            fov_right += 360
                        if fov_right > 360:
                            fov_right -= 360

                        if fov_right > fov_left:
                            fov_right -= 360
                        fov_range = [fov_right, fov_left]

                    self.test_result['General']['camera_position'][cam_name] = {
                        'x': x, 'y': y, 'z': z, 'fov': fov_range,
                    }
                    send_log(self, f'{cam_name} 位于({x}, {y}, {z}, {fov_range})')

            else:
                for cam_par_path in glob.glob(os.path.join(yaml_folder, '*.yaml')):
                    cam_name = os.path.splitext(os.path.basename(cam_par_path))[0]

                    with open(cam_par_path, 'r', encoding='utf-8') as f:
                        cam_par = yaml.safe_load(f)
                        x, y, z = [
                            float(cam_par[f'pos_x']),
                            float(cam_par[f'pos_y']),
                            float(cam_par[f'pos_z']),
                        ]
                        image_width = float(cam_par[f'image_width'])
                        focal_x = float(cam_par[f'focal_x'])
                        yaw = np.rad2deg(float(cam_par[f'yaw']))
                        h_fov = np.rad2deg(2 * np.arctan(image_width / 2 / focal_x))

                        fov_left = float(yaw + h_fov / 2)
                        fov_right = float(yaw - h_fov / 2)

                        if fov_left < 0:
                            fov_left += 360
                        if fov_left > 360:
                            fov_left -= 360

                        if fov_right < 0:
                            fov_right += 360
                        if fov_right > 360:
                            fov_right -= 360

                        if fov_right > fov_left:
                            fov_right -= 360
                        fov_range = [fov_right, fov_left]

                    self.test_result['General']['camera_position'][cam_name] = {
                        'x': x, 'y': y, 'z': z, 'fov': fov_range,
                    }
                    send_log(self, f'{cam_name} 位于('
                                   f'{round(x, 2)}, {round(y, 2)}, {round(z, 2)}, '
                                   f'{fov_range})')

        new_scenario_info_folder = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo')
        if os.path.exists(new_scenario_info_folder):
            shutil.rmtree(new_scenario_info_folder)
        os.rename(os.path.join(self.scenario_unit_folder, 'scenario_info'), new_scenario_info_folder)

    @sync_test_result
    def load_gt_data(self):
        gt_config_path = os.path.join(self.gt_raw_folder, 'yaml_management.yaml')
        if not os.path.exists(gt_config_path):
            send_log(self, 'No yaml_management.yaml is found. Please check')
            return

        with open(gt_config_path) as f:
            gt_config = yaml.load(f, Loader=yaml.FullLoader)

        if self.test_topic == 'Obstacles':
            gt_topic = 'od'
        elif self.test_topic == 'Lines':
            gt_topic = 'line'
        else:
            send_log(self, '未知的测试主题')
            return

        flag = '{:s}_csv_flag'.format(gt_topic)
        timestamp_csv_tag = '{:s}_timestamp'.format(gt_topic.split('_')[0])

        if not (flag in gt_config.keys() and gt_config[flag]):
            send_log(self, f'GroundTruth 未找到{gt_topic}对应的真值文件')
            return

        send_log(self, f'GroundTruth 正在读取{gt_topic}')

        # 读取原始数据
        csv_data = []
        for f in os.listdir(os.path.join(self.gt_raw_folder, f'{gt_topic}_csv')):
            csv_path = os.path.join(self.gt_raw_folder, f'{gt_topic}_csv', f)
            csv_data.append(pd.read_csv(csv_path, index_col=False))
        gt_data = pd.concat(csv_data)

        if self.test_topic == 'Obstacles':
            gt_data = gt_data.rename(columns={
                'timestamp': 'time_stamp',
                'od_json_sequence': 'frame_id',
                'subtype': 'sub_type',
                'type_conf_3d': 'confidence',
            }).sort_values(by=['time_stamp'])
            gt_data['vx_rel'] = gt_data['vx'] - gt_data['ego_v']
            gt_data['vy_rel'] = gt_data['vy']
            gt_data = gt_data[gt_data['type'].isin([1, 2, 18])]

        elif self.test_topic == 'Lines':
            gt_data = gt_data.rename(columns={
                'timestamp': 'time_stamp',
                'line_json_sequence': 'frame_id',
                'x_list': 'x_points',
                'y_list': 'y_points',
            }).sort_values(by=['time_stamp'])
            gt_data['confidence'] = 0.5

        # 读取时间辍, 用于时间辍匹配
        gt_timestamp_csv = \
            glob.glob(os.path.join(self.gt_raw_folder, f'{timestamp_csv_tag}*hz.csv'))[0]
        gt_timestamp = pd.read_csv(gt_timestamp_csv).sort_values(
            by=['timestamp']).rename(columns={'timestamp': 'time_stamp'}) \
            .drop_duplicates(subset=['time_stamp'], keep='first')
        gt_hz = float(os.path.basename(gt_timestamp_csv).split('_')[2])

        raw_column = self.test_information['raw_column']

        raw_folder = os.path.join(self.DataFolder, self.test_topic, 'GroundTruth', 'raw')
        create_folder(raw_folder, False)
        self.test_result[self.test_topic]['GroundTruth'] = {
            'frequency': gt_hz, 'raw': {},
        }

        path = os.path.join(raw_folder, 'gt_data.csv')
        gt_data[raw_column].to_csv(path, index=False, encoding='utf_8_sig')
        self.test_result[self.test_topic]['GroundTruth']['raw']['gt_data'] = self.get_relpath(path)

        path = os.path.join(raw_folder, 'gt_timestamp.csv')
        gt_timestamp.to_csv(path, index=False, encoding='utf_8_sig')
        self.test_result[self.test_topic]['GroundTruth']['raw']['gt_timestamp'] = self.get_relpath(path)

        # 自车速度用于对齐
        ego_data_path = os.path.join(self.gt_raw_folder, 'od_ego.csv')
        if not os.path.exists(ego_data_path):
            send_log(self, 'GroundTruth未找到自车速度数据')
            self.gt_ego_flag = False
            ego_data = pd.DataFrame(columns=['time_stamp', 'ego_vx'])
            ego_data['time_stamp'] = gt_timestamp['time_stamp']
            ego_data['ego_vx'] = 0
        else:
            send_log(self, f'GroundTruth 正在读取{os.path.basename(ego_data_path)}, 用于时间同步')
            self.gt_ego_flag = True
            ego_data = pd.read_csv(os.path.join(self.gt_raw_folder, 'od_ego.csv'))[
                ['ego_timestamp', 'INS_Speed']].rename(
                columns={'ego_timestamp': 'time_stamp', 'INS_Speed': 'ego_vx'})
            create_folder(os.path.join(self.DataFolder, 'General'), update=False)
            path = os.path.join(self.DataFolder, 'General', 'gt_ego.csv')

        ego_data.to_csv(path, index=False, encoding='utf_8_sig')
        self.test_result['General']['gt_ego'] = self.get_relpath(path)

    @sync_test_result
    def sync_timestamp(self):

        # 调用计算时间差的接口
        baseline_data_path = self.get_abspath(self.test_result['General']['gt_ego'])
        baseline_data = pd.read_csv(baseline_data_path, index_col=False)
        baseline_time_series = baseline_data['time_stamp'].to_list()
        calibrated_data_path = self.get_abspath(self.test_result['General']['pred_ego'])
        calibrated_data = pd.read_csv(calibrated_data_path, index_col=False)
        calibrated_time_series = calibrated_data['time_stamp'].to_list()

        cmd = [
            f"{bench_config['master']['sys_interpreter']}",
            "Api_GetTimeGap.py",
            "-b", baseline_data_path,
            "-c", calibrated_data_path
        ]

        print(cmd)
        cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
        result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
        if result.stderr and 'warning' not in result.stderr:
            send_log(self, f'Api_GetTimeGap 发生错误 {result.stderr}')
        t_delta, v_error = result.stdout.strip().split('\n')[-1].split(' ')
        t_delta = float(t_delta)

        send_log(self, f'使用速度平移获得的最佳时间间隔 = {t_delta}, 平均速度误差 = {v_error}')
        self.test_result['General']['vel_time_gap'] = t_delta

        if 'sa_time_gap' in self.test_result['General']:
            t_delta = self.test_result['General']['sa_time_gap']

        self.test_result['General']['time_gap'] = t_delta

        time_start = max(min(calibrated_time_series) + t_delta, min(baseline_time_series)) + 5
        time_end = min(max(calibrated_time_series) + t_delta, max(baseline_time_series)) - 5
        send_log(self, f'测试指标的数据时间段为{time_start}-{time_end}')
        self.test_result['General']['time_start'] = float(time_start)
        self.test_result['General']['time_end'] = float(time_end)

        # 将同步后的自车速度保存, 后续插值车速需要
        calibrated_data['time_stamp'] += t_delta
        sync_ego_data = calibrated_data[(calibrated_data['time_stamp'] <= time_end + 1)
                                        & (calibrated_data['time_stamp'] >= time_start - 1)]
        path = os.path.join(self.DataFolder, 'General', 'SyncEgoVx.csv')
        sync_ego_data.to_csv(path, index=False, encoding='utf_8_sig')
        self.test_result['General']['sync_ego_data'] = self.get_relpath(path)

        # 将同步后的自车速度可视化
        fig = plt.figure(figsize=(10, 5.625))
        fig.tight_layout()
        plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1)
        grid = plt.GridSpec(1, 1, wspace=0.2, hspace=0.25)
        ax = fig.add_subplot(grid[0, 0])

        ax.plot(calibrated_data['time_stamp'],
                calibrated_data['ego_vx'] * 3.6,
                color='red', linestyle='dashed', linewidth=2,
                label=f'ego_speed of {self.product}')
        ax.plot(baseline_data['time_stamp'],
                baseline_data['ego_vx'] * 3.6,
                color='green', linestyle='dashed', linewidth=2,
                label=f'ego_speed of GroundTruth')
        ax.set_title(f'{self.scenario_id}  time_gap = {t_delta} sec', fontsize=font_size * 1.3)
        ax.set_xlabel('time[second]', fontdict=axis_font)
        ax.set_ylabel('ego_vx[km/h]', fontdict=axis_font)
        ax.set_xlim(time_start, time_end)
        ax.set_ylim(0, 120)
        ax.set_yticks(np.arange(0, 121, 20))
        ax.grid('--', color='gainsboro')
        ax.tick_params(direction='out', labelsize=font_size / 1.1, length=2)
        ax.legend(loc=0, prop=legend_font)

        path = os.path.join(self.DataFolder, 'General', 'SyncEgoVx.png')
        canvas = FigureCanvas(fig)
        canvas.print_figure(path, facecolor='white', dpi=100)
        self.test_result['General']['sync_ego_figure'] = self.get_relpath(path)
        fig.clf()
        plt.close()

    @sync_test_result
    def promote_additional(self):
        time_gap = self.test_result['General']['time_gap']
        time_start = self.test_result['General']['time_start']
        time_end = self.test_result['General']['time_end']

        additional_column = self.test_information['raw_column'] + self.test_information['additional_column']
        send_log(self, f'{self.test_topic}, 使用{self.test_topic}Preprocess')
        preprocess_instance = eval(f'PreProcess.{self.test_topic}Preprocess()')

        # 先处理真值
        raw = self.test_result[self.test_topic]['GroundTruth']['raw']
        additional_folder = os.path.join(self.DataFolder, self.test_topic, 'GroundTruth', 'additional')
        create_folder(additional_folder)
        if 'additional' not in self.test_result[self.test_topic]['GroundTruth']:
            self.test_result[self.test_topic]['GroundTruth']['additional'] = {}

        # 时间辍补齐
        send_log(self, f'{self.test_topic} GroundTruth 时间辍同步')
        data = pd.read_csv(self.get_abspath(raw['gt_timestamp']), index_col=False)
        data = data[(data['time_stamp'] <= time_end) & (data['time_stamp'] >= time_start)]
        path = os.path.join(additional_folder, 'gt_timestamp.csv')
        self.test_result[self.test_topic]['GroundTruth']['additional']['gt_timestamp'] = self.get_relpath(
            path)
        data.to_csv(path, index=False, encoding='utf_8_sig')

        # 预处理原始数据, 增加列
        send_log(self, f'{self.test_topic} GroundTruth 预处理步骤 {preprocess_instance.preprocess_types}')
        data = pd.read_csv(self.get_abspath(raw['gt_data']), index_col=False)
        data = data[(data['time_stamp'] <= time_end) & (data['time_stamp'] >= time_start)]
        path = os.path.join(additional_folder, 'gt_data.csv')
        data.to_csv(path, index=False, encoding='utf_8_sig')
        self.test_result[self.test_topic]['GroundTruth']['additional']['gt_data'] = {}

        # 按照每个topic的ROI预处理真值
        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth':
                continue

            topic_tag = topic.replace('/', '')

            if self.test_topic == 'Obstacles':
                input_parameter_container = {
                    'camera': self.test_result['General']['camera_position'],
                    'coverage_reference_point': self.test_config['coverage_reference_point'],
                    'coverage_threshold': self.test_config['coverage_threshold'],
                    'ROI': self.test_config['detected_ROI'][topic],
                    'lane_width': 3.6,
                    'moving_threshold': 2,
                    'key_coverage_threshold': 0.1,
                    'test_topic': self.test_topic,
                }
            elif self.test_topic == 'Lines':
                input_parameter_container = {
                    'lane_width': 3.6,
                    'test_topic': self.test_topic,
                    'if_gt': True,
                }
            else:
                return

            parameter_json_path = os.path.join(get_project_path(), 'Temp', 'process_api_parameter.json')
            with open(parameter_json_path, 'w', encoding='utf-8') as f:
                json.dump(input_parameter_container, f, ensure_ascii=False, indent=4)
            topic_gt_data_path = os.path.join(additional_folder, f'{topic_tag}_gt_data.csv')
            self.test_result[self.test_topic]['GroundTruth']['additional']['gt_data'][topic] \
                = self.get_relpath(topic_gt_data_path)

            cmd = [
                f"{bench_config['master']['sys_interpreter']}",
                "Api_ProcessRawData.py",
                "-r", path,
                "-j", parameter_json_path,
                "-p", topic_gt_data_path,
            ]

            print(cmd)
            cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
            result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
            # os.remove(parameter_json_path)
            if result.stderr and 'warning' not in result.stderr:
                send_log(self, f'ProcessRawData 发生错误 {result.stderr}')

            data = pd.read_csv(topic_gt_data_path, index_col=False)[additional_column]
            data.to_csv(topic_gt_data_path, index=False, encoding='utf_8_sig')

        # 按照每个topic的ROI预处理感知结果
        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth':
                continue

            raw = self.test_result[self.test_topic][topic]['raw']
            topic_tag = topic.replace('/', '')
            additional_folder = os.path.join(self.DataFolder, self.test_topic, topic_tag, 'additional')
            create_folder(additional_folder)
            if 'additional' not in self.test_result[self.test_topic][topic]:
                self.test_result[self.test_topic][topic]['additional'] = {}

            # 时间辍补齐
            send_log(self, f'{self.test_topic} {topic} 时间辍同步')
            data = pd.read_csv(self.get_abspath(raw['pred_timestamp']), index_col=False)
            data['time_stamp'] += time_gap
            data = data[(data['time_stamp'] <= time_end) & (data['time_stamp'] >= time_start)]
            path = os.path.join(additional_folder, 'pred_timestamp.csv')
            self.test_result[self.test_topic][topic]['additional']['pred_timestamp'] = self.get_relpath(path)
            data.to_csv(path, index=False, encoding='utf_8_sig')

            # 预处理原始数据, 增加列
            send_log(self, f'{self.test_topic} {topic} 预处理步骤 {preprocess_instance.preprocess_types}')
            data = pd.read_csv(self.get_abspath(raw['pred_data']), index_col=False)

            # 如果为空，比如说没有探测到任何目标，需要额外处理
            path = os.path.join(additional_folder, 'pred_data.csv')
            self.test_result[self.test_topic][topic]['additional']['pred_data'] = self.get_relpath(path)
            if len(data):
                data['time_stamp'] += time_gap
                data = data[(data['time_stamp'] <= time_end) & (data['time_stamp'] >= time_start)]
                data.to_csv(path, index=False, encoding='utf_8_sig')

                if self.test_topic == 'Obstacles':
                    input_parameter_container = {
                        'camera': self.test_result['General']['camera_position'],
                        'coverage_reference_point': self.test_config['coverage_reference_point'],
                        'coverage_threshold': self.test_config['coverage_threshold'],
                        'ROI': self.test_config['detected_ROI'][topic],
                        'lane_width': 3.6,
                        'moving_threshold': 2,
                        'key_coverage_threshold': 0.1,
                        'test_topic': self.test_topic,
                    }
                elif self.test_topic == 'Lines':
                    input_parameter_container = {
                        'lane_width': 3.6,
                        'test_topic': self.test_topic,
                        'if_gt': False,
                    }
                else:
                    return

                parameter_json_path = os.path.join(get_project_path(), 'Temp', 'process_api_parameter.json')
                with open(parameter_json_path, 'w', encoding='utf-8') as f:
                    json.dump(input_parameter_container, f, ensure_ascii=False, indent=4)

                cmd = [
                    f"{bench_config['master']['sys_interpreter']}",
                    "Api_ProcessRawData.py",
                    "-r", path,
                    "-j", parameter_json_path,
                    "-p", path,
                ]

                print(cmd)
                cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
                result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
                # os.remove(parameter_json_path)
                if result.stderr and 'warning' not in result.stderr:
                    send_log(self, f'ProcessRawData 发生错误 {result.stderr}')

                data = pd.read_csv(path, index_col=False)[additional_column]

            else:
                data = pd.DataFrame(columns=additional_column)

            data.to_csv(path, index=False, encoding='utf_8_sig')

    @sync_test_result
    def match_timestamp(self):
        gt_timestamp_path = self.get_abspath(
            self.test_result[self.test_topic]['GroundTruth']['additional']['gt_timestamp'])
        gt_timestamp = pd.read_csv(gt_timestamp_path, index_col=False)['time_stamp'].to_list()
        gt_hz = self.test_result[self.test_topic]['GroundTruth']['frequency']

        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth':
                continue

            additional = self.test_result[self.test_topic][topic]['additional']
            topic_tag = topic.replace('/', '')
            match_folder = os.path.join(self.DataFolder, self.test_topic, topic_tag, 'match')
            create_folder(match_folder)
            if 'match' not in self.test_result[self.test_topic][topic]:
                self.test_result[self.test_topic][topic]['match'] = {}

            pred_timestamp_path = self.get_abspath(additional['pred_timestamp'])
            pred_timestamp = pd.read_csv(pred_timestamp_path, index_col=False)['time_stamp'].to_list()
            pred_hz = self.test_result[self.test_topic][topic]['frequency']

            match_tolerance = self.test_config['timestamp_matching_tolerance'] / max(pred_hz, gt_hz)
            send_log(self, f'{self.test_topic} {topic} 时间差低于{match_tolerance} sec的尝试匹配, '
                           f'进一步选择局部最优')
            path = os.path.join(match_folder, 'match_timestamp.csv')

            cmd = [
                f"{bench_config['master']['sys_interpreter']}",
                "Api_MatchTimestamp.py",
                "-p", pred_timestamp_path,
                "-g", gt_timestamp_path,
                "-t", str(match_tolerance),
                "-f", path
            ]

            print(cmd)
            cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
            result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
            if result.stderr and 'warning' not in result.stderr:
                send_log(self, f'MatchTimestamp 发生错误 {result.stderr}')

            match_timestamp_data = pd.read_csv(path, index_col=False)
            send_log(self, f'{self.test_topic} {topic} GroundTruth 时间戳总计{len(gt_timestamp)}个, '
                           f'对齐{len(match_timestamp_data)} '
                           f'比例{len(match_timestamp_data) / len(gt_timestamp):.2%}')
            send_log(self, f'{self.test_topic} {topic} Prediction 时间戳总计{len(pred_timestamp)}个, '
                           f'对齐{len(match_timestamp_data)} '
                           f'比例{len(match_timestamp_data) / len(pred_timestamp):.2%}')

            self.test_result[self.test_topic][topic]['match']['match_timestamp'] = self.get_relpath(path)
            self.test_result[self.test_topic][topic]['match_frequency'] \
                = round(self.test_result[self.test_topic]['GroundTruth']['frequency']
                        * len(match_timestamp_data) / len(gt_timestamp), 2)

    @sync_test_result
    def match_object(self):
        additional_column = self.test_information['raw_column'] + self.test_information['additional_column']
        send_log(self, f'{self.test_topic}, 使用{self.test_topic}MatchTool')
        match_column = ['corresponding_index', 'gt.flag', 'pred.flag']
        for col in additional_column:
            for kind in ['gt', 'pred']:
                match_column.append(f'{kind}.{col}')

        if self.test_topic == 'Lines':
            match_column.append('IOU')

        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth':
                continue

            topic_tag = topic.replace('/', '')
            match_folder = os.path.join(self.DataFolder, self.test_topic, topic_tag, 'match')

            match_timestamp_path = self.get_abspath(
                self.test_result[self.test_topic][topic]['match']['match_timestamp'])

            pred_data_path = self.get_abspath(
                self.test_result[self.test_topic][topic]['additional']['pred_data'])

            gt_data_path = self.get_abspath(
                self.test_result[self.test_topic]['GroundTruth']['additional']['gt_data'][topic])

            if self.test_topic == 'Obstacles':
                input_parameter_container = {
                    'object_matching_tolerance': self.test_config['object_matching_tolerance'],
                    'test_topic': self.test_topic,
                }
            elif self.test_topic == 'Lines':
                input_parameter_container = {
                    'lane_matching_width': self.test_config['lane_matching_width'],
                    'test_topic': self.test_topic,
                }
            else:
                return

            parameter_json_path = os.path.join(get_project_path(), 'Temp', 'match_api_parameter.json')
            with open(parameter_json_path, 'w', encoding='utf-8') as f:
                json.dump(input_parameter_container, f, ensure_ascii=False, indent=4)
            path = os.path.join(match_folder, 'match_data.csv')

            cmd = [
                f"{bench_config['master']['sys_interpreter']}",
                "Api_MatchData.py",
                "-p", pred_data_path,
                "-g", gt_data_path,
                "-t", match_timestamp_path,
                "-j", parameter_json_path,
                "-m", path,
            ]

            print(cmd)
            cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
            result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
            # os.remove(parameter_json_path)
            if result.stderr and 'warning' not in result.stderr:
                send_log(self, f'MatchObstacles 发生错误 {result.stderr}')

            send_log(self, f'{self.test_topic} {topic} 目标匹配')
            data = pd.read_csv(path, index_col=False)[match_column]
            self.test_result[self.test_topic][topic]['match']['match_data'] = self.get_relpath(path)
            data.to_csv(path, index=False, encoding='utf_8_sig')

    def get_relpath(self, path: str) -> str:
        return os.path.relpath(path, self.scenario_unit_folder)

    def get_abspath(self, path: str) -> str:
        return os.path.join(self.scenario_unit_folder, path)

    def save_test_result(self):
        with open(self.test_result_yaml, 'w', encoding='utf-8') as f:
            yaml.dump(self.test_result,
                      f, encoding='utf-8', allow_unicode=True, sort_keys=False)

    def load_test_result(self):
        with open(self.test_result_yaml) as f:
            self.test_result = yaml.load(f, Loader=yaml.FullLoader)


class DataGrinderOneTask:

    def __init__(self, task_folder):
        # 参数初始化
        self.report_path = None

        print('=' * 25 + self.__class__.__name__ + '=' * 25)
        scenario_config_yaml = os.path.join(task_folder, 'TestConfig.yaml')
        with open(scenario_config_yaml, 'r', encoding='utf-8') as file:
            self.test_config = yaml.safe_load(file)
        self.task_folder = task_folder

        # 加载测试相关的参数
        self.product = self.test_config['product']
        self.version = self.test_config['version']
        self.test_topic = self.test_config['test_topic']
        self.truth_source = self.test_config['test_action']['ros2bag']['truth_source']
        self.test_date = str(self.test_config['test_date'])
        self.kpi_date_label = self.test_config['kpi_date_label']
        self.test_information = test_encyclopaedia['Information'][self.test_topic]
        self.scenario_update = self.test_config['test_action']['scenario_unit']['scenario_update']

        self.test_encyclopaedia = test_encyclopaedia[self.product]
        self.test_action = self.test_config['test_action']
        self.scenario_unit_folder = os.path.join(task_folder, '01_ScenarioUnit')
        self.tag_combination_folder = os.path.join(task_folder, '02_TagCombination')
        self.output_result_folder = os.path.join(task_folder, '03_OutputResult')
        topic_output_statistics_path = os.path.join(self.test_config['pred_folder'], 'topic_output_statistics.csv')
        self.scenario_statistics = pd.read_csv(topic_output_statistics_path, index_col=0)
        self.valid_scenario_list = list(self.scenario_statistics[self.scenario_statistics['isValid'] == 1].index)
        send_log(self, f'Valid Scenario为{self.valid_scenario_list}, 参与结果分析')

        self.test_result_yaml = os.path.join(self.task_folder, 'TestResult.yaml')
        if not os.path.exists(self.test_result_yaml):
            self.test_result = {
                'ScenarioUnit': {},
                'TagCombination': {},
                'OutputResult': {},
            }

            self.save_test_result()

    @sync_test_result
    def analyze_scenario_unit(self):
        # 依次创建scenario_unit测试信息
        scenario_run_list = {}
        scenario_list = []
        for scenario_tag in self.test_config['scenario_tag']:
            for scenario_id in scenario_tag['scenario_id']:
                if scenario_id in scenario_list or scenario_id not in self.valid_scenario_list:
                    continue

                scenario_list.append(scenario_id)
                scenario_test_config = {
                    'product': self.test_config['product'],
                    'version': self.test_config['version'],
                    'test_topic': self.test_config['test_topic'],
                    'test_date': str(self.test_config['test_date']),
                    'kpi_date_label': self.test_config['kpi_date_label'],
                    'pred_folder': os.path.join(self.test_config['pred_folder'], scenario_id),
                    'gt_folder': os.path.join(self.test_config['gt_folder'], scenario_id),
                    'truth_source': self.truth_source,
                    'test_action': self.test_action['scenario_unit'],
                    'test_item': self.test_config['test_item'],
                    'scenario_tag': scenario_tag['tag'],
                    'scenario_id': scenario_id,
                    'timestamp_matching_tolerance': self.test_config['timestamp_matching_tolerance'],
                }

                if self.test_topic == 'Obstacles':
                    scenario_test_config['target_characteristic'] = ['is_coverageValid', 'is_keyObj']
                    for item in [
                        'detected_ROI', 'region_division', 'coverage_reference_point',
                        'coverage_threshold', 'object_matching_tolerance',
                    ]:
                        scenario_test_config[item] = self.test_config[item]
                elif self.test_topic == 'Lines':
                    for item in [
                        'lane_matching_width',
                    ]:
                        scenario_test_config[item] = self.test_config[item]

                scenario_run_list[scenario_id] = {
                    'scenario_unit_folder': os.path.join(self.scenario_unit_folder, scenario_id),
                    'scenario_test_config': scenario_test_config,
                    'scenario_config_yaml': os.path.join(self.scenario_unit_folder, scenario_id, 'TestConfig.yaml')
                }

        for scenario_id, scenario_run_info in scenario_run_list.items():

            if os.path.exists(scenario_run_info['scenario_config_yaml']) and (not self.scenario_update):
                send_log(self, f'{scenario_id}已经完成计算,跳过')
                continue

            self.test_result['ScenarioUnit'][scenario_id] = (
                self.get_relpath(scenario_run_info['scenario_unit_folder']))
            create_folder(scenario_run_info['scenario_unit_folder'], False)
            with open(scenario_run_info['scenario_config_yaml'], 'w', encoding='utf-8') as f:
                yaml.dump(scenario_run_info['scenario_test_config'],
                          f, encoding='utf-8', allow_unicode=True, sort_keys=False)
            DataGrinderPilotObstaclesOneCase(scenario_run_info['scenario_unit_folder']).start()

    def get_relpath(self, path: str) -> str:
        return os.path.relpath(path, self.task_folder)

    def get_abspath(self, path: str) -> str:
        return os.path.join(self.task_folder, path)

    def save_test_result(self):
        with open(self.test_result_yaml, 'w', encoding='utf-8') as f:
            yaml.dump(self.test_result,
                      f, encoding='utf-8', allow_unicode=True, sort_keys=False)

    def load_test_result(self):
        with open(self.test_result_yaml) as f:
            self.test_result = yaml.load(f, Loader=yaml.FullLoader)


class DataGrinderPilotObstaclesOneCase(DataGrinderOneCase):

    @sync_test_result
    def evaluate_metrics(self):
        send_log(self, f'{self.test_topic}, 使用{self.test_topic}MetricEvaluator')

        for topic in self.test_result[self.test_topic].keys():

            if topic == 'GroundTruth':
                continue

            topic_tag = topic.replace('/', '')
            metric_folder = os.path.join(self.DataFolder, self.test_topic, topic_tag, 'metric')
            create_folder(metric_folder)
            if 'metric' not in self.test_result[self.test_topic][topic]:
                self.test_result[self.test_topic][topic]['metric'] = {}

            # 确认是否需要继续计算
            raw = self.test_result[self.test_topic][topic]['raw']
            data = pd.read_csv(self.get_abspath(raw['pred_data']), index_col=False)
            if not len(data):
                send_log(self, f'{self.scenario_id} {topic} 不计算metric数据')
                continue

            match_data_path = self.test_result[self.test_topic][topic]['match']['match_data']

            input_parameter_container = {
                'metric_type': self.test_config['test_item'][topic],
                'characteristic_type': self.test_config['target_characteristic'],
                'kpi_date_label': self.kpi_date_label,
            }
            parameter_json_path = os.path.join(get_project_path(), 'Temp', 'evaluate_api_parameter.json')
            with open(parameter_json_path, 'w', encoding='utf-8') as f:
                json.dump(input_parameter_container, f, ensure_ascii=False, indent=4)

            cmd = [
                f"{bench_config['master']['sys_interpreter']}",
                "Api_EvaluateMetrics.py",
                "-m", self.get_abspath(match_data_path),
                "-j", parameter_json_path,
                "-f", metric_folder,
            ]

            print(cmd)
            send_log(self, f'{self.test_topic} {topic} 指标评估')
            cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
            result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
            # os.remove(parameter_json_path)
            if result.stderr and 'warning' not in result.stderr:
                send_log(self, f'EvaluateMetrics 发生错误 {result.stderr}')

            for characteristic in os.listdir(metric_folder):
                self.test_result[self.test_topic][topic]['metric'][characteristic] = {}
                for metric in os.listdir(os.path.join(metric_folder, characteristic)):
                    self.test_result[self.test_topic][topic]['metric'][characteristic][metric.split('.')[0]] \
                        = self.get_relpath(os.path.join(metric_folder, characteristic, metric))

    @sync_test_result
    def load_scenario_info(self):

        def resize_image_by_height(image_path, height):
            img = Image.open(image_path)
            width = int(img.width * height / img.height)
            img = img.resize((width, height))

            return img

        # 图片拼接: 预览图+地图+自车车速
        overview_pic_path = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo', f'{self.scenario_id}_3000.png')
        overview_img = resize_image_by_height(overview_pic_path, 900)
        map_pic_path = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo', f'{self.scenario_id}_map.png')
        map_pic = resize_image_by_height(map_pic_path, 900)
        ego_vx_pic = self.get_abspath(self.test_result['General']['sync_ego_figure'])
        ego_vx = resize_image_by_height(ego_vx_pic, 900)

        # 创建一个新的空白图片用于拼接
        total_width = overview_img.width + map_pic.width + ego_vx.width
        concat_img = Image.new('RGB', (total_width, 900))

        # 拼接图片
        x_offset = 0
        for resized_img in [overview_img, map_pic, ego_vx]:
            concat_img.paste(resized_img, (x_offset, 0))
            x_offset += resized_img.width

        # 保存拼接后的图片
        path = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo', f'{self.scenario_id}_Info.jpg')
        concat_img.save(path)
        img = mpimg.imread(path)
        fig, ax = plt.subplots()
        ax.axis('off')
        ax.imshow(img)
        ax.set_title(self.scenario_id, fontsize=6, color='black', pad=2)  # pad参数调整标题与图片之间的距离
        plt.savefig(path, bbox_inches='tight', pad_inches=0, dpi=400)
        self.test_result['General']['scenario_info'] = self.get_relpath(path)

    @sync_test_result
    def sketch_bug(self):

        # 对于每个频繁的ID，找到时间戳位于中间的行的索引
        def get_middle_index_for_bug(bug_data, sort_value, count_threshold):
            id_counts = bug_data[sort_value].value_counts()
            frequent_ids = id_counts[id_counts >= count_threshold].index

            corresponding_indices = []
            for id_ in frequent_ids:
                id_df = bug_data[bug_data[sort_value] == id_]
                id_df_sorted = id_df.sort_values(by='gt.time_stamp')
                middle_index = len(id_df_sorted) // 2
                corresponding_indices.append(id_df.at[id_df_sorted.index[middle_index], 'corresponding_index'])

            return sorted(corresponding_indices)

        # 时间戳换算为帧数，进行视频截图
        video_info_path = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo', 'video_info.yaml')
        with open(video_info_path, 'r', encoding='utf-8') as file:
            video_info = yaml.safe_load(file)
        video_start_time, fps = video_info['start_time'] + self.cut_frame_offset, video_info['fps']
        # 将所有需要截图的时间辍都保存起来，统一交给ReplayClient视频截图
        video_snap_dict = {}

        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth':
                continue

            topic_tag = topic.replace('/', '')
            sketch_folder = os.path.join(self.BugFolder, self.test_topic, topic_tag, 'sketch')
            self.test_result[self.test_topic][topic]['bug'] = {}
            create_folder(sketch_folder)

            # 异常至少存在frame_threshold帧，才会被识别为bug用作分析
            frame_threshold = round(self.test_result[self.test_topic][topic]['match_frequency'])
            send_log(self, f'出现次数低于{frame_threshold}的bug会被忽视')

            # 使用total中的recall_precision数据可视化，但只抓取部分特镇的bug
            if 'total' not in self.test_result[self.test_topic][topic]['metric']:
                send_log(self, f'{self.scenario_id} {topic} 不存在metric数据')
                continue

            total_data_path = self.test_result[self.test_topic][topic]['metric']['total']['recall_precision']
            total_data = pd.read_csv(self.get_abspath(total_data_path), index_col=False).reset_index(drop=True)

            for characteristic in self.test_result[self.test_topic][topic]['metric']:
                if characteristic not in self.test_action['bug_characteristic']:
                    continue

                characteristic_folder = os.path.join(sketch_folder, characteristic)
                create_folder(characteristic_folder)
                self.test_result[self.test_topic][topic]['bug'][characteristic] = {}
                data_for_bug = self.test_result[self.test_topic][topic]['metric'][characteristic]
                bug_index_dict = self.get_bug_index_dict(data_for_bug)

                for bug_type, corresponding_index in bug_index_dict.items():
                    bug_type_folder = os.path.join(characteristic_folder, bug_type)
                    bug_data = total_data[total_data['corresponding_index'].isin(corresponding_index)]
                    # 根据id的出现次数排序
                    if bug_type == 'false_positive':
                        sorted_id = 'pred.id'
                    else:
                        sorted_id = 'gt.id'

                    bug_corresponding_indices = get_middle_index_for_bug(bug_data, sorted_id, frame_threshold)
                    for bug_corresponding_index in bug_corresponding_indices:
                        row = bug_data[bug_data['corresponding_index'] == bug_corresponding_index].iloc[0]
                        time_stamp = row['gt.time_stamp']
                        frame_data = total_data[total_data['gt.time_stamp'] == time_stamp]
                        # 如果这一帧内没有gt或者没有pred，暂时先不提bug
                        if frame_data['gt.flag'].sum() == 0 or frame_data['pred.flag'].sum() == 0:
                            continue

                        one_bug_folder = os.path.join(bug_type_folder, f'{time_stamp}')
                        create_folder(one_bug_folder)
                        if bug_type not in self.test_config['test_item']:
                            self.test_result[self.test_topic][topic]['bug'][characteristic][bug_type] = (
                                self.get_relpath(bug_type_folder))

                        plot_path = os.path.join(one_bug_folder, 'bug_sketch.jpg')
                        frame_bug_info_bev = {
                            'gt': {bug_type: [row['gt.id']]}, 'pred': {bug_type: [row['pred.id']]},
                        }

                        send_log(self, f'保存 {topic} {bug_type} {time_stamp}的图片')
                        self.plot_one_frame_for_obstacles(topic, frame_data, plot_path, frame_bug_info_bev)

                        # 选择截图的相机
                        if row['gt.flag'] == 1:
                            cameras = self.which_camera_saw_you(row['gt.x'], row['gt.y'])
                        else:
                            cameras = self.which_camera_saw_you(row['pred.x'], row['pred.y'])

                        frame_index = round((time_stamp - video_start_time) * fps)

                        for camera in cameras:
                            if camera not in video_snap_dict:
                                video_snap_dict[camera] = []
                            if frame_index not in video_snap_dict[camera]:
                                video_snap_dict[camera].append(frame_index)

                        send_log(self, f'保存 {topic} {bug_type} {time_stamp}的异常信息')
                        bug_info = row.to_dict()
                        bug_info['rosbag_time'] = bug_info['pred.time_stamp'] - self.test_result['General']['time_gap']
                        bug_info['frame_index'] = frame_index
                        bug_info['camera'] = cameras
                        bug_info['bug_type'] = bug_type
                        with open(os.path.join(one_bug_folder, 'bug_info.json'), 'w', encoding='utf-8') as f:
                            json.dump(bug_info, f, ensure_ascii=False, indent=4)

        # 启用ssh_client, 按照相机批量截图并复制到本机
        replay_client = SSHClient()
        video_snap_folder = os.path.join(self.BugFolder, 'General')
        create_folder(video_snap_folder)

        for camera, frame_index_list in video_snap_dict.items():
            camera_folder = os.path.join(video_snap_folder, camera)
            create_folder(camera_folder)

            replay_client.cut_frames(scenario_id=self.scenario_id,
                                     frame_index_list=sorted(frame_index_list),
                                     camera=camera,
                                     local_folder=camera_folder)

        # 将bug需要ps的内容保存为一个json，批量处理
        bug_label_info_list = []
        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth':
                continue

            for characteristic in self.test_result[self.test_topic][topic]['bug']:
                for bug_type, bug_type_folder in self.test_result[self.test_topic][topic]['bug'][
                    characteristic].items():

                    for time_stamp in os.listdir(self.get_abspath(bug_type_folder)):

                        send_log(self, f'汇总 {topic} {characteristic} {bug_type} {time_stamp} camera截图数据')
                        one_bug_folder = os.path.join(self.get_abspath(bug_type_folder), time_stamp)
                        bug_info_json = os.path.join(one_bug_folder, 'bug_info.json')
                        with open(bug_info_json, 'r', encoding='utf-8') as f:
                            bug_info = json.load(f)

                        frame_index = round((float(time_stamp) - video_start_time) * fps)
                        bug_label_info = {
                            'scenario_id': self.scenario_id,
                            'frame_index': frame_index,
                            'time_stamp': float(time_stamp),
                            'camera_label_info': {}
                        }

                        for camera in bug_info['camera']:
                            bug_label_info['camera_label_info'][camera] = {
                                'origin_shot': os.path.join(self.BugFolder, 'General', camera,
                                                            f'{frame_index}.jpg'),
                                'process_shot': os.path.join(one_bug_folder,
                                                             f'{camera}-{self.scenario_id}-{frame_index}.jpg'),
                                'label_info': []
                            }
                            one_label_info = {
                                'bug_type': bug_type,
                            }

                            # 箭头
                            if bug_info['gt.flag'] == 1:
                                one_label_info['gt_arrow'] = [bug_info['gt.x'], bug_info['gt.y'],
                                                              bug_info['gt.height']]

                                # 8个角点也放进去
                                one_label_info['gt_corner'] = {
                                    'bottom': [
                                        [bug_info['gt.pt_0_x'], bug_info['gt.pt_0_y'], 0],
                                        [bug_info['gt.pt_1_x'], bug_info['gt.pt_1_y'], 0],
                                        [bug_info['gt.pt_2_x'], bug_info['gt.pt_2_y'], 0],
                                        [bug_info['gt.pt_3_x'], bug_info['gt.pt_3_y'], 0],
                                    ],
                                    'top': [
                                        [bug_info['gt.pt_0_x'], bug_info['gt.pt_0_y'], bug_info['gt.height']],
                                        [bug_info['gt.pt_1_x'], bug_info['gt.pt_1_y'], bug_info['gt.height']],
                                        [bug_info['gt.pt_2_x'], bug_info['gt.pt_2_y'], bug_info['gt.height']],
                                        [bug_info['gt.pt_3_x'], bug_info['gt.pt_3_y'], bug_info['gt.height']],
                                    ]
                                }

                            if bug_info['pred.flag'] == 1:
                                one_label_info['pred_arrow'] = [bug_info['pred.x'], bug_info['pred.y'],
                                                                bug_info['pred.height']]

                                # 8个角点也放进去
                                one_label_info['pred_corner'] = {
                                    'bottom': [
                                        [bug_info['pred.pt_0_x'], bug_info['pred.pt_0_y'], 0],
                                        [bug_info['pred.pt_1_x'], bug_info['pred.pt_1_y'], 0],
                                        [bug_info['pred.pt_2_x'], bug_info['pred.pt_2_y'], 0],
                                        [bug_info['pred.pt_3_x'], bug_info['pred.pt_3_y'], 0],
                                    ],
                                    'top': [
                                        [bug_info['pred.pt_0_x'], bug_info['pred.pt_0_y'],
                                         bug_info['pred.height']],
                                        [bug_info['pred.pt_1_x'], bug_info['pred.pt_1_y'],
                                         bug_info['pred.height']],
                                        [bug_info['pred.pt_2_x'], bug_info['pred.pt_2_y'],
                                         bug_info['pred.height']],
                                        [bug_info['pred.pt_3_x'], bug_info['pred.pt_3_y'],
                                         bug_info['pred.height']],
                                    ]
                                }

                            if bug_info['gt.flag'] == 1:
                                one_label_info['center'] = [
                                    bug_info['gt.x'], bug_info['gt.y'], bug_info['gt.height']
                                ]
                            else:
                                one_label_info['center'] = [
                                    bug_info['pred.x'], bug_info['pred.y'], bug_info['pred.height']
                                ]

                            bug_label_info['camera_label_info'][camera]['label_info'].append(one_label_info)

                        bug_label_info_list.append(bug_label_info)

        bug_label_info_json = os.path.join(self.BugFolder, 'General', 'bug_label_info.json')
        with open(bug_label_info_json, 'w') as json_file:
            json.dump(bug_label_info_list, json_file, ensure_ascii=False, indent=4)

        calibration = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo', 'yaml_calib')

        # 调用给视频截图增加箭头的端口
        cmd = [
            f"{bench_config['master']['sys_interpreter']}",
            "Api_ProcessVideoShot.py",
            "-c",
            calibration,
            "-a",
            bug_label_info_json
        ]

        print(cmd)
        cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
        result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
        if result.stderr and 'warning' not in result.stderr:
            send_log(self, f'ProcessVideoShot 发生错误 {result.stderr}')

    @sync_test_result
    def bug_report(self):

        # 建立bug report，并汇总
        bug_jira_column = [
            'sw_version', 'test_object', 'scenario_type', 'topic', 'bug_type', 'target_type', 'gt_target_id',
            'scenario_id', 'summary', 'description', 'assignee', 'attachment_path', 'contained_by',
            'project_id', 'uuid_content', 'uuid', 'is_valid'
        ]

        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth':
                continue

            bug_jira_rows = []
            topic_tag = topic.replace('/', '')

            for characteristic in self.test_result[self.test_topic][topic]['bug']:
                bug_report_folder = os.path.join(self.BugFolder, self.test_topic, topic_tag, 'bug_report',
                                                 characteristic)
                create_folder(bug_report_folder)

                for bug_type, bug_type_folder in self.test_result[self.test_topic][topic]['bug'][
                    characteristic].items():

                    # 模型输出没有速度
                    if bug_type in ['vx_error', 'vy_error'] and topic != '/VA/Obstacles':
                        continue

                    for time_stamp in os.listdir(self.get_abspath(bug_type_folder)):
                        one_bug_folder = os.path.join(self.get_abspath(bug_type_folder), time_stamp)

                        bug_info_json = os.path.join(one_bug_folder, 'bug_info.json')
                        with open(bug_info_json, 'r', encoding='utf-8') as f:
                            bug_info = json.load(f)

                        # 获取时间
                        gt_timestamp = bug_info['gt.time_stamp']
                        rosbag_time = bug_info['rosbag_time']

                        # 获得目标类型
                        target_type = bug_info['gt.type_classification'] if bug_info['gt.flag'] \
                            else bug_info['pred.type_classification']

                        # 获得目标类型
                        target_x = bug_info['gt.x'] if bug_info['gt.flag'] else bug_info['pred.x']

                        # 获得目标id
                        target_id = bug_info['gt.id'] if bug_info['gt.flag'] else bug_info['pred.id']

                        # 获得目标特征
                        target_characteristic = []
                        if bug_info['gt.flag']:
                            if bug_info['gt.is_coverageValid']:
                                target_characteristic.append(
                                    f'是全局目标(遮挡率≤{self.test_config["coverage_threshold"]:.0%})')
                            else:
                                target_characteristic.append(
                                    f'非全局目标(遮挡率≤{self.test_config["coverage_threshold"]:.0%})')
                            if bug_info['gt.is_keyObj']:
                                target_characteristic.append('是关键目标')
                            else:
                                target_characteristic.append('非关键目标')
                        else:
                            if bug_info['pred.is_coverageValid']:
                                target_characteristic.append(
                                    f'是全局目标(遮挡率≤{self.test_config["coverage_threshold"]:.0%})')
                            else:
                                target_characteristic.append(
                                    f'非全局目标(遮挡率≤{self.test_config["coverage_threshold"]:.0%})')
                            if bug_info['pred.is_keyObj']:
                                target_characteristic.append('是关键目标')
                            else:
                                target_characteristic.append('非关键目标')
                        target_characteristic = '&'.join(target_characteristic)

                        # 开始生成报告
                        uuid_content = f'{self.scenario_id}-{bug_type}-{int(target_id)}|{gt_timestamp}_{target_x}'
                        uuid = generate_unique_id(f'{self.version}-{uuid_content}')
                        report_title = f'{self.product}测试异常报告({uuid})'
                        send_log(self, f'开始生成 {report_title}')
                        send_log(self, f'{one_bug_folder} 开始生成 {report_title}')

                        title_background = os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure',
                                                        'TitlePage.png')
                        logo = os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'ZoneLogo.png')
                        report_generator = PDFReportTemplate(report_title=report_title,
                                                             test_time=f'{topic} {bug_type}',
                                                             tester=self.scenario_id,
                                                             version=self.version,
                                                             title_page=title_background,
                                                             title_summary_img=None,
                                                             logo=logo)

                        img_list = glob.glob(os.path.join(one_bug_folder, 'CAM_*.jpg'))
                        if len(img_list):
                            img_list = [img_list]
                        else:
                            img_list = None

                        report_generator.addOnePage(
                            heading=f'{bug_type} 视频截图',
                            text_list=[
                                f'场景名: {self.scenario_id}',
                                f'topic: {topic}, 目标类型: {target_type}, 目标特征: {target_characteristic}',
                                f'发生时刻: {round(float(time_stamp), 3)} sec / {bug_info["frame_index"]} frame / rosbag_time: {round(float(rosbag_time), 3)} sec',
                                '红色为GroundTruth，蓝色为Prediction',
                            ],
                            img_list=img_list,
                        )

                        text_list = ['id, type, x, y, vx, vy, yaw, length, width, height 信息如下:']
                        if bug_info['gt.flag']:
                            text_list.append(
                                f"GroundTruth: id-{int(bug_info['gt.id'])}, {bug_info['gt.type_classification']}, "
                                f"({round(bug_info['gt.x'], 2)}m, {round(bug_info['gt.y'], 2)}m), "
                                f"({round(bug_info['gt.vx'], 2)}m/s, {round(bug_info['gt.vy'], 2)}m/s), "
                                f"{round(bug_info['gt.yaw'] * 57.3, 1)}°, "
                                f"{round(bug_info['gt.length'], 2)}m × {round(bug_info['gt.width'], 2)}m × {round(bug_info['gt.height'], 2)}m"
                            )
                        if bug_info['pred.flag']:
                            text_list.append(
                                f"Prediction: id-{int(bug_info['pred.id'])}, {bug_info['pred.type_classification']}, "
                                f"({round(bug_info['pred.x'], 2)}m, {round(bug_info['pred.y'], 2)}m), "
                                f"({round(bug_info['pred.vx'], 2)}m/s, {round(bug_info['pred.vy'], 2)}m/s), "
                                f"{round(bug_info['pred.yaw'] * 57.3, 1)}°, "
                                f"{round(bug_info['pred.length'], 2)}m × {round(bug_info['pred.width'], 2)}m × {round(bug_info['pred.height'], 2)}m"
                            )

                        report_generator.addOnePage(
                            heading=f'{bug_type} 真值与感知的对比',
                            text_list=text_list,
                            img_list=[[os.path.join(one_bug_folder, 'bug_sketch.jpg')]],
                        )

                        bug_report_path = report_generator.genReport(
                            compress=1,
                            report_path=os.path.join(bug_report_folder,
                                                     f'{self.product}-{self.scenario_id}-{target_type}-{topic_tag}-{bug_type}({uuid}).pdf')
                        )

                        info = test_encyclopaedia['Information'][self.test_topic]
                        jira_summary = f'数据回灌感知测试({self.product}-{self.version}-{info["name"]})'
                        text = [
                            '测试版本: {:s}, {:s}'.format(self.product, self.version),
                            '测试时间: {:s}'.format(self.test_config['test_date']),
                            '异常类型: {:s}'.format(info['bug_items'][bug_type]['name']),
                            '异常发生时刻的场景和截图见附件',
                            '-------------',
                        ]
                        jira_description = '\n'.join(text)
                        project_key = self.test_encyclopaedia['project_id'][0]
                        project_id = self.test_encyclopaedia['project_id'][1]

                        if bug_info['gt.flag']:
                            bug_gt_id = bug_info['gt.id']
                        else:
                            bug_gt_id = 0

                        jira_row = [
                            self.version, info['name'], '&'.join(self.scenario_tag.values()),
                            topic, info['bug_items'][bug_type]['name'], target_type, bug_gt_id,
                            self.scenario_id, jira_summary, jira_description,
                            '', bug_report_path, f'{project_key}-0', project_id,
                            uuid_content, f'uuid-{uuid}', 0
                        ]

                        bug_jira_rows.append(jira_row)

                bug_jira_summary_path = os.path.join(bug_report_folder, 'bug_jira_summary.csv')
                bug_jira_summary = pd.DataFrame(bug_jira_rows, columns=bug_jira_column)
                bug_jira_summary.to_csv(bug_jira_summary_path, index=False)
                send_log(self, '{:s} 保存完毕'.format(bug_jira_summary_path))
                self.test_result[self.test_topic][topic]['bug_jira_summary'] = self.get_relpath(bug_jira_summary_path)

        # 删除图片，减少磁盘占用
        general_folder = os.path.join(self.BugFolder, 'General')
        if os.path.exists(general_folder):
            shutil.rmtree(general_folder)

    @sync_test_result
    def sketch_render(self):

        # 时间戳换算为帧数，进行视频截图
        video_info_path = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo', 'video_info.yaml')
        with open(video_info_path, 'r', encoding='utf-8') as file:
            video_info = yaml.safe_load(file)
        video_start_time, fps = video_info['start_time'] + self.cut_frame_offset, video_info['fps']
        # 将所有需要截图的时间辍都保存起来，统一交给ReplayClient视频截图
        video_snap_dict = {}

        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth' or topic not in self.test_action['render_topic']:
                continue

            topic_tag = topic.replace('/', '')
            sketch_folder = os.path.join(self.RenderFolder, self.test_topic, topic_tag, 'sketch')
            self.test_result[self.test_topic][topic]['render'] = {}
            create_folder(sketch_folder)

            # 使用total中的recall_precision数据可视化，但只抓取keyObj的bug
            total_data_path = self.test_result[self.test_topic][topic]['metric']['total']['recall_precision']
            total_data = pd.read_csv(self.get_abspath(total_data_path), index_col=False).reset_index(drop=True)

            for characteristic in self.test_result[self.test_topic][topic]['metric']:
                if characteristic not in self.test_action['bug_characteristic']:
                    continue

                characteristic_folder = os.path.join(sketch_folder, characteristic)
                create_folder(characteristic_folder)
                self.test_result[self.test_topic][topic]['render'][characteristic] = {
                    'sketch': self.get_relpath(characteristic_folder)
                }
                data_for_render = self.test_result[self.test_topic][topic]['metric'][characteristic]
                bug_index_dict = self.get_bug_index_dict(data_for_render)

                for time_stamp in (total_data.sort_values(by='gt.time_stamp')
                        .drop_duplicates(subset=['gt.time_stamp'], keep='first')['gt.time_stamp'].values):

                    frame_index = round((time_stamp - video_start_time) * fps)
                    frame_data = total_data[total_data['gt.time_stamp'] == time_stamp]
                    frame_corresponding_index_list = frame_data['corresponding_index'].to_list()

                    one_render_folder = os.path.join(characteristic_folder, f'{time_stamp}')
                    create_folder(one_render_folder)
                    plot_path = os.path.join(one_render_folder, 'render_sketch.jpg')

                    frame_bug_info_bev = {'gt': {}, 'pred': {}}
                    frame_bug_info_camera = []
                    for bug_type, corresponding_index_list in bug_index_dict.items():
                        for corresponding_index in corresponding_index_list:
                            if corresponding_index not in frame_corresponding_index_list:
                                continue

                            row = frame_data[frame_data['corresponding_index'] == corresponding_index].iloc[0]
                            if row['gt.flag'] == 1:
                                if bug_type not in frame_bug_info_bev['gt']:
                                    frame_bug_info_bev['gt'][bug_type] = []
                                frame_bug_info_bev['gt'][bug_type].append(row['gt.id'])

                            if row['pred.flag'] == 1:
                                if bug_type not in frame_bug_info_bev['pred']:
                                    frame_bug_info_bev['pred'][bug_type] = []
                                frame_bug_info_bev['pred'][bug_type].append(row['pred.id'])

                            # 选择截图的相机
                            if row['gt.flag'] == 1:
                                cameras = self.which_camera_saw_you(row['gt.x'], row['gt.y'])
                            else:
                                cameras = self.which_camera_saw_you(row['pred.x'], row['pred.y'])

                            bug_info = row.to_dict()
                            bug_info['frame_index'] = frame_index
                            bug_info['camera'] = cameras
                            bug_info['bug_type'] = bug_type
                            frame_bug_info_camera.append(bug_info)

                    send_log(self, f'保存 {self.scenario_id} {topic} {time_stamp}的全部信息')
                    with open(os.path.join(one_render_folder, 'bug_info.json'), 'w', encoding='utf-8') as f:
                        json.dump(frame_bug_info_camera, f, ensure_ascii=False, indent=4)

                    send_log(self, f'保存 {self.scenario_id} {topic} {time_stamp} render图片')
                    self.plot_one_frame_for_obstacles(topic, frame_data, plot_path, frame_bug_info_bev)

                    for camera in self.camera_list:
                        if camera not in video_snap_dict:
                            video_snap_dict[camera] = []
                        if frame_index not in video_snap_dict[camera]:
                            video_snap_dict[camera].append(frame_index)

        # 启用ssh_client, 按照相机批量截图并复制到本机
        replay_client = SSHClient()
        video_snap_folder = os.path.join(self.RenderFolder, 'General')
        create_folder(video_snap_folder)

        for camera, frame_index_list in video_snap_dict.items():
            camera_folder = os.path.join(video_snap_folder, camera)
            create_folder(camera_folder)

            replay_client.cut_frames(scenario_id=self.scenario_id,
                                     frame_index_list=sorted(frame_index_list),
                                     camera=camera,
                                     local_folder=camera_folder)

        # 将bug需要ps的内容保存为一个json，批量处理
        bug_label_info_list = []
        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth' or topic not in self.test_action['render_topic']:
                continue

            for characteristic in self.test_result[self.test_topic][topic]['render']:
                characteristic_folder = self.get_abspath(self.test_result[self.test_topic][topic]['render'][
                                                             characteristic]['sketch'])

                for time_stamp in os.listdir(characteristic_folder):
                    one_render_folder = os.path.join(characteristic_folder, time_stamp)
                    bug_info_json = os.path.join(one_render_folder, 'bug_info.json')
                    with open(bug_info_json, 'r', encoding='utf-8') as f:
                        bug_info_list = json.load(f)

                    frame_index = round((float(time_stamp) - video_start_time) * fps)
                    for camera in self.camera_list:
                        origin_shot = os.path.join(self.RenderFolder, 'General', camera, f'{frame_index}.jpg')
                        process_shot = os.path.join(one_render_folder,
                                                    f'{camera}-{self.scenario_id}-{frame_index}.jpg')
                        shutil.copy(origin_shot, process_shot)

                    if not len(bug_info_list):
                        continue

                    bug_label_info = {
                        'scenario_id': self.scenario_id,
                        'frame_index': frame_index,
                        'time_stamp': float(time_stamp),
                        'camera_label_info': {}
                    }

                    for bug_info in bug_info_list:
                        for camera in bug_info['camera']:
                            if camera not in bug_label_info['camera_label_info']:
                                bug_label_info['camera_label_info'][camera] = {
                                    'origin_shot': os.path.join(one_render_folder,
                                                                f'{camera}-{self.scenario_id}-{frame_index}.jpg'),
                                    'process_shot': os.path.join(one_render_folder,
                                                                 f'{camera}-{self.scenario_id}-{frame_index}.jpg'),
                                    'label_info': []
                                }

                            one_label_info = {
                                'bug_type': bug_info['bug_type'],
                            }

                            # 箭头
                            if bug_info['gt.flag'] == 1:
                                one_label_info['gt_arrow'] = [bug_info['gt.x'], bug_info['gt.y'],
                                                              bug_info['gt.height']]

                                # 8个角点也放进去
                                one_label_info['gt_corner'] = {
                                    'bottom': [
                                        [bug_info['gt.pt_0_x'], bug_info['gt.pt_0_y'], 0],
                                        [bug_info['gt.pt_1_x'], bug_info['gt.pt_1_y'], 0],
                                        [bug_info['gt.pt_2_x'], bug_info['gt.pt_2_y'], 0],
                                        [bug_info['gt.pt_3_x'], bug_info['gt.pt_3_y'], 0],
                                    ],
                                    'top': [
                                        [bug_info['gt.pt_0_x'], bug_info['gt.pt_0_y'], bug_info['gt.height']],
                                        [bug_info['gt.pt_1_x'], bug_info['gt.pt_1_y'], bug_info['gt.height']],
                                        [bug_info['gt.pt_2_x'], bug_info['gt.pt_2_y'], bug_info['gt.height']],
                                        [bug_info['gt.pt_3_x'], bug_info['gt.pt_3_y'], bug_info['gt.height']],
                                    ]
                                }

                            if bug_info['pred.flag'] == 1:
                                one_label_info['pred_arrow'] = [bug_info['pred.x'], bug_info['pred.y'],
                                                                bug_info['pred.height']]

                                # 8个角点也放进去
                                one_label_info['pred_corner'] = {
                                    'bottom': [
                                        [bug_info['pred.pt_0_x'], bug_info['pred.pt_0_y'], 0],
                                        [bug_info['pred.pt_1_x'], bug_info['pred.pt_1_y'], 0],
                                        [bug_info['pred.pt_2_x'], bug_info['pred.pt_2_y'], 0],
                                        [bug_info['pred.pt_3_x'], bug_info['pred.pt_3_y'], 0],
                                    ],
                                    'top': [
                                        [bug_info['pred.pt_0_x'], bug_info['pred.pt_0_y'],
                                         bug_info['pred.height']],
                                        [bug_info['pred.pt_1_x'], bug_info['pred.pt_1_y'],
                                         bug_info['pred.height']],
                                        [bug_info['pred.pt_2_x'], bug_info['pred.pt_2_y'],
                                         bug_info['pred.height']],
                                        [bug_info['pred.pt_3_x'], bug_info['pred.pt_3_y'],
                                         bug_info['pred.height']],
                                    ]
                                }

                            if bug_info['gt.flag'] == 1:
                                one_label_info['center'] = [
                                    bug_info['gt.x'], bug_info['gt.y'], bug_info['gt.height']
                                ]
                            else:
                                one_label_info['center'] = [
                                    bug_info['pred.x'], bug_info['pred.y'], bug_info['pred.height']
                                ]

                            bug_label_info['camera_label_info'][camera]['label_info'].append(one_label_info)

                    send_log(self, f'汇总 {self.scenario_id} {topic} {characteristic} {time_stamp} camera 截图数据')
                    bug_label_info_list.append(bug_label_info)

        bug_label_info_json = os.path.join(self.RenderFolder, 'General', 'bug_label_info.json')
        with open(bug_label_info_json, 'w') as json_file:
            json.dump(bug_label_info_list, json_file, ensure_ascii=False, indent=4)

        calibration = os.path.join(self.scenario_unit_folder, '00_ScenarioInfo', 'yaml_calib')

        # 调用给视频截图增加箭头的端口
        cmd = [
            f"{bench_config['master']['sys_interpreter']}",
            "Api_ProcessVideoShot.py",
            "-c",
            calibration,
            "-a",
            bug_label_info_json
        ]

        print(cmd)
        cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
        result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
        if result.stderr and 'warning' not in result.stderr:
            send_log(self, f'ProcessVideoShot 发生错误 {result.stderr}')

    @sync_test_result
    def generate_image_and_video(self):

        def combine_image(center_pic_path, camera_path_dict, combined_pic_path):
            center_image = Image.open(center_pic_path)
            center_width, center_height = center_image.size

            camera_image = {}
            for camera, path in camera_path_dict.items():
                if path is not None:
                    image = Image.open(path)
                    width, height = image.size
                    new_height = round(center_width / 3 * height / width)
                    resized_img = image.resize((round(center_width / 3), new_height))
                    camera_image[camera] = {'width': round(center_width / 3), 'height': new_height,
                                            'image': resized_img}
                else:
                    camera_image[camera] = {'width': round(center_width / 3), 'height': 0, 'image': None}

            upper_height = 0
            for i in range(3):
                if list(camera_image.values())[i]['height'] > upper_height:
                    upper_height = list(camera_image.values())[i]['height']

            lower_height = 0
            for i in range(3, 6):
                if list(camera_image.values())[i]['height'] > lower_height:
                    lower_height = list(camera_image.values())[i]['height']

            total_height = upper_height + lower_height + center_height
            total_width = center_width

            canvas = Image.new('RGB', (total_width, total_height), 'white')
            canvas.paste(center_image, (0, upper_height))

            for i in range(3):
                camera = list(camera_image.keys())[i]
                if camera_image[camera]['image'] is not None:
                    width = camera_image[camera]['width']
                    image = camera_image[camera]['image'].resize((width, upper_height))
                    canvas.paste(image, (width * i, 0))

            for i in range(3, 6):
                camera = list(camera_image.keys())[i]
                if camera_image[camera]['image'] is not None:
                    width = camera_image[camera]['width']
                    image = camera_image[camera]['image'].resize((width, upper_height))
                    canvas.paste(image, (width * (i - 3), upper_height + center_height))

            canvas.save(combined_pic_path)

            return total_width, total_height

        for topic in self.test_result[self.test_topic].keys():
            if topic == 'GroundTruth' or topic not in self.test_action['render_topic']:
                continue

            if 'render' not in self.test_result[self.test_topic][topic]:
                continue

            topic_tag = topic.replace('/', '')
            for characteristic in self.test_result[self.test_topic][topic]['render']:
                characteristic_folder = self.get_abspath(self.test_result[self.test_topic][topic]['render'][
                                                             characteristic]['sketch'])
                image_folder = os.path.join(self.RenderFolder, self.test_topic, topic_tag, 'image', characteristic)
                self.test_result[self.test_topic][topic]['render'][characteristic]['image'] = (
                    self.get_relpath(image_folder))
                create_folder(image_folder)

                width, height = 2500, 2132
                time_stamp_list = sorted(os.listdir(characteristic_folder), key=float)
                for t_idx, time_stamp in enumerate(time_stamp_list):
                    combined_pic_path = os.path.join(image_folder, 'img{:05d}.jpg'.format(t_idx + 1))
                    one_sketch_folder = os.path.join(characteristic_folder, time_stamp)
                    camera_path_dict = {}
                    for camera in self.camera_list:
                        c = glob.glob(os.path.join(one_sketch_folder, f'{camera}-*.jpg'))
                        if len(c):
                            camera_path_dict[camera] = c[0]
                        else:
                            camera_path_dict[camera] = None

                    send_log(self, f'保存 {self.scenario_id} {topic} {characteristic} bev和camera合并图片 {os.path.basename(combined_pic_path)}')
                    width, height = combine_image(
                        center_pic_path=os.path.join(one_sketch_folder, 'render_sketch.jpg'),
                        camera_path_dict=camera_path_dict,
                        combined_pic_path=combined_pic_path,
                    )

                video_folder = os.path.join(self.RenderFolder, self.test_topic, topic_tag, 'video', characteristic)
                create_folder(video_folder)
                combined_video = os.path.join(video_folder, f'{self.scenario_id}-{topic_tag}-{characteristic}.mp4')
                self.test_result[self.test_topic][topic]['render'][characteristic]['video'] = self.get_relpath(combined_video)
                fps = self.test_result[self.test_topic][topic]['match_frequency']
                image2video(image_folder, fps, combined_video, 1400, round(1400 * height / width))
                shutil.rmtree(image_folder)

        # 删除图片，减少磁盘占用
        general_folder = os.path.join(self.RenderFolder, 'General')
        if os.path.exists(general_folder):
            shutil.rmtree(general_folder)
        for folder in glob.glob(os.path.join(self.RenderFolder, self.test_topic, '*', 'sketch')):
            shutil.rmtree(folder)

    def get_bug_index_dict(self, metric_data_group):

        def filter_valid_bug_data(data):

            def check_regions(x, y, ru):
                def check_region(pt, region):
                    is_valid_list = []
                    for range_type in ['x', 'y']:
                        if isinstance(region[range_type][0], float) or isinstance(region[range_type][0], int):
                            is_valid = region[range_type][0] <= pt[range_type] < region[range_type][1]
                            is_valid_list.append(is_valid)

                        elif isinstance(region[range_type][0], list) or isinstance(region[range_type][0], tuple):
                            is_valid = any(
                                [sub_range_value[0] <= pt[range_type] < sub_range_value[1] for sub_range_value in
                                 region[range_type]])
                            is_valid_list.append(is_valid)

                    return all(is_valid_list)

                pt = {'x': x, 'y': y}
                if ru == 'DRU':
                    regions = self.test_config['region_division']['DRU'] + self.test_config['region_division']['VRU']
                else:
                    regions = self.test_config['region_division']['VRU']

                for region in regions:
                    if check_region(pt, region):
                        return 1
                return 0

            if 'gt.flag' in data.columns:
                data = data[~((data['gt.flag'] == 1) & (data['gt.id'].isin(bug_excluding_ids)))]
                data['is_bugArea'] = data.apply(
                    lambda row: check_regions(row['gt.x'], row['gt.y'], row['gt.road_user'])
                    if row['gt.flag'] == 1 else check_regions(row['pred.x'], row['pred.y'], row['pred.road_user']),
                    axis=1)
            else:
                data = data[~(data['gt.id'].isin(bug_excluding_ids))]
                data['is_bugArea'] = data.apply(
                    lambda row: check_regions(row['gt.x'], row['gt.y'], row['gt.road_user']),
                    axis=1)

            return data[data['is_bugArea'] == 1]

        bug_excluding_ids = []
        if os.path.exists(scenario_test_record_path):
            scenario_test_record = pd.read_csv(scenario_test_record_path, index_col=False)
            bug_excluding_row = scenario_test_record[(scenario_test_record['scenario_id'] == self.scenario_id)
                                                 & (scenario_test_record['truth_source'] == self.truth_source)]
            if len(bug_excluding_row):
                index = bug_excluding_row.index[0]
                bug_excluding_ids = json.loads(scenario_test_record.at[index, 'bug_excluding'])
        send_log(self, f'{self.scenario_id} 排除以下gt.id的bug {bug_excluding_ids}')

        bug_count = self.test_action['bug_count']
        bug_index_dict = {}
        for metric, data_path in metric_data_group.items():
            metric_data = pd.read_csv(self.get_abspath(data_path), index_col=False)
            metric_data = filter_valid_bug_data(metric_data)

            # 提bug的内容:
            # 这里不能知道场景，所以无法筛选场景相关的内容
            # 1.所有bug仅关注-50米到100米的范围，行人和两轮车只关注-50米到50米
            # 2.行人不测长，宽，高，航向角，车速
            # 3.两轮车不测长，宽，高；低速时不测航向角，车速
            # 4.长宽高误差必须针对类型判断正确的样本
            if metric == 'recall_precision':
                # 进一步筛选出现每种类型帧数最大的3个目标
                bug_index_dict['false_positive'] = []
                FP_data = metric_data[(metric_data['gt.flag'] == 0)
                                      & (metric_data['pred.flag'] == 1)
                                      & (metric_data['pred.x'] <= 100)
                                      & (metric_data['pred.x'] >= -50)
                                      & (metric_data['pred.y'] <= 8)
                                      & (metric_data['pred.y'] >= -8)]
                FP_data = FP_data[~((FP_data['pred.type_classification'].isin(['pedestrian', 'cyclist']))
                                    & (FP_data['pred.x'] > 50))]
                for key, group in FP_data.groupby('pred.type_classification'):
                    top_values = group['pred.id'].value_counts().head(bug_count).index
                    top_group = group[group['pred.id'].isin(top_values)]
                    bug_index_dict['false_positive'].extend(top_group['corresponding_index'].to_list())

                # 对漏检需要观察100-150米之内的bug
                bug_index_dict['false_negative'] = []
                FN_data = metric_data[(metric_data['gt.flag'] == 1)
                                      & (metric_data['pred.flag'] == 0)
                                      & (metric_data['gt.y'] <= 8)
                                      & (metric_data['gt.y'] >= -8)]
                FN_data = FN_data[~((FN_data['gt.type_classification'].isin(['pedestrian', 'cyclist']))
                                    & ((FN_data['gt.x'] > 50) | (FN_data['gt.x'] < -50)))]
                FN_data_near = FN_data[(metric_data['gt.x'] <= 100)
                                      & (metric_data['gt.x'] >= -100)]
                for key, group in FN_data_near.groupby('gt.type_classification'):
                    top_values = group['gt.id'].value_counts().head(bug_count * 2).index
                    top_group = group[group['gt.id'].isin(top_values)]
                    bug_index_dict['false_negative'].extend(top_group['corresponding_index'].to_list())
                FN_data_far = FN_data[(metric_data['gt.x'] <= 150)
                                      & (metric_data['gt.x'] > 100)]
                for key, group in FN_data_far.groupby('gt.type_classification'):
                    top_values = group['gt.id'].value_counts().head(bug_count * 2).index
                    top_group = group[group['gt.id'].isin(top_values)]
                    bug_index_dict['false_negative'].extend(top_group['corresponding_index'].to_list())

                bug_index_dict['false_type'] = []
                NCTP_data = metric_data[(metric_data['CTP'] == 0)
                                        & (metric_data['gt.flag'] == 1)
                                        & (metric_data['pred.flag'] == 1)
                                        & (metric_data['gt.x'] <= 100)
                                        & (metric_data['gt.x'] >= -50)
                                        & (metric_data['gt.y'] <= 8)
                                        & (metric_data['gt.y'] >= -8)]
                NCTP_data = NCTP_data[~((NCTP_data['gt.type_classification'].isin(['pedestrian', 'cyclist']))
                                        & (NCTP_data['gt.x'] > 50))]
                for key, group in NCTP_data.groupby('gt.type_classification'):
                    top_values = group['gt.id'].value_counts().head(bug_count).index
                    top_group = group[group['gt.id'].isin(top_values)]
                    bug_index_dict['false_type'].extend(top_group['corresponding_index'].to_list())

            else:
                distance_n_f = 50
                error_data = metric_data[(metric_data['is_abnormal'] == 1)
                                         & (metric_data['gt.x'] <= 100)
                                         & (metric_data['gt.x'] >= -50)
                                         & (metric_data['gt.y'] <= 8)
                                         & (metric_data['gt.y'] >= -8)
                                         & (metric_data['gt.type'] == metric_data['pred.type'])]
                error_data = error_data[~((error_data['gt.type'].isin(['pedestrian', 'cyclist']))
                                          & (error_data['gt.x'] > 50))]
                if metric in ['width_error', 'length_error', 'height_error']:
                    error_data = error_data[~error_data['gt.type'].isin(['pedestrian', 'cyclist'])]
                if metric in ['vx_error', 'vy_error', 'yaw_error']:
                    error_data = error_data[~error_data['gt.type'].isin(['pedestrian'])]
                    error_data = error_data[~((error_data['gt.type'].isin(['cyclist']))
                                              & (error_data['gt.vel'] <= 2))]
                if metric == 'yaw_error':
                    error_data = error_data[error_data['yaw.error_abs'] > 20]

                bug_index_dict[metric] = []
                for key, group in error_data.groupby('gt.type'):
                    # 进一步筛选出20米内误差绝对值最大，20米外误差相对值最大的前5个
                    near_group = group[(group['gt.x'] <= distance_n_f) & (group['gt.x'] >= -distance_n_f)]
                    id_counts = near_group['gt.id'].value_counts()
                    frequent_ids = id_counts[id_counts > 10].index
                    filtered_df = near_group[near_group['gt.id'].isin(frequent_ids)]
                    mean_values = filtered_df.groupby('gt.id')[f'{metric.replace("_", ".")}_abs'].mean().reset_index()
                    top_5_ids = mean_values.sort_values(by=f'{metric.replace("_", ".")}_abs', ascending=False).head(bug_count)['gt.id']
                    near_top_5 = filtered_df[filtered_df['gt.id'].isin(top_5_ids)]

                    far_group = group[(group['gt.x'] > distance_n_f) | (group['gt.x'] < -distance_n_f)]
                    if metric != 'yaw_error':
                        col = f'{metric.replace("_", ".")}%_abs'
                    else:
                        col = f'{metric.replace("_", ".")}_abs'
                    id_counts = far_group['gt.id'].value_counts()
                    frequent_ids = id_counts[id_counts > 10].index
                    filtered_df = far_group[far_group['gt.id'].isin(frequent_ids)]
                    mean_values = filtered_df.groupby('gt.id')[col].mean().reset_index()
                    top_5_ids = mean_values.sort_values(by=col, ascending=False).head(bug_count)['gt.id']
                    far_top_5 = filtered_df[filtered_df['gt.id'].isin(top_5_ids)]

                    top_group = pd.concat([near_top_5, far_top_5])
                    bug_index_dict[metric].extend(top_group['corresponding_index'].to_list())

        return bug_index_dict

    def plot_one_frame_for_obstacles(self, topic, frame_data, plot_path, frame_bug_info_bev=None):

        def plot_rectangle(center_x, center_y, yaw, length, width, text, color, visibility, fill):
            res = np.array([[np.cos(yaw), -np.sin(yaw)], [np.sin(yaw), np.cos(yaw)]]) @ np.array(
                [[-length / 2], [-width / 2]])
            xy = (res[0][0] + center_x, res[1][0] + center_y)
            rotation = yaw * 57.3

            ax.add_patch(pc.Rectangle(
                xy=xy,
                width=length, height=width,
                angle=rotation, alpha=visibility,
                rotation_point='xy',
                facecolor='lightgrey',
                fill=fill,
                edgecolor=color,
                linewidth=2))

            arrow_length = length / 2 + 2
            dx, dy = arrow_length * np.cos(yaw) / 2, arrow_length * np.sin(yaw) / 2
            ax.arrow(center_x + dx, center_y + dy, dx, dy,
                     length_includes_head=True,
                     head_width=0.5, head_length=1.3, fc='r', ec='r')

            while rotation < -90:
                rotation += 180
            while rotation > 90:
                rotation -= 180
            ax.text(center_x, center_y, text, rotation=rotation,
                    va='center', ha='center',
                    fontdict={
                        'family': 'Ubuntu',
                        'style': 'normal',
                        'weight': 'normal',
                        'color': 'black',
                        'size': 8,
                    })

        def plot_outline(center_x, center_y, yaw, length, width, text):
            length += 3
            width += 2

            res = np.array([[np.cos(yaw), -np.sin(yaw)], [np.sin(yaw), np.cos(yaw)]]) @ np.array(
                [[-length / 2], [-width / 2]])
            xy = (res[0][0] + center_x, res[1][0] + center_y)
            rotation = yaw * 57.3

            ax.add_patch(pc.Rectangle(
                xy=xy,
                width=length, height=width,
                angle=rotation, alpha=1,
                rotation_point='xy',
                fill=False,
                edgecolor='lightcoral',
                linestyle='dashed',
                linewidth=2))

            while rotation < -90:
                rotation += 180
            while rotation > 90:
                rotation -= 180
            res = np.array([[np.cos(yaw), -np.sin(yaw)], [np.sin(yaw), np.cos(yaw)]]) @ np.array(
                [[0], [width / 2 + 1]])
            ax.text(res[0][0] + center_x, res[1][0] + center_y, text, rotation=rotation,
                    va='center', ha='center',
                    fontdict={
                        'family': 'Ubuntu',
                        'style': 'normal',
                        'weight': 'normal',
                        'color': 'darkred',
                        'size': 11,
                    })

        def plot_one_frame(ax, data, title, ego_velocity_generator=None, bug_info=None):
            if bug_info is None:
                bug_info = {}
            ax.patch.set_facecolor('white')
            for pos in ['right', 'left']:
                ax.spines[pos].set_visible(False)

            ax.add_patch(pc.Rectangle(
                (-1.0, -0.95), 4.6, 1.9, edgecolor='mediumslateblue', facecolor='limegreen'))
            ax.arrow(3, 0, 1.3, 0,
                     length_includes_head=True,
                     head_width=0.5, head_length=1.3, fc='r', ec='r')

            time_stamp = data.iloc[0]['time_stamp']
            if ego_velocity_generator is not None:
                ego_velocity = ego_velocity_generator([time_stamp])[0]
                ax.text(0, 20, f'{round(ego_velocity * 3.6)} km/h',
                        va='center', ha='left',
                        fontdict={
                            'family': 'Ubuntu',
                            'style': 'normal',
                            'weight': 'bold',
                            'color': 'orange',
                            'size': 20
                        })

            for idx, row in data.iterrows():
                if row['flag'] == 0:
                    continue

                color = color_type[row['type_classification']]
                fill = False
                if row['is_keyObj'] == 1:
                    visibility = 1
                    fill = True
                elif row['is_coverageValid'] == 1:
                    visibility = 0.5
                elif row['is_detectedValid'] == 1:
                    visibility = 0.3
                else:
                    visibility = 0.1

                plot_rectangle(center_x=row['x'], center_y=row['y'],
                               yaw=row['yaw'], length=row['length'],
                               width=row['width'], text=int(row['id']),
                               color=color, visibility=visibility, fill=fill)

                # 画出cipv
                if row['is_cipv'] == 1:
                    ax.annotate('', xy=(row['x'], row['y'] + 1.2), xytext=(row['x'], row['y'] + 3),
                                arrowprops=dict(facecolor='red', shrink=0.05))

            for bug_type, id_list in bug_info.items():
                for id_ in id_list:
                    id_data = data[data['id'] == id_]
                    if len(id_data):
                        row = id_data.iloc[0]
                        plot_outline(center_x=row['x'], center_y=row['y'],
                                     yaw=row['yaw'], length=row['length'],
                                     width=row['width'], text=bug_type)

                        ax.plot([row['x'] - 7, row['x'] + 7], [row['y'], row['y']],
                                linestyle='--', color='lightcoral', alpha=0.5)
                        ax.plot([row['x'], row['x']], [row['y'] - 4, row['y'] + 4],
                                linestyle='--', color='lightcoral', alpha=0.5)

            ax.set_xlim(-100, 150)
            ax.set_xticks(np.arange(-100, 151, 50))
            ax.set_xticklabels([f'{x:.0f} m' for x in np.arange(-100, 151, 50)])
            ax.set_ylim(-25, 25)
            ax.set_yticks([-16, -8, 0, 8, 16])
            ax.set_yticklabels([f'{y:.0f} m' for y in [-16, -8, 0, 8, 16]])
            ax.tick_params(direction='out', labelsize=font_size * 1.2, length=4)
            ax.set_title(title, fontdict=title_font, y=0.9, loc='left')
            red_triangle = mlines.Line2D([], [], color='red', marker='v', linestyle='None',
                                         markersize=10, label='cipv')
            ego_rectangle = mlines.Line2D([], [], color='none', marker='s', linestyle='None',
                                          markerfacecolor='limegreen', markersize=12, label='ego_car')
            car_rectangle = mlines.Line2D([], [], color='none', marker='s', linestyle='None',
                                          mec='#3682be', mfc='lightgrey', markersize=12, label='small-medium car')
            pedestrian_rectangle = mlines.Line2D([], [], color='none', marker='s', linestyle='None',
                                                 mec='#45a776', mfc='lightgrey', markersize=12, label='pedestrian')
            bus_rectangle = mlines.Line2D([], [], color='none', marker='s', linestyle='None',
                                          mec='#f05326', mfc='lightgrey', markersize=12, label='truck_bus')
            truck_rectangle = mlines.Line2D([], [], color='none', marker='s', linestyle='None',
                                            mec='#800080', mfc='lightgrey', markersize=12, label='truck')
            cyclist_rectangle = mlines.Line2D([], [], color='none', marker='s', linestyle='None',
                                              mec='#334f65', mfc='lightgrey', markersize=12, label='cyclist')

            # 将这两个图形添加到图例中
            ax.legend(handles=[red_triangle, ego_rectangle, car_rectangle,
                               pedestrian_rectangle, bus_rectangle, truck_rectangle, cyclist_rectangle], fontsize=13)
            ax.grid(linestyle='-', linewidth=0.5, color='lightgray', alpha=0.6)

        color_type = {
            'car': '#3682be',
            'pedestrian': '#45a776',
            'truck_bus': '#f05326',
            'truck': '#800080',
            'cyclist': '#334f65',
        }

        if self.ego_velocity_generator is None:
            if self.gt_ego_flag:
                ego_data = pd.read_csv(self.get_abspath(self.test_result['General']['gt_ego']), index_col=False)
                self.ego_velocity_generator = interp1d(ego_data['time_stamp'].values, ego_data['ego_vx'].values, kind='linear')
            else:
                ego_data = pd.read_csv(self.get_abspath(self.test_result['General']['pred_ego']), index_col=False)
                self.ego_velocity_generator = interp1d(ego_data['time_stamp'].values + self.test_result['General']['time_gap'],
                                                       ego_data['ego_vx'].values, kind='linear')

        # 开始画图
        fig = plt.figure(figsize=(25, 12))
        fig.tight_layout()
        plt.subplots_adjust(left=0.03, right=0.97, top=0.97, bottom=0.03)
        grid = plt.GridSpec(2, 1, wspace=0.1, hspace=0.15)

        # 上图为gt，下图为pred
        ax = fig.add_subplot(grid[0, 0])
        data = pd.DataFrame()
        for col in frame_data.columns:
            if 'pred.' in col:
                continue
            if 'gt.' in col:
                data[col.split('.')[-1]] = frame_data[col]
            else:
                data[col] = frame_data[col]
        if frame_bug_info_bev is not None:
            bug_info = frame_bug_info_bev['gt']
        else:
            bug_info = None
        time_stamp = data.iloc[0]['time_stamp']
        title = f'GroundTruth <{topic}@{round(time_stamp, 3)}s >'
        plot_one_frame(ax, data, title, self.ego_velocity_generator, bug_info)

        ax = fig.add_subplot(grid[1, 0])
        data = pd.DataFrame()
        for col in frame_data.columns:
            if 'gt.' in col:
                continue
            if 'pred.' in col:
                data[col.split('.')[-1]] = frame_data[col]
            else:
                data[col] = frame_data[col]
        if frame_bug_info_bev is not None:
            bug_info = frame_bug_info_bev['pred']
        else:
            bug_info = None
        time_stamp = data.iloc[0]['time_stamp']
        title = f'Prediction <{topic}@{round(time_stamp, 3)}s >'
        plot_one_frame(ax, data, title, self.ego_velocity_generator, bug_info)

        canvas = FigureCanvas(fig)
        canvas.print_figure(plot_path, facecolor='white', dpi=100)
        fig.clf()
        plt.close()

    def start(self):

        if self.test_action['preprocess']:
            self.load_pred_data()
            self.load_gt_data()
            self.sync_timestamp()
            self.promote_additional()

        if self.test_action['match']:
            self.match_timestamp()
            self.match_object()

        if self.test_action['metric']:
            self.evaluate_metrics()

        if self.test_action['bug']:
            self.load_scenario_info()
            self.sketch_bug()
            self.bug_report()

        if self.test_action['render']:
            self.sketch_render()
            self.generate_image_and_video()

    def which_camera_saw_you(self, x, y):
        valid_camera = []
        for camera in self.camera_list:
            camera_parameter = self.test_result['General']['camera_position'][camera]
            azimuth = np.arctan2(y - camera_parameter['y'], x - camera_parameter['x'])
            if azimuth < 0:
                azimuth += 2 * np.pi
            azimuth = np.rad2deg(azimuth)

            angle_min, angle_max = camera_parameter['fov'][0], camera_parameter['fov'][1]
            if angle_min >= 0:
                if angle_min <= azimuth < angle_max:
                    valid_camera.append(camera)
            else:
                if angle_min + 360 <= azimuth or azimuth <= angle_max:
                    valid_camera.append(camera)

        return valid_camera


class DataGrinderPilotObstaclesOneTask(DataGrinderOneTask):

    @sync_test_result
    def combine_scenario_tag(self):
        # 合并各个场景的match_data, 重新corresponding_index编号
        # 获得每一种特征的分类结果
        for scenario_tag in self.test_config['scenario_tag']:
            # 需要去除Broken Scenario
            valid_scenario_list = [scenario_id for scenario_id in scenario_tag['scenario_id']
                                   if scenario_id in self.valid_scenario_list]
            if len(valid_scenario_list):
                tag_key = '&'.join(scenario_tag['tag'].values())
                self.test_result['TagCombination'][tag_key] = {
                    'tag': scenario_tag['tag'],
                    'scenario_id': valid_scenario_list,
                    self.test_topic: {},
                }

        match_data_dict = {}
        for tag_key in self.test_result['TagCombination'].keys():

            # 汇总数据
            tag_combination_folder = os.path.join(self.tag_combination_folder, tag_key)
            create_folder(tag_combination_folder)
            match_data_dict[tag_key] = {}

            for scenario_id in self.test_result['TagCombination'][tag_key]['scenario_id']:
                scenario_unit_folder = os.path.join(self.scenario_unit_folder, scenario_id)
                with open(os.path.join(scenario_unit_folder, 'TestResult.yaml')) as f:
                    scenario_test_result = yaml.load(f, Loader=yaml.FullLoader)

                # 遍历全部测试结果，按照topic和metrics合并
                for topic in scenario_test_result[self.test_topic].keys():
                    if topic == 'GroundTruth':
                        continue

                    topic_tag = topic.replace('/', '')
                    # 用于设置评判样本量最小限度的帧数阈值
                    if topic not in self.test_result['TagCombination'][tag_key][self.test_topic]:
                        self.test_result['TagCombination'][tag_key][self.test_topic][topic] = {
                            'frequency': scenario_test_result[self.test_topic][topic]['match_frequency']
                        }
                        topic_folder = os.path.join(tag_combination_folder, topic_tag)
                        create_folder(topic_folder)

                    if topic not in match_data_dict[tag_key]:
                        match_data_dict[tag_key][topic] = []

                    print(scenario_id)
                    match_data_path = scenario_test_result[self.test_topic][topic]['match']['match_data']
                    match_data = pd.read_csv(os.path.join(scenario_unit_folder, match_data_path), index_col=False)
                    match_data.insert(0, 'scenario_id', scenario_id)
                    match_data_dict[tag_key][topic].append(match_data)

        for tag_key in match_data_dict.keys():

            send_log(self, f'{self.test_topic}, 使用{self.test_topic}MetricEvaluator')

            for topic, df_list in match_data_dict[tag_key].items():
                topic_tag = topic.replace('/', '')
                metric_folder = os.path.join(self.tag_combination_folder, tag_key, topic_tag)
                create_folder(metric_folder)

                total_match_data = pd.concat(df_list).reset_index(drop=True)
                total_match_data['corresponding_index'] = total_match_data.index
                total_match_data_path = os.path.join(get_project_path(), 'Temp', 'total_match_data.csv')
                total_match_data.to_csv(total_match_data_path, index=False, encoding='utf_8_sig')

                input_parameter_container = {
                    'metric_type': self.test_config['test_item'][topic],
                    'characteristic_type': self.test_config['target_characteristic'],
                    'kpi_date_label': self.kpi_date_label,
                }
                parameter_json_path = os.path.join(get_project_path(), 'Temp', 'evaluate_api_parameter.json')
                with open(parameter_json_path, 'w', encoding='utf-8') as f:
                    json.dump(input_parameter_container, f, ensure_ascii=False, indent=4)

                cmd = [
                    f"{bench_config['master']['sys_interpreter']}",
                    "Api_EvaluateMetrics.py",
                    "-m", total_match_data_path,
                    "-j", parameter_json_path,
                    "-f", metric_folder,
                ]

                print(cmd)
                send_log(self, f'{self.test_topic} {topic} 指标评估')
                cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
                result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
                # os.remove(parameter_json_path)
                os.remove(total_match_data_path)
                if result.stderr and 'warning' not in result.stderr:
                    print(result.stderr)
                    send_log(self, f'EvaluateMetrics 发生错误 {result.stderr}')

                for characteristic in os.listdir(metric_folder):
                    self.test_result['TagCombination'][tag_key][self.test_topic][topic][characteristic] = {}
                    for metric in os.listdir(os.path.join(metric_folder, characteristic)):
                        self.test_result['TagCombination'][tag_key][self.test_topic][topic][characteristic][
                            metric.split('.')[0]] = self.get_relpath(
                            os.path.join(metric_folder, characteristic, metric))

    @sync_test_result
    def summary_bug_items(self):

        bugItem_folder = os.path.join(self.output_result_folder, 'bugItems')
        create_folder(bugItem_folder)
        dataframes = []

        for scenario_tag in self.test_result['TagCombination']:
            scenario_ids = self.test_result['TagCombination'][scenario_tag]['scenario_id']
            for scenario_id in scenario_ids:
                for root, dirs, files in os.walk(os.path.join(self.scenario_unit_folder, scenario_id)):
                    for file in files:
                        if 'bug_jira_summary.csv' in file:
                            file_path = os.path.join(root, file)
                            df = pd.read_csv(file_path, index_col=False)
                            if len(df):
                                dataframes.append(df)

        if len(dataframes):
            bug_summary = pd.concat(dataframes, ignore_index=True)
            bug_report_path_list = []
            for i, row in bug_summary.iterrows():
                topic = row['topic'].replace('/', '')
                bug_type = row['bug_type']
                target_type = row['target_type']
                folder = os.path.join(bugItem_folder, bug_type, topic, target_type)
                if not os.path.exists(folder):
                    os.makedirs(folder)

                bug_report_path = row['attachment_path']
                shutil.copy(bug_report_path, folder)
                bug_report_path_list.append(os.path.join(folder, os.path.basename(bug_report_path)))

            bug_summary['attachment_path'] = bug_report_path_list
            bug_summary_path = os.path.join(bugItem_folder, 'bug_summary.csv')
            (bug_summary.sort_values(by=['bug_type', 'scenario_id', 'gt_target_id'])
             .to_csv(bug_summary_path, index=False, encoding='utf_8_sig'))
            self.test_result['OutputResult']['bugItems'] = self.get_relpath(bug_summary_path)

        # 获取版本对比的文件夹
        version_comparison_folder = self.test_config['version_comparison']
        if version_comparison_folder is not None and os.path.exists(version_comparison_folder):
            for root, dirs, files in os.walk(version_comparison_folder):
                for file in files:
                    if 'bug_summary.csv' in file and self.test_topic in root:
                        print(os.path.join(root, file))

        # 1. 上个版本没有，这个版本出现

    @sync_test_result
    def compile_statistics(self):
        # 按照数据库数据单元的方式保存数据
        # 格式为json，在文件夹内平铺
        statistics_folder = os.path.join(self.output_result_folder, 'statistics')
        create_folder(statistics_folder)

        # 统计目标个数
        total_scenario_analysis_path = os.path.join(project_path, 'Docs', 'Resources', 'scenario_info', 'scenario_analysis.json')
        with open(total_scenario_analysis_path, 'r', encoding='utf-8') as f:
            total_scenario_analysis = json.load(f)

        scenario_analysis = {}
        for tag_key in self.test_result['TagCombination'].keys():
            scenario_list = self.test_result['TagCombination'][tag_key]['scenario_id']
            scenario_analysis[tag_key] = {}

            for scenario_id in scenario_list:
                if f'{scenario_id}-{self.truth_source}' not in total_scenario_analysis.keys():
                    continue

                scenario_analysis[tag_key][scenario_id] = {}
                for characteristic in self.test_config['test_action']['output_characteristic']:
                    scenario_analysis[tag_key][scenario_id][characteristic] = (
                        total_scenario_analysis)[f'{scenario_id}-{self.truth_source}']['obstacles'][characteristic]

        scenario_analysis_path = os.path.join(statistics_folder, 'scenario_analysis.json')
        with open(scenario_analysis_path, 'w', encoding='utf-8') as f:
            json.dump(scenario_analysis, f, ensure_ascii=False, indent=4)

        self.test_result['OutputResult']['scenario_analysis'] = self.get_relpath(scenario_analysis_path)
        # ====================

        # 生成用于上传数据库的json文件，以及汇总结果OutputResult
        json_count = 0
        json_rows = []
        for tag_key in self.test_result['TagCombination'].keys():
            tag = self.test_result['TagCombination'][tag_key]['tag']
            scenario_list = self.test_result['TagCombination'][tag_key]['scenario_id']

            metric_statistics = eval(f'MetricStatistics.{self.test_topic}MetricStatistics()')

            for topic in self.test_result['TagCombination'][tag_key][self.test_topic].keys():
                topic_tag = topic.replace('/', '')

                for characteristic in self.test_config['target_characteristic']:
                    if characteristic not in self.test_result['TagCombination'][tag_key][self.test_topic][topic]:
                        continue

                    if characteristic not in self.test_config['test_action']['output_characteristic']:
                        continue

                    test_result = self.test_result['TagCombination'][tag_key][self.test_topic][topic][
                        characteristic]

                    info_json_data = {
                        'Product': self.product,
                        'Version': self.version,
                        'StartTime': str(self.test_date),
                        **{tag_type: tag_value for tag_type, tag_value in tag.items()},
                        'ScenarioGroup': scenario_list,
                        'TopicName': topic}

                    data = {
                        metric: pd.read_csv(self.get_abspath(data_path), index_col=False)
                        for metric, data_path in test_result.items()
                    }

                    input_parameter_container = {
                        'region_division': self.test_config['region_division'],
                        'characteristic': characteristic,
                    }

                    json_datas = metric_statistics.run(data, input_parameter_container)

                    for json_data in json_datas:
                        json_data = {**info_json_data, **json_data}

                        # 在生成测试结果表格的时候，做一下筛选:
                        # 1.速度仅在后处理存在
                        # 2.行人不存在航向角
                        # 3.快速和高速场景不存在行人和两轮车
                        # 4.行人和两轮车没有那么远的测试距离

                        if '速度' in json_data['MetricTypeName'] and json_data['TopicName'] != '/VA/Obstacles':
                            continue

                        if json_data['MetricTypeName'] == '航向角误差' and json_data['ObstacleName'] == '行人':
                            continue

                        if (('快速' in json_data['RoadTypeCondition'] or '高速' in json_data['RoadTypeCondition'])
                                and (json_data['ObstacleName'] in ['行人', '两轮车'])):
                            continue

                        if ('100~150' in json_data['RangeDetails'] and json_data['ObstacleName'] in ['行人', '两轮车']) \
                                or ('-100~-50' in json_data['RangeDetails'] and json_data['ObstacleName'] == '行人'):
                            continue

                        if json_data['TopicName'] == '/VA/PedResult' and json_data['ObstacleName'] in ['小车', '大车']:
                            continue

                        if json_data['TopicName'] == '/VA/VehicleResult' and json_data['ObstacleName'] in ['两轮车', '行人']:
                            continue

                        # 保存单个json文件
                        json_count += 1
                        json_name = (f'{json_count:06d}--{self.version}--{tag_key}--{topic_tag}--{json_data["ObstacleName"]}'
                                     f'--{json_data["RangeDetails"]}--{json_data["FeatureDetail"]}--{json_data["MetricTypeName"]}.json')
                        json_path = os.path.join(statistics_folder, json_name)

                        frequency_threshold = self.test_result['TagCombination'][tag_key][self.test_topic][topic][
                                                  'frequency'] * 2
                        if json_data['Output']['sample_count'] < frequency_threshold:
                            send_log(self, f'{json_count} {json_name} 样本少于{frequency_threshold}，不保存')
                            continue

                        json_data['ScenarioGroup'] = '&'.join(json_data['ScenarioGroup'])
                        json_data['uuid'] = str(uuid.uuid4())
                        send_log(self, f'{json_count} {json_name} 已保存')
                        with open(json_path, 'w', encoding='utf-8') as f:
                            json.dump(json_data, f, ensure_ascii=False, indent=4)

                        # 保存json文件的目录
                        json_rows.append(
                            [
                                json_count, self.version, tag_key, topic,
                                json_data['ObstacleName'], json_data['RangeDetails'],
                                json_data['FeatureDetail'], json_data['MetricTypeName'],
                                json_data['Output']['sample_count'], json.dumps(json_data['Output'])
                            ]
                        )

        path = os.path.join(statistics_folder, 'output_result.csv')
        output_result = pd.DataFrame(json_rows, columns=[
            'index', 'version', 'scenario_tag', 'topic', 'obstacle_type', 'region',
            'characteristic', 'metric', 'sample_count', 'result'])
        output_result.to_csv(path, index=False, encoding='utf_8_sig')
        self.test_result['OutputResult']['statistics'] = self.get_relpath(path)
        # ====================

        # 生成以场景类型作为sheet的xlsx
        characteristic_key = [test_encyclopaedia['Information'][self.test_topic]['characteristic'][c]['name']
                              for c in self.test_config['test_action']['output_characteristic']]
        topic_key = output_result['topic'].unique()
        scenario_tag_key = output_result['scenario_tag'].unique()
        obstacle_type_key = output_result['obstacle_type'].unique()
        region_key = output_result['region'].unique()
        metric_key = [v['name'] for v in test_encyclopaedia['Information'][self.test_topic]['metrics'].values()
                      if v['name'] in output_result['metric'].unique()]
        # 首先生成columns
        columns = [
            ('基本信息', '基本信息', 'OD类型'),
            ('基本信息', '基本信息', '目标物范围'),
            ('基本信息', '基本信息', 'start_distance'),
            ('基本信息', '基本信息', '场景类型'),
        ]
        for metric in metric_key:
            temp_data = output_result[output_result['metric'] == metric]
            for res_key in json.loads(temp_data.iloc[0]['result']).keys():
                if (
                        (metric == '准召信息') or
                        (metric in ['长度误差', '宽度误差', '高度误差'] and 'abs_95' in res_key) or
                        ('距离误差' in metric and 'abs_95' in res_key) or
                        ('abs_95' in res_key and '%' not in res_key)
                ):
                    if res_key not in ['TP', 'FP', 'FN', 'CTP', 'sample_count']:
                        columns.append(
                            (metric, res_key, 'kpi_target')
                        )
                    for topic in topic_key:
                        columns.append(
                            (metric, res_key, topic)
                        )

        self.test_result['OutputResult']['report_table'] = {}
        for i, characteristic in enumerate(characteristic_key):
            path = os.path.join(statistics_folder, f'{characteristic}-指标汇总.xlsx')
            self.test_result['OutputResult']['report_table'][characteristic] = self.get_relpath(path)

            # 使用ExcelWriter将DataFrame保存到不同的sheet中
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                for scenario_tag in scenario_tag_key:
                    rows = []
                    for obstacle_type in obstacle_type_key:
                        if '快速' in scenario_tag and obstacle_type in ['行人', '两轮车']:
                            continue

                        for region in region_key:
                            if ('100~150' in region and obstacle_type in ['行人', '两轮车']) \
                                or ('-100~-50' in region and obstacle_type == '行人'):
                                continue

                            start_distance = int(region.split('~')[0].split('(')[1])
                            row = [obstacle_type, region, start_distance, scenario_tag]
                            filter_data = output_result[
                                (output_result['characteristic'] == characteristic)
                                & (output_result['scenario_tag'] == scenario_tag)
                                & (output_result['obstacle_type'] == obstacle_type)
                                & (output_result['region'] == region)
                            ]

                            for metric, res_key, topic in columns[4:]:
                                if topic == 'kpi_target':
                                    kpi_threshold = get_obstacles_kpi_threshold(metric, res_key, obstacle_type, region=region)
                                    if kpi_threshold is None:
                                        row.append(np.nan)
                                    else:
                                        kpi_ratio = get_obstacles_kpi_ratio(metric, res_key, obstacle_type, self.kpi_date_label)
                                        row.append(kpi_threshold * kpi_ratio)

                                else:
                                    filter2_data = filter_data[
                                        (filter_data['metric'] == metric)
                                        & (filter_data['topic'] == topic)
                                    ]

                                    if len(filter2_data):
                                        c = json.loads(filter2_data.iloc[0]['result'])
                                        row.append(c[res_key])
                                    else:
                                        row.append(np.nan)

                            rows.append(row)

                    df = pd.DataFrame(rows, columns=pd.MultiIndex.from_tuples(columns))
                    df.sort_values(by=[('基本信息', '基本信息', 'OD类型'), ('基本信息', '基本信息', 'start_distance')], inplace=True)
                    df.drop(('基本信息', '基本信息', 'start_distance'), axis=1, inplace=True)
                    df.to_excel(writer, sheet_name=scenario_tag, merge_cells=True)

            workbook = Workbook()
            workbook.LoadFromFile(path)
            for sheet in workbook.Worksheets:
                sheet.DeleteRow(4)
                sheet.DeleteColumn(1)
            workbook.SaveToFile(path)
            workbook.Dispose()
            sheet_name1 = 'Evaluation Warning'
            workbook = openpyxl.load_workbook(path)
            worksheet = workbook[sheet_name1]
            workbook.remove(worksheet)
            workbook.save(path)
        # ====================

    @sync_test_result
    def visualize_output(self):

        def plot_table(axes, columns, columns_width, values, index=None, index_width=0):
            if index is None:
                index = []
            axes.patch.set_facecolor('white')
            axes.patch.set_alpha(0.3)
            for pos in ['right', 'top', 'left', 'bottom']:
                axes.spines[pos].set_visible(False)

            # 画index
            if index_width and len(index):
                axes.add_patch(pc.Rectangle(
                    xy=(-index_width, 0),
                    width=index_width, height=1,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=True,
                    edgecolor='grey',
                    facecolor='bisque',
                    linewidth=1))
                axes.text(-index_width / 2, 0.5, 'Test Case',
                          va='center', ha='center',
                          fontdict={
                              'family': 'Ubuntu',
                              'style': 'normal',
                              'weight': 'bold',
                              'color': 'black',
                              'size': font_size * 1.5,
                          })

                for i, idx in enumerate(index):
                    axes.add_patch(pc.Rectangle(
                        xy=(-index_width, -1 - i),
                        width=index_width, height=1,
                        angle=0, alpha=1,
                        rotation_point='xy',
                        fill=True,
                        edgecolor='grey',
                        facecolor='linen',
                        linewidth=1))

                    axes.text(-index_width * 0.95, -1 - i + 0.5, idx,
                              va='center', ha='left',
                              fontdict={
                                  'family': 'sans-serif',
                                  'style': 'normal',
                                  'weight': 'normal',
                                  'color': 'black',
                                  'size': font_size * 1.5,
                              })

            # 画columns
            current_x = 0
            for i, (col, width) in enumerate(zip(columns, columns_width)):
                axes.add_patch(pc.Rectangle(
                    xy=(current_x, 0),
                    width=width, height=1,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=True,
                    edgecolor='grey',
                    facecolor='lightsteelblue',
                    linewidth=1))

                axes.text(current_x + width * 0.5, 0.5, col,
                          va='center', ha='center',
                          fontdict={
                              'family': 'Ubuntu',
                              'style': 'normal',
                              'weight': 'bold',
                              'color': 'black',
                              'size': font_size * 1.5,
                          })

                current_x += width

            # 画values
            ii, jj = values.shape
            for i in range(ii):
                current_x = 0
                for j in range(jj):
                    width = columns_width[j]
                    color = 'white'
                    value = values[i, j]

                    if '%' in columns[j]:
                        value = '{:.2%}'.format(value)

                    elif isinstance(value, float) and (not np.isnan(value)):
                        if int(value) == value:
                            value = int(value)
                        else:
                            value = round(value, 4)

                    if contains_chinese(str(value)):
                        fontdict = {
                            'family': 'sans-serif',
                            'style': 'normal',
                            'weight': 'normal',
                            'color': 'black',
                            'size': font_size * 1.5,
                        }
                    else:
                        fontdict = {
                            'family': 'Ubuntu',
                            'style': 'normal',
                            'weight': 'normal',
                            'color': 'black',
                            'size': font_size * 1.5,
                        }

                    axes.add_patch(pc.Rectangle(
                        xy=(current_x, -i - 1),
                        width=width, height=1,
                        angle=0, alpha=1,
                        rotation_point='xy',
                        fill=True,
                        edgecolor='grey',
                        facecolor=color,
                        linewidth=1))
                    axes.text(current_x + width * 0.5, -i - 1 + 0.5, value,
                              va='center', ha='center',
                              fontdict=fontdict)

                    current_x += width

            axes.set_xlim(-index_width - 0.1, np.sum(columns_width) + 0.1)
            axes.set_ylim(-len(df) - 0.1, 2.1)
            axes.set_xticks([])
            axes.set_yticks([])

        def wrap_text(text, max_length):
            wrapped_lines = []
            for i in range(0, len(text), max_length):
                wrapped_lines.append(text[i:i + max_length])
            return '\n'.join(wrapped_lines)

        visualization_folder = os.path.join(self.output_result_folder, 'visualization')
        create_folder(visualization_folder)

        # 生成测试信息总览
        def plot_test_info(test_info, plot_path):

            def plot_rect(left_bottom_pt, width, height,
                          facecolor, text, text_angle,
                          font_size, fill=True):
                ax.add_patch(pc.Rectangle(
                    xy=left_bottom_pt,
                    width=width, height=height,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=fill,
                    edgecolor='grey',
                    facecolor=facecolor,
                    linewidth=1))
                va, ha = 'center', 'center'
                ax.text(left_bottom_pt[0] + width / 2, left_bottom_pt[1] + height / 2, text,
                        va=va, ha=ha, rotation=text_angle,
                        fontsize=font_size)

            cell_w = 2
            cell_h = 0.6
            font_size = 13

            fig_height = 0
            for i, (key, info) in enumerate(test_info.items()):
                if not isinstance(info, dict):
                    fig_height += cell_h
                else:
                    fig_height += cell_h * len(info.keys()) * 0.9
            fig_width = cell_w * 6

            fig_size = (fig_width + 0.1, fig_height + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            ax.axis('off')

            left_bottom_y = 0
            for i, (key, info) in enumerate(test_info.items()):
                if key == '测试版本':
                    left_bottom_y -= cell_h
                    plot_rect(left_bottom_pt=[0, left_bottom_y],
                              width=fig_width, height=cell_h, facecolor='lightsteelblue',
                              text=f'<{info}>测试报告信息总览', text_angle=0, font_size=font_size * 1.2)
                elif not isinstance(info, dict):
                    left_bottom_y -= cell_h
                    plot_rect(left_bottom_pt=[cell_w, left_bottom_y],
                              width=cell_w * 5, height=cell_h, facecolor='white',
                              text=info, text_angle=0, font_size=font_size)
                    plot_rect(left_bottom_pt=[0, left_bottom_y],
                              width=cell_w, height=cell_h, facecolor='lightsteelblue',
                              text=key, text_angle=0, font_size=font_size)
                else:
                    left_bottom_y -= cell_h * len(info.keys()) * 0.9
                    text = '\n\n'.join([f'{k}:{v}' for k, v in info.items()])
                    plot_rect(left_bottom_pt=[cell_w, left_bottom_y],
                              width=cell_w * 5, height=cell_h * len(info.keys()) * 0.9, facecolor='white',
                              text=text, text_angle=0, font_size=font_size)
                    plot_rect(left_bottom_pt=[0, left_bottom_y],
                              width=cell_w, height=cell_h * len(info.keys()) * 0.9, facecolor='lightsteelblue',
                              text=key, text_angle=0, font_size=font_size)

            ax.set_xlim(-0.1, fig_width + 0.1)
            ax.set_ylim(-fig_height - 0.1, 0.1)
            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)

        def plot_title_table(title_dict, plot_path):
            fig = plt.figure(figsize=(4 * len(title_dict), 3.5))
            fig.tight_layout()
            plt.subplots_adjust(left=0.05, right=0.95, top=0.95, bottom=0.05)
            grid = plt.GridSpec(1, 1, wspace=0.05, hspace=0.05)
            ax_title_page = fig.add_subplot(grid[0, 0])
            for pos in ['right', 'top', 'left', 'bottom']:
                ax_title_page.spines[pos].set_visible(False)

            center_x = 2
            for k, v in title_dict.items():
                ax_title_page.text(center_x, 0.875, k,
                                   va='center', ha='center',
                                   fontdict={
                                       'style': 'normal',
                                       'weight': 'normal',
                                       'color': 'black',
                                       'size': font_size * 2.5,
                                   })
                ax_title_page.text(center_x, -0.875, str(v),
                                   va='center', ha='center',
                                   fontdict={
                                       'family': 'Ubuntu',
                                       'style': 'normal',
                                       'weight': 'bold',
                                       'color': 'orange',
                                       'size': font_size * 6,
                                   })
                center_x += 4

            ax_title_page.set_xlim(0, len(title_dict) * 4)
            ax_title_page.set_ylim(-1.75, 1.75)
            ax_title_page.set_xticks([])
            ax_title_page.set_yticks([])

            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)
            fig.clf()
            plt.close()

        current_version = self.test_config['version']
        test_topic = self.test_config['test_topic']
        version_comparison_folder = self.test_config['version_comparison']
        old_version = ''
        if version_comparison_folder is not None and os.path.exists(version_comparison_folder):
            for root, dirs, files in os.walk(version_comparison_folder):
                for file in files:
                    if 'TestConfig.yaml' in file:
                        file_path = os.path.join(root, file)
                        with open(file_path, 'r', encoding='utf-8') as file:
                            old_test_config = yaml.safe_load(file)
                        if 'version_comparison' in old_test_config and old_test_config['test_topic'] == test_topic:
                            old_version = old_test_config['version']

        output_result = pd.read_csv(self.get_abspath(self.test_result['OutputResult']['statistics']), index_col=False)
        region_key = ';'.join(output_result['region'].unique())
        metric_key = ';'.join([v['name'] for v in test_encyclopaedia['Information'][self.test_topic]['metrics'].values()
                      if v['name'] in output_result['metric'].unique()])
        characteristic_key = {test_encyclopaedia['Information'][self.test_topic]['characteristic'][c]['name']:
                                  test_encyclopaedia['Information'][self.test_topic]['characteristic'][c]['description']
                              for c in self.test_config['test_action']['output_characteristic']}

        with open(self.get_abspath(self.test_result['OutputResult']['scenario_analysis']), 'r', encoding='utf-8') as file:
            scenario_analysis = json.load(file)

        target_num, ru_num = {}, {}
        for scenario_tag in scenario_analysis:
            for scenario_id in scenario_analysis[scenario_tag]:
                for target_type in scenario_analysis[scenario_tag][scenario_id]['is_coverageValid']:
                    if target_type in self.test_information['type']:
                        if target_type in ['cyclist', 'pedestrian']:
                            ru = 'VRU'
                        else:
                            ru = 'DRU'
                        if ru not in ru_num:
                            ru_num[ru] = 0
                        ru_num[ru] += scenario_analysis[scenario_tag][scenario_id]['is_coverageValid'][target_type]['all_region']

                        target_type_text = self.test_information['type'][target_type]['name']
                        if target_type_text not in target_num:
                            target_num[target_type_text] = 0
                        target_num[target_type_text] += scenario_analysis[scenario_tag][scenario_id]['is_coverageValid'][target_type]['all_region']

        scenario_list = {}
        total_frame = 0
        total_distance = 0
        for i, scenario_tag in enumerate(self.test_result['TagCombination']):
            scenario_list[scenario_tag] = {}
            scenario_ids = self.test_result['TagCombination'][scenario_tag]['scenario_id']

            frame, distance = 0, 0
            for scenario_id in scenario_ids:
                video_info_path = os.path.join(self.scenario_unit_folder, scenario_id, '00_ScenarioInfo', 'video_info.yaml')
                with open(video_info_path, 'r', encoding='utf-8') as file:
                    video_info = yaml.safe_load(file)
                frame += round(video_info['duration'] * video_info['fps'] * 291 / 300)
                ego_vx_path = os.path.join(self.scenario_unit_folder, scenario_id, '01_Data', 'General', 'pred_ego.csv')
                ego_vx = pd.read_csv(ego_vx_path, index_col=False)
                distance += np.ceil((ego_vx['time_stamp'].max() - ego_vx['time_stamp'].min()) * ego_vx['ego_vx'].mean() / 1000)

            total_frame += frame
            total_distance += distance
            scenario_list[scenario_tag] = f'帧数-{frame:.0f} frame, 里程-{distance:.0f} km     ------- Page {3 + 12 * i}'

        test_info = {
            '测试版本': current_version,
            '版本对比': f'{current_version} <--> {old_version}',
            # 'test_date': self.test_config['test_date'],
            '测试对象': ';'.join(self.test_config['test_item'].keys()),
            '目标类型': ';'.join([f'{k}-{v}' for k, v in target_num.items()]),
            '指标类型': metric_key,
            '区域划分': region_key,
            '特征类型': characteristic_key,
            '测试场景类型': scenario_list,
        }
        plot_path = os.path.join(visualization_folder, '测试信息汇总.png')
        send_log(self, f'{plot_path} 图片已保存')
        self.test_result['OutputResult']['test_info'] = self.get_relpath(plot_path)
        plot_test_info(test_info, plot_path)

        title_summary = {
            '回灌帧数\n(千帧)': f'{(total_frame / 1000):.0f}',
            '场景里程\n(千米)': f'{total_distance:.0f}',
            '场景类型\n(类)': len(self.test_result['TagCombination'].keys()),
        }
        for ru, num in ru_num.items():
            title_summary[f'{ru}\n(千)'] = f'{(num / 1000):.0f}'
        plot_path = os.path.join(visualization_folder, '测试报告首页.png')
        plot_title_table(title_summary, plot_path)
        self.test_result['OutputResult']['test_title'] = self.get_relpath(plot_path)
        # ====================

        # 将速度图和地图都移动过来
        for scenario_tag in self.test_result['TagCombination']:
            scenario_ids = self.test_result['TagCombination'][scenario_tag]['scenario_id']
            for scenario_id in scenario_ids:
                for root, dirs, files in os.walk(os.path.join(self.scenario_unit_folder, scenario_id)):
                    for file in files:
                        if '_Info.jpg' in file:
                            file_path = os.path.join(root, file)
                            os.makedirs(os.path.join(visualization_folder, scenario_tag), exist_ok=True)
                            new_file_path = os.path.join(visualization_folder, scenario_tag, f'{scenario_id}.jpg')
                            shutil.copy(file_path, new_file_path)
        # ====================

        # 按照场景和目标特征，生成各个指标的场景
        statistics = pd.read_csv(self.get_abspath(self.test_result['OutputResult']['statistics']), index_col=False)
        group_columns = ['scenario_tag', 'topic', 'obstacle_type', 'characteristic', 'metric']
        stat_group = statistics.groupby(group_columns)
        self.test_result['OutputResult']['visualization'] = {}

        df_tp_error = {}
        for df_name, df in stat_group:
            scenario_tag, topic, obstacle_type, characteristic, metric = df_name
            if characteristic not in self.test_result['OutputResult']['report_table'].keys():
                continue

            if characteristic == '静止目标' and '速度' in metric:
                continue

            characteristic_folder = os.path.join(visualization_folder, scenario_tag, characteristic)
            if not os.path.exists(characteristic_folder):
                os.makedirs(characteristic_folder)

            for index, row in df.iterrows():
                result_dict = json.loads(row['result'])
                for key, value in result_dict.items():
                    df.at[index, key] = value
            df['type_cate'] = pd.Categorical(df['obstacle_type'],
                                             categories=['小车', '大车', '两轮车', '行人'], ordered=True)
            df.sort_values(by=['type_cate', 'region'], inplace=True)
            drop_columns = group_columns + ['index', 'type_cate', 'result']
            df.drop(drop_columns, axis=1, inplace=True)
            df.insert(0, 'obstacle_type', obstacle_type)
            df.rename(columns={'region': 'Test Area[m]'}, inplace=True)

            columns = list(df.columns)
            columns_width = [max(get_string_display_length(c) / 5.2, 2) for c in columns]
            values = df.values

            fig_size = (np.sum(columns_width) + 0.1, (len(df) + 2) + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.02, right=0.98, top=0.98, bottom=0.02)
            plt.axis('off')
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            plot_table(ax, columns, columns_width, values)
            ax.text(np.sum(columns_width) / 2, 1.5, '-'.join(
                [scenario_tag, topic, characteristic, obstacle_type, metric]),
                    va='center', ha='center', fontsize=font_size * 2)

            pic_name = '--'.join(
                [scenario_tag, topic, characteristic, obstacle_type, metric]).replace('/', '')
            pic_path = os.path.join(characteristic_folder, f'{pic_name}.jpg')
            send_log(self, f'{pic_name} 已保存')
            canvas = FigureCanvas(fig)
            canvas.print_figure(pic_path, facecolor='white', dpi=100)
            fig.clf()
            plt.close()

            if scenario_tag not in df_tp_error.keys():
                df_tp_error[scenario_tag] = {}
            if topic not in df_tp_error[scenario_tag].keys():
                df_tp_error[scenario_tag][topic] = {}
            if characteristic not in df_tp_error[scenario_tag][topic].keys():
                df_tp_error[scenario_tag][topic][characteristic] = {}
            if obstacle_type not in df_tp_error[scenario_tag][topic][characteristic].keys():
                df_tp_error[scenario_tag][topic][characteristic][obstacle_type] = []

            # 汇总准召率
            if '准召信息' in metric:
                df.rename(columns={
                    'recall%': '召回率',
                    'precision%': '准确率',
                    'type_accuracy%': '类型准确率'}, inplace=True)
                df_tp_error[scenario_tag][topic][characteristic][obstacle_type].append(
                    df[['Test Area[m]', '召回率', '准确率', '类型准确率']])

            # 汇总tp_error
            else:
                df.rename(columns={'pass_ratio%': metric}, inplace=True)
                df_tp_error[scenario_tag][topic][characteristic][obstacle_type].append(
                    df[['Test Area[m]', metric]])
        # ====================

        # 合并tp_error的metric数据，以及可视化为饼图
        for scenario_tag in df_tp_error.keys():
            if scenario_tag not in self.test_result['OutputResult']['visualization'].keys():
                self.test_result['OutputResult']['visualization'][scenario_tag] = {}

            for topic in df_tp_error[scenario_tag].keys():
                if topic not in self.test_result['OutputResult']['visualization'][scenario_tag].keys():
                    self.test_result['OutputResult']['visualization'][scenario_tag][topic] = {}

                for characteristic in df_tp_error[scenario_tag][topic].keys():
                    if characteristic not in self.test_result['OutputResult']['visualization'][scenario_tag][
                        topic].keys():
                        self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic] = {}

                    for obstacle_type in df_tp_error[scenario_tag][topic][characteristic].keys():
                        if obstacle_type not in self.test_result['OutputResult']['visualization'][scenario_tag][topic][
                            characteristic].keys():
                            self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic][
                                obstacle_type] = {}

                        # 合并表格
                        merged_df = pd.DataFrame(columns=['Test Area[m]'])
                        for one_df in df_tp_error[scenario_tag][topic][characteristic][obstacle_type]:
                            merged_df = merged_df.merge(one_df, on='Test Area[m]', how='outer')

                        # 重新排列结果的顺序
                        merged_df = merged_df.set_index('Test Area[m]', drop=True)
                        result_column = test_encyclopaedia['Information'][self.test_topic]['result_column']

                        merged_df = merged_df[[c for c in result_column if c in merged_df.columns]]

                        csv_name = '--'.join([scenario_tag, topic, characteristic, obstacle_type]).replace('/', '')
                        csv_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{csv_name}.csv')
                        (merged_df.sort_values(by=['Test Area[m]'])
                         .to_csv(csv_path, index=True, encoding='utf_8_sig'))
                        send_log(self, f'汇总百分比 {csv_name} 数据已保存')

                        self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic][
                            obstacle_type]['data'] = self.get_relpath(csv_path)

                        # 画结果大饼图
                        height_ratios = [1] + [3] * len(merged_df)  # 第一行高度是1/3，其余行高度是1
                        width_ratios = [2] + [3] * len(merged_df.columns)  # 第一列宽度是1/3，其余列宽度是1
                        fig = plt.figure(figsize=(sum(width_ratios), sum(height_ratios)))
                        fig.tight_layout()
                        plt.subplots_adjust(left=0.02, right=0.98, top=0.98, bottom=0.02)

                        # 创建GridSpec对象
                        grid = plt.GridSpec(len(height_ratios), len(width_ratios),
                                            height_ratios=height_ratios, width_ratios=width_ratios,
                                            wspace=0, hspace=0)

                        # 先画index和columns
                        for i, col in enumerate(merged_df.columns):
                            ax = fig.add_subplot(grid[0, i + 1])
                            ax.xaxis.set_visible(False)
                            ax.yaxis.set_visible(False)
                            for spine in ax.spines.values():
                                spine.set_edgecolor('black')
                                spine.set_linewidth(2)

                            ax.text(0.5, 0.5, col, va='center', ha='center',
                                    fontsize=font_size * 2.5)

                        for i, idx in enumerate(merged_df.index):
                            ax = fig.add_subplot(grid[i + 1, 0])
                            ax.xaxis.set_visible(False)
                            ax.yaxis.set_visible(False)
                            for spine in ax.spines.values():
                                spine.set_edgecolor('black')
                                spine.set_linewidth(0)

                            ax.text(0.5, 0.5, idx.replace(',', '\n'), va='center', ha='center',
                                    rotation=45, fontsize=font_size * 2.5)

                        # 画单独的饼图
                        cmap = LinearSegmentedColormap.from_list('red_to_green', ['darkred', 'darkgreen'])
                        value_shape = merged_df.values.shape
                        for i in range(value_shape[0]):
                            for j in range(value_shape[1]):
                                value = merged_df.values[i, j]
                                if not np.isnan(value):
                                    ax = fig.add_subplot(grid[i + 1, j + 1])
                                    ax.axis('off')
                                    sizes = [value, 1 - value]
                                    colors = ['limegreen', 'lightcoral']
                                    wedgeprops = {'width': 0.4}  # 设置扇区宽度为0.3，得到一个空心的效果
                                    ax.pie(sizes, colors=colors, startangle=90, wedgeprops=wedgeprops)
                                    ax.text(0, 0, f'{value:.2%}', va='center', ha='center',
                                            fontsize=font_size * 2, color=cmap(value))

                        pic_name = '--'.join(
                            [scenario_tag, topic, characteristic, obstacle_type]).replace('/', '')
                        pic_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{pic_name}.jpg')
                        send_log(self, f'{pic_name} 已保存')
                        canvas = FigureCanvas(fig)
                        canvas.print_figure(pic_path, facecolor='white', dpi=100)
                        fig.clf()
                        plt.close()
                        send_log(self, f'汇总百分比 {pic_name} 图片已保存')

                        self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic][
                            obstacle_type]['plot'] = self.get_relpath(pic_path)
        # ====================

        # 生成用于汇报的图
        def plot_single_metric(sub_data, old_data, plot_path):

            def plot_rect(left_bottom_pt, width, height,
                          facecolor, text, text_angle,
                          font_size, alpha=1, comparison=None):
                ratio = 1
                if (isinstance(comparison, float) or isinstance(comparison, int)) and len(text):
                    ratio = 0.75
                    if comparison < 0:
                        comparison_color = ['darkred', 'darkgreen'][metric_name != '准召信息']
                    else:
                        comparison_color = ['darkgreen', 'darkred'][metric_name != '准召信息']

                    if '%' in text:
                        comparison_text = f'{comparison:.2%}'
                    elif round(comparison, 0) == comparison:
                        comparison_text = f'{comparison:.0f}'
                    else:
                        comparison_text = f'{comparison:.2f}'

                    ax.add_patch(pc.Rectangle(
                        xy=(left_bottom_pt[0] + ratio * width, left_bottom_pt[1]),
                        width=width - ratio * width, height=height,
                        angle=0, alpha=alpha,
                        rotation_point='xy',
                        fill=True,
                        edgecolor='grey',
                        facecolor='white',
                        linewidth=1))
                    ax.text(left_bottom_pt[0] + (1 + ratio) * width / 2, left_bottom_pt[1] + height / 2,
                            comparison_text,
                            va='center', ha='center', rotation=270,
                            fontdict={
                                'family': 'sans-serif',  # 字体家族
                                'color': comparison_color,  # 字体颜色
                                'weight': 'bold',  # 字体粗细
                                'size': font_size * 0.7,  # 字体大小
                                'style': 'normal'  # 字体样式
                            })

                ax.add_patch(pc.Rectangle(
                    xy=left_bottom_pt,
                    width=width * ratio, height=height,
                    angle=0, alpha=alpha,
                    rotation_point='xy',
                    fill=True,
                    edgecolor='grey',
                    facecolor=facecolor,
                    linewidth=1))
                va, ha = 'center', 'center'
                ax.text(left_bottom_pt[0] + ratio * width / 2, left_bottom_pt[1] + height / 2, text,
                        va=va, ha=ha, rotation=text_angle,
                        fontsize=font_size)

            color_for_sheet = []
            cell_w = 1
            cell_h = 0.5
            font_size = 11
            fig_size = ((len(sub_data.columns) + 2) * cell_w + 0.1, (len(sub_data) + 3) * cell_h + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            ax.axis('off')

            metric_name = sub_data.columns[0][0]
            tag_name = sub_data.index[0][2]
            ax.text((len(sub_data.columns) - 2) / 2, cell_h * 2.5, f'<{tag_name}> {metric_name}',
                    va='center', ha='center', rotation=0,
                    fontsize=font_size * 2, color='peru')

            # 原点放置在数据的左上角
            # 画index
            plot_rect(left_bottom_pt=(-2 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='目标类型', text_angle=0, font_size=font_size)
            plot_rect(left_bottom_pt=(-1 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='测试范围[m]', text_angle=0, font_size=font_size)

            last_index_0 = sub_data.index[0][0]
            index_0_x, index_0_y = -2 * cell_w, 0
            index_width, index_height = cell_w, 0
            for i, index in enumerate(sub_data.index):
                if last_index_0 != index[0]:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)
                    index_height = 0
                index_height += cell_h
                index_0_y -= cell_h
                last_index_0 = index[0]
                if i == len(sub_data.index) - 1:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)

                plot_rect(left_bottom_pt=(-1 * cell_w, - (i + 1) * cell_h),
                          width=cell_w, height=cell_h, facecolor='bisque',
                          text=index[1].replace(',', '\n'), text_angle=0, font_size=font_size)

            # 画columns和data
            last_col_0 = sub_data.columns[0][1]
            col_0_x, col_0_y = 0, cell_h
            col_width, col_height = 0, cell_h
            for i, col in enumerate(sub_data.columns):
                if last_col_0 != col[1]:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)
                    col_0_x += col_width
                    col_width = 0
                col_width += cell_w
                last_col_0 = col[1]
                if i == len(sub_data.columns) - 1:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)

                if col[2] != 'kpi_target':
                    plot_rect(left_bottom_pt=(i * cell_w, 0),
                              width=cell_w, height=cell_h, facecolor='lightsteelblue',
                              text=wrap_text(col[2], 10), text_angle=0, font_size=font_size)
                    for j in range(len(sub_data.index)):
                        value = sub_data[col].tolist()[j]
                        if (col[0], col[1], 'kpi_target') in sub_data.columns:
                            kpi_target = sub_data[(col[0], col[1], 'kpi_target')].tolist()[j]
                            if np.isnan(value):
                                text = ''
                                color = 'lightgrey'
                            else:
                                if '%' in col[1]:
                                    text = f'{value:.2%}'
                                elif round(value, 0) == value:
                                    text = f'{value:.0f}'
                                else:
                                    text = f'{value:.2f}'
                                if np.isnan(kpi_target):
                                    color = 'white'
                                else:
                                    if value < kpi_target:
                                        color = ['lightcoral', 'limegreen'][metric_name != '准召信息']
                                    else:
                                        color = ['limegreen', 'lightcoral'][metric_name != '准召信息']
                        else:
                            if np.isnan(value):
                                text = ''
                                color = 'lightgrey'
                            else:
                                color = 'white'
                                if '%' in col[1]:
                                    text = f'{value:.2%}'
                                elif round(value, 0) == value:
                                    text = f'{value:.0f}'
                                else:
                                    text = f'{value:.2f}'

                        # 用于为excel表格涂色
                        color_for_sheet.append([list(col), j, color])

                        #取出old_data中相应格子的值
                        if old_data is None:
                            comparison = None
                        else:
                            if col not in old_data.columns:
                                comparison = None
                            else:
                                old_value = old_data[col].tolist()[j]
                                if np.isnan(old_value) or np.isnan(value):
                                    comparison = None
                                else:
                                    comparison = value - old_value

                        plot_rect(left_bottom_pt=(i * cell_w, - j * cell_h - cell_h),
                                  width=cell_w, height=cell_h, facecolor=color,
                                  text=text, text_angle=0, font_size=font_size * 1.2, comparison=comparison)

                else:
                    plot_rect(left_bottom_pt=(i * cell_w, 0),
                              width=cell_w, height=cell_h, facecolor='lightsteelblue',
                              text='指标期望', text_angle=0, font_size=font_size * 1.3)
                    for j in range(len(sub_data.index)):
                        value = sub_data.values[j, i]
                        color = 'lightcyan'
                        if np.isnan(value):
                            text = ''
                        elif '%' in col[1]:
                            text = f'{value:.2%}'
                        elif round(value, 0) == value:
                            text = f'{value:.0f}'
                        else:
                            text = f'{value:.2f}'
                        plot_rect(left_bottom_pt=(i * cell_w, - j * cell_h - cell_h),
                                  width=cell_w, height=cell_h, facecolor=color,
                                  text=text, text_angle=0, font_size=font_size * 1.2)

                        # 用于为excel表格涂色
                        color_for_sheet.append([list(col), j, color])

            ax.set_xlim(-2 * cell_w - 0.1, len(sub_data.columns) * cell_w + 0.1)
            ax.set_ylim(- len(sub_data) * cell_h - 0.1, 3 * cell_h + 0.1)
            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)

            return color_for_sheet

        def plot_summary_metric(summary_data, plot_path):

            def plot_rect(left_bottom_pt, width, height,
                          facecolor, text, text_angle,
                          font_size, fill=True):

                ax.add_patch(pc.Rectangle(
                    xy=left_bottom_pt,
                    width=width, height=height,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=fill,
                    edgecolor='grey',
                    facecolor=facecolor,
                    linewidth=1))
                va, ha = 'center', 'center'
                ax.text(left_bottom_pt[0] + width / 2, left_bottom_pt[1] + height / 2, text,
                        va=va, ha=ha, rotation=text_angle,
                        fontsize=font_size)

            cell_w = 1
            cell_h = 0.5
            font_size = 11
            fig_size = ((len(summary_data.columns) + 2) * cell_w + 0.1, (len(summary_data) + 3) * cell_h + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            ax.axis('off')

            tag_name = summary_data.index[0][2]
            ax.text((len(summary_data.columns) - 2) / 2, cell_h * 2.5, tag_name,
                    va='center', ha='center', rotation=0,
                    fontsize=font_size * 2, color='peru')

            # 原点放置在数据的左上角
            # 画index
            plot_rect(left_bottom_pt=(-2 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='目标类型', text_angle=0, font_size=font_size)
            plot_rect(left_bottom_pt=(-1 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='测试范围[m]', text_angle=0, font_size=font_size)

            last_index_0 = summary_data.index[0][0]
            index_0_x, index_0_y = -2 * cell_w, 0
            index_width, index_height = cell_w, 0
            for i, index in enumerate(summary_data.index):
                if last_index_0 != index[0]:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)
                    index_height = 0
                index_height += cell_h
                index_0_y -= cell_h
                last_index_0 = index[0]
                if i == len(summary_data.index) - 1:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)

                plot_rect(left_bottom_pt=(-1 * cell_w, - (i + 1) * cell_h),
                          width=cell_w, height=cell_h, facecolor='bisque',
                          text=index[1].replace(',', '\n'), text_angle=0, font_size=font_size)

            # 画columns和data
            last_col_0 = summary_data.columns[0][1]
            col_0_x, col_0_y = 0, cell_h
            col_width, col_height = 0, cell_h
            for i, col in enumerate(summary_data.columns):
                if last_col_0 != col[1]:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)
                    col_0_x += col_width
                    col_width = 0
                col_width += cell_w
                last_col_0 = col[1]
                if i == len(summary_data.columns) - 1:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)

                plot_rect(left_bottom_pt=(i * cell_w, 0),
                          width=cell_w, height=cell_h, facecolor='lightsteelblue',
                          text=wrap_text(col[2], 10), text_angle=0, font_size=font_size)

                good_face_image = Image.open(os.path.join(project_path, 'Docs', 'Resources', 'Icon', 'good_face.png'))
                bad_face_image = Image.open(os.path.join(project_path, 'Docs', 'Resources', 'Icon', 'bad_face.png'))

                for j in range(len(summary_data.index)):
                    plot_rect(left_bottom_pt=(i * cell_w, -j * cell_h - cell_h),
                              width=col_width, height=col_height, facecolor='white',
                              text='', text_angle=0, fill=False, font_size=font_size * 1.2)

                    value = summary_data.values[j, i]
                    extend = (i * cell_w + (cell_w - cell_h) / 2, i * cell_w + cell_w - (cell_w - cell_h) / 2,
                              -j * cell_h - cell_h, -j * cell_h)
                    if value == 1:
                        ax.imshow(good_face_image, extent=extend)
                    elif value == 0:
                        ax.imshow(bad_face_image, extent=extend)

            ax.set_xlim(-2 * cell_w - 0.1, len(summary_data.columns) * cell_w + 0.1)
            ax.set_ylim(- len(summary_data) * cell_h - 0.1, 3 * cell_h + 0.1)
            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)

        # 获取版本对比的文件夹
        version_comparison_folder = self.test_config['version_comparison']
        if version_comparison_folder is not None and os.path.exists(version_comparison_folder):
            comparison_valid = 1
        else:
            comparison_valid = 0

        metric_pass_ratio_by_characteristic_scenario = {} # 用于统计各特征和场景下的指标合格率
        color_for_sheets = {} # excel的涂色字典
        self.test_result['OutputResult']['report_plot'] = {}
        for characteristic, excel_path in self.test_result['OutputResult']['report_table'].items():
            excel_path = self.get_abspath(excel_path)
            self.test_result['OutputResult']['report_plot'][characteristic] = {}
            color_for_sheets[characteristic] = {}
            metric_pass_ratio_by_characteristic_scenario[characteristic] = {}

            # 寻找其他版本相同的文件
            old_excel_path = None
            if comparison_valid:
                for root, dirs, files in os.walk(version_comparison_folder):
                    for file in files:
                        if os.path.basename(excel_path) in file and self.test_topic in root:
                            old_excel_path = os.path.join(root, file)

            for _, scenario_tag in enumerate(pd.ExcelFile(excel_path).sheet_names):
                data = pd.read_excel(excel_path, sheet_name=scenario_tag, header=[0, 1, 2], index_col=[0, 1, 2])
                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag] = {}
                color_for_sheets[characteristic][scenario_tag] = []

                old_data = None
                if old_excel_path is not None and scenario_tag in pd.ExcelFile(excel_path).sheet_names:
                    old_data = pd.read_excel(old_excel_path, sheet_name=scenario_tag, header=[0, 1, 2], index_col=[0, 1, 2])

                sub_data_col = {}
                for col in data.columns:
                    if col[0] not in sub_data_col:
                        sub_data_col[col[0]] = []
                    sub_data_col[col[0]].append(col)

                # 获取汇总笑脸图，以及分指标kpi图
                pass_or_fail_summary = []
                for metric, cols in sub_data_col.items():

                    if metric == '准召信息':
                        new_cols = []
                        for col in cols:
                            if col[1] not in ['TP', 'CTP', 'FP', 'FN', 'sample_count']:
                                new_cols.append(col)
                        sub_data = data[new_cols]

                        # 将sample数量也单独绘制
                        new_cols = []
                        for col in cols:
                            if col[1] in ['TP', 'FP', 'FN', 'CTP']:
                                new_cols.append(col)
                        sample_data = data[new_cols]
                        sample_plot_path = os.path.join(visualization_folder, scenario_tag, characteristic,
                                                 f'{scenario_tag}--{characteristic}--样本数量.png')
                        plot_single_metric(sample_data, old_data, sample_plot_path)
                        self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag][
                            '样本数量'] = self.get_relpath(sample_plot_path)
                        send_log(self, f'{sample_plot_path} 图片已保存')

                        __metrics = list(dict.fromkeys(sub_data.columns.get_level_values(1)))
                        for __metric in __metrics:
                            temp_data = sub_data.loc[:,
                                        [(col_0, col_1, col_2) for col_0, col_1, col_2 in sub_data.columns if col_1 == __metric]]
                            pass_or_fails = []
                            for idx, row in temp_data.iterrows():
                                pass_or_fail = []
                                for t0, t1, topic in row.index:
                                    if topic == 'kpi_target':
                                        continue
                                    if np.isnan(row[(t0, t1, topic)]) or np.isnan(row[(t0, t1, 'kpi_target')]):
                                        pass_or_fail.append(-1)
                                    elif row[(t0, t1, topic)] > row[(t0, t1, 'kpi_target')]:
                                        pass_or_fail.append(1)
                                    else:
                                        pass_or_fail.append(0)
                                pass_or_fails.append(pass_or_fail)

                            new_data = sub_data.loc[:, [(col_0, col_1, col_2) for col_0, col_1, col_2 in temp_data.columns if
                                                        col_2 != 'kpi_target']]
                            pass_or_fail_data = pd.DataFrame(pass_or_fails, columns=new_data.columns, index=new_data.index)
                            pass_or_fail_summary.append(pass_or_fail_data)

                    else:
                        sub_data = data[cols]
                        __metrics = list(dict.fromkeys(sub_data.columns.get_level_values(1)))
                        pass_or_fails_compare = []
                        for __metric in __metrics:
                            temp_data = sub_data.loc[:,
                                        [(col_0, col_1, col_2) for col_0, col_1, col_2 in sub_data.columns if col_1 == __metric]]
                            pass_or_fails = []
                            for idx, row in temp_data.iterrows():
                                pass_or_fail = []
                                for t0, t1, topic in row.index:
                                    if topic == 'kpi_target':
                                        continue
                                    if np.isnan(row[(t0, t1, topic)]) or np.isnan(row[(t0, t1, 'kpi_target')]):
                                        pass_or_fail.append(-1)
                                    elif row[(t0, t1, topic)] > row[(t0, t1, 'kpi_target')]:
                                        pass_or_fail.append(0)
                                    else:
                                        pass_or_fail.append(1)
                                pass_or_fails.append(pass_or_fail)
                            pass_or_fails_compare.append(np.array(pass_or_fails))

                        # 与或合并测试结果
                        compare_res = np.zeros_like(pass_or_fails_compare[0])
                        for ii in range(compare_res.shape[0]):
                            for jj in range(compare_res.shape[1]):
                                res = [pass_or_fails[ii, jj] for pass_or_fails in pass_or_fails_compare]
                                if 1 in res:
                                    compare_res[ii, jj] = 1
                                elif 0 in res:
                                    compare_res[ii, jj] = 0
                                else:
                                    compare_res[ii, jj] = -1

                        pass_or_fail_columns = []
                        for col_0, col_1, col_2 in sub_data.columns:
                            if col_2 == 'kpi_target':
                                continue
                            if (col_0, col_0, col_2) not in pass_or_fail_columns:
                                pass_or_fail_columns.append((col_0, col_0, col_2))
                        pass_or_fail_data = pd.DataFrame(compare_res, columns=pass_or_fail_columns, index=sub_data.index)
                        pass_or_fail_summary.append(pass_or_fail_data)

                    plot_path = os.path.join(visualization_folder, scenario_tag, characteristic,
                                             f'{scenario_tag}--{characteristic}--{metric}.png')
                    self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag][
                        metric] = self.get_relpath(plot_path)
                    color_for_sheets[characteristic][scenario_tag].extend(plot_single_metric(sub_data, old_data, plot_path))
                    send_log(self, f'{plot_path} 图片已保存')

                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary'] = []
                plot_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{scenario_tag}--{characteristic}_summary_1.png')
                metric_summary = pd.concat(pass_or_fail_summary[:5], axis=1)
                plot_summary_metric(metric_summary, plot_path)
                send_log(self, f'{plot_path} 图片已保存')
                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary'].append(self.get_relpath(plot_path))

                plot_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{scenario_tag}--{characteristic}_summary_2.png')
                metric_summary = pd.concat(pass_or_fail_summary[5:], axis=1)
                plot_summary_metric(metric_summary, plot_path)
                send_log(self, f'{plot_path} 图片已保存')
                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary'].append(self.get_relpath(plot_path))

                metric_pass_count = np.sum(pd.concat(pass_or_fail_summary, axis=1).values == 1)
                metric_fail_count = np.sum(pd.concat(pass_or_fail_summary, axis=1).values == 0)
                metric_pass_ratio_by_characteristic_scenario[characteristic][scenario_tag] = {
                    'pass_count': int(metric_pass_count), 'fail_count': int(metric_fail_count),
                    'ratio': float(metric_pass_count / (metric_pass_count + metric_fail_count))}

        self.test_result['OutputResult']['metric_pass_ratio'] = metric_pass_ratio_by_characteristic_scenario
        # ====================

        # 根据比对结果，对excel源文件涂色
        with open(os.path.join(visualization_folder, 'color_for_sheets.yaml'), 'w', encoding='utf-8') as f:
            yaml.dump(color_for_sheets,
                      f, encoding='utf-8', allow_unicode=True, sort_keys=False)
        fill_type = {
            'lightcoral': PatternFill(fill_type="solid", start_color="FFB6C1"),
            'limegreen': PatternFill(fill_type="solid", start_color="32CD32"),
            'lightcyan': PatternFill(fill_type="solid", start_color="E0FFFF"),
            'lightgrey': PatternFill(fill_type="solid", start_color="D3D3D3"),
            'white': PatternFill(fill_type="solid", start_color="FFFFFF")
        }
        color_list = [
                         'CC4422', '33FF57', '2244CC', 'CC2288', '8822CC',
                         'CCAA00', '8B4513', 'D45A4A', '6600A3', 'A68C67',
                     ] * 20

        for characteristic, file in self.test_result['OutputResult']['report_table'].items():
            workbook = openpyxl.load_workbook(self.get_abspath(file))
            for scenario_tag in workbook.sheetnames:
                sheet = workbook[scenario_tag]
                sheet_colors = color_for_sheets[characteristic][scenario_tag]
                max_column = sheet.max_column

                column, i = 1, 0
                while column <= max_column:
                    if sheet.cell(row=1, column=column).value is not None:
                        i += 1
                    sheet.cell(row=1, column=column).fill = PatternFill(fill_type="solid", start_color=color_list[i])
                    column += 1

                for color_item in sheet_colors:
                    column = 1
                    while sheet.cell(row=1, column=column).value != color_item[0][0] and column <= max_column:
                        column += 1
                    while sheet.cell(row=2, column=column).value != color_item[0][1] and column <= max_column:
                        column += 1
                    while sheet.cell(row=3, column=column).value != color_item[0][2] and column <= max_column:
                        column += 1
                    if column <= max_column:
                        cell = sheet.cell(row=color_item[1] + 4, column=column)
                        cell.fill = fill_type[color_item[2]]
                        cell.border = Border(
                            left=Side(style='thin', color="000000"),
                            right=Side(style='thin', color="000000"),
                            top=Side(style='thin', color="000000"),
                            bottom=Side(style='thin', color="000000")
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            workbook.save(os.path.join(visualization_folder, f'{characteristic}.xlsx'))

    @sync_test_result
    def gen_report(self):

        report_title = f'{self.product}_{self.test_topic}_数据回灌感知测试报告'
        send_log(self, f'开始生成报告 {report_title}')

        title_background = os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'TitlePage.png')
        logo = os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'ZoneLogo.png')
        report_generator = PDFReportTemplate(report_title=report_title,
                                             test_time=self.test_date.split(' ')[0],
                                             tester='Hil_DataReplay_TestStand',
                                             version=self.version,
                                             title_page=title_background,
                                             title_summary_img=self.get_abspath(self.test_result['OutputResult']['test_title']),
                                             logo=logo)

        heading = '报告信息总览'
        send_log(self, f'生成页面 {heading}')
        report_generator.addOnePage(
            heading=heading,
            text_list=None,
            img_list=[[self.get_abspath(self.test_result['OutputResult']['test_info'])]],
        )

        for characteristic in self.test_result['OutputResult']['report_plot'].keys():

            if characteristic != '关键目标':
                continue

            for scenario_tag in self.test_result['OutputResult']['report_plot'][characteristic]:

                send_log(self, f'生成页面 {characteristic} {scenario_tag}')
                report_generator.addTitlePage(
                    title=f'{scenario_tag}<{characteristic}>',
                    page_type='sequence',
                    stamp=None,
                    sub_title_list=None,
                )

                for img in self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary']:

                    send_log(self, f'生成页面 {characteristic} {scenario_tag} Overview')
                    report_generator.addOnePage(
                        heading=f'{scenario_tag}<{characteristic}> Overview',
                        text_list=['绿色笑脸 = Pass, 红色哭脸 = Fail'],
                        img_list=[[self.get_abspath(img)]],
                    )

                for metric, img in self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag].items():
                    if metric == 'summary':
                        continue

                    send_log(self, f'生成页面 {characteristic} {scenario_tag} {metric}')
                    report_generator.addOnePage(
                        heading=f'{scenario_tag}<{characteristic}> {metric}',
                        text_list=['绿色代表“Pass”和“好于”，红色代表“Fail”和“差于”', '侧边竖写数字为与对比版本的差值'],
                        img_list=[[self.get_abspath(img)]],
                    )

        report_generator.addTitlePage(title='附 录',
                                      page_type='sequence',
                                      stamp=None,
                                      sub_title_list=None)

        for scenario_tag in self.test_result['OutputResult']['visualization'].keys():

            # 测试场景展示页面
            img_count = 0
            img_list = []
            for i, scenario_id in enumerate(self.test_result['TagCombination'][scenario_tag]['scenario_id']):
                scenario_info_img = os.path.join(self.output_result_folder, 'visualization', scenario_tag, f'{scenario_id}.jpg')
                img_list.append([scenario_info_img])
                img_count += 1
                if img_count == 3 or i == len(self.test_result['TagCombination'][scenario_tag]['scenario_id'])-1:
                    report_generator.addOnePage(heading=f'{scenario_tag} 测试场景',
                                                text_list=None,
                                                img_list=img_list)
                    img_list = []
                    img_count = 0

        # 测试环境
        heading = '测试环境——硬件在环 | 数采回灌'
        text = ['将真实控制器部署在仿真测试台架中，将实车路采的传感器数据, '
                '通过回灌测试设备回放及注入智驾控制器, 模拟运行场景和控制对象:',
                'A. 实车路采数据: 实车运行时采集的视频图片, 总线Can, Lidar点云等实时数据.',
                'B. 数据回灌设备: 数采数据后处理, 信号同步, 编码解码等, 模拟传感器输出, 将数据注入控制器.',
                'C. 智驾控制器: 接收模拟传感器信号, 模拟实车运行场景, 输出算法感知结果.',
                'D. ECU-Test，实现测试环境自动化运行及数据自动化分析，用于测试用例的执行和测试结果的生成.', ]
        img = [
            [os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'DataReplayTestEnv.png')]
        ]
        report_generator.addOnePage(heading=heading,
                                    text_list=text,
                                    img_list=img)

        # 测试原理
        heading = '测试原理——硬件在环 | 数采回灌'
        text = ['对比感知算法结果与参考真值, 计算性能指标, 输出评价结果:',
                'A. 以标注平台或具备更高感知能力的工具基于数采数据提供标注真值结果作为GroundTruth(GT).',
                'B. 控制器感知节点订阅信息提供场景各要素特征测量值，MeasuredValue(MV).',
                'C. 对比并评价感知信息是否符合预期，包括准召率，定位精度, 几何精度, 类型准确度等.',
                '--------------------',
                '测试结果局限性:',
                'A. 测试结果的置信度受限于真值结果对于真实世界的准确程度.',
                'B. 测试结果的全面性依赖于数采数据的样本规模, 需要时间的积累.']
        img = [
            [os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'DataReplayTestPrinciple.png')]
        ]
        report_generator.addOnePage(heading=heading,
                                    text_list=text,
                                    img_list=img)

        report_generator.page_count += 0
        self.report_path = os.path.join(self.task_folder, f'{self.product}_{self.test_topic}_数据回灌测试报告({self.test_config["version"]}).pdf')
        report_generator.genReport(report_path=self.report_path, compress=1)

    def start(self):
        if (self.test_action['scenario_unit']['preprocess']
                or self.test_action['scenario_unit']['match']
                or self.test_action['scenario_unit']['metric']
                or self.test_action['scenario_unit']['bug']
                or self.test_action['scenario_unit']['render']):
            self.analyze_scenario_unit()

        if self.test_action['tag_combination']:
            self.combine_scenario_tag()

        if self.test_action['statistics']:
            self.summary_bug_items()
            self.compile_statistics()

        if self.test_action['visualization']:
            self.visualize_output()

        if self.test_action['gen_report']:
            self.gen_report()


class DataGrinderPilotLinesOneCase(DataGrinderOneCase):

    def start(self):

        if self.test_action['preprocess']:
            self.load_pred_data()
            self.load_gt_data()
            self.sync_timestamp()
            self.promote_additional()

        if self.test_action['match']:
            self.match_timestamp()
            self.match_object()


class DataGrinderPilotLinesOneTask(DataGrinderOneTask):

    @sync_test_result
    def combine_scenario_tag(self):
        # 合并各个场景的match_data, 重新corresponding_index编号
        # 获得每一种特征的分类结果
        for scenario_tag in self.test_config['scenario_tag']:
            # 需要去除Broken Scenario
            valid_scenario_list = [scenario_id for scenario_id in scenario_tag['scenario_id']
                                   if scenario_id in self.valid_scenario_list]
            if len(valid_scenario_list):
                tag_key = '&'.join(scenario_tag['tag'].values())
                self.test_result['TagCombination'][tag_key] = {
                    'tag': scenario_tag['tag'],
                    'scenario_id': valid_scenario_list,
                    self.test_topic: {},
                }

        match_data_dict = {}
        for tag_key in self.test_result['TagCombination'].keys():

            # 汇总数据
            tag_combination_folder = os.path.join(self.tag_combination_folder, tag_key)
            create_folder(tag_combination_folder)
            match_data_dict[tag_key] = {}

            for scenario_id in self.test_result['TagCombination'][tag_key]['scenario_id']:
                scenario_unit_folder = os.path.join(self.scenario_unit_folder, scenario_id)
                with open(os.path.join(scenario_unit_folder, 'TestResult.yaml')) as f:
                    scenario_test_result = yaml.load(f, Loader=yaml.FullLoader)

                # 遍历全部测试结果，按照topic和metrics合并
                for topic in scenario_test_result[self.test_topic].keys():
                    if topic == 'GroundTruth':
                        continue

                    topic_tag = topic.replace('/', '')
                    # 用于设置评判样本量最小限度的帧数阈值
                    if topic not in self.test_result['TagCombination'][tag_key][self.test_topic]:
                        self.test_result['TagCombination'][tag_key][self.test_topic][topic] = {
                            'frequency': scenario_test_result[self.test_topic][topic]['match_frequency']
                        }
                        topic_folder = os.path.join(tag_combination_folder, topic_tag)
                        create_folder(topic_folder)

                    if topic not in match_data_dict[tag_key]:
                        match_data_dict[tag_key][topic] = []

                    print(scenario_id)
                    match_data_path = scenario_test_result[self.test_topic][topic]['match']['match_data']
                    match_data = pd.read_csv(os.path.join(scenario_unit_folder, match_data_path), index_col=False)
                    match_data.insert(0, 'scenario_id', scenario_id)
                    match_data_dict[tag_key][topic].append(match_data)

        for tag_key in match_data_dict.keys():

            send_log(self, f'{self.test_topic}, 使用{self.test_topic}MetricEvaluator')

            for topic, df_list in match_data_dict[tag_key].items():
                topic_tag = topic.replace('/', '')
                metric_folder = os.path.join(self.tag_combination_folder, tag_key, topic_tag)
                create_folder(metric_folder)

                total_match_data = pd.concat(df_list).reset_index(drop=True)
                total_match_data['corresponding_index'] = total_match_data.index
                total_match_data_path = os.path.join(get_project_path(), 'Temp', 'total_match_data.csv')
                total_match_data.to_csv(total_match_data_path, index=False, encoding='utf_8_sig')

                input_parameter_container = {
                    'metric_type': self.test_config['test_item'][topic],
                    'characteristic_type': self.test_config['target_characteristic'],
                    'kpi_date_label': self.kpi_date_label,
                }
                parameter_json_path = os.path.join(get_project_path(), 'Temp', 'evaluate_api_parameter.json')
                with open(parameter_json_path, 'w', encoding='utf-8') as f:
                    json.dump(input_parameter_container, f, ensure_ascii=False, indent=4)

                cmd = [
                    f"{bench_config['master']['sys_interpreter']}",
                    "Api_EvaluateMetrics.py",
                    "-m", total_match_data_path,
                    "-j", parameter_json_path,
                    "-f", metric_folder,
                ]

                print(cmd)
                send_log(self, f'{self.test_topic} {topic} 指标评估')
                cwd = os.path.join(get_project_path(), 'Envs', 'Master', 'Interfaces')
                result = subprocess.run(cmd, cwd=cwd, capture_output=True, text=True)
                # os.remove(parameter_json_path)
                os.remove(total_match_data_path)
                if result.stderr and 'warning' not in result.stderr:
                    print(result.stderr)
                    send_log(self, f'EvaluateMetrics 发生错误 {result.stderr}')

                for characteristic in os.listdir(metric_folder):
                    self.test_result['TagCombination'][tag_key][self.test_topic][topic][characteristic] = {}
                    for metric in os.listdir(os.path.join(metric_folder, characteristic)):
                        self.test_result['TagCombination'][tag_key][self.test_topic][topic][characteristic][
                            metric.split('.')[0]] = self.get_relpath(
                            os.path.join(metric_folder, characteristic, metric))

    @sync_test_result
    def summary_bug_items(self):

        bugItem_folder = os.path.join(self.output_result_folder, 'bugItems')
        create_folder(bugItem_folder)
        dataframes = []

        for scenario_tag in self.test_result['TagCombination']:
            scenario_ids = self.test_result['TagCombination'][scenario_tag]['scenario_id']
            for scenario_id in scenario_ids:
                for root, dirs, files in os.walk(os.path.join(self.scenario_unit_folder, scenario_id)):
                    for file in files:
                        if 'bug_jira_summary.csv' in file:
                            file_path = os.path.join(root, file)
                            df = pd.read_csv(file_path, index_col=False)
                            if len(df):
                                dataframes.append(df)

        if len(dataframes):
            bug_summary = pd.concat(dataframes, ignore_index=True)
            bug_report_path_list = []
            for i, row in bug_summary.iterrows():
                topic = row['topic'].replace('/', '')
                bug_type = row['bug_type']
                target_type = row['target_type']
                folder = os.path.join(bugItem_folder, bug_type, topic, target_type)
                if not os.path.exists(folder):
                    os.makedirs(folder)

                bug_report_path = row['attachment_path']
                shutil.copy(bug_report_path, folder)
                bug_report_path_list.append(os.path.join(folder, os.path.basename(bug_report_path)))

            bug_summary['attachment_path'] = bug_report_path_list
            bug_summary_path = os.path.join(bugItem_folder, 'bug_summary.csv')
            (bug_summary.sort_values(by=['bug_type', 'scenario_id', 'gt_target_id'])
             .to_csv(bug_summary_path, index=False, encoding='utf_8_sig'))
            self.test_result['OutputResult']['bugItems'] = self.get_relpath(bug_summary_path)

        # 获取版本对比的文件夹
        version_comparison_folder = self.test_config['version_comparison']
        if version_comparison_folder is not None and os.path.exists(version_comparison_folder):
            for root, dirs, files in os.walk(version_comparison_folder):
                for file in files:
                    if 'bug_summary.csv' in file and self.test_topic in root:
                        print(os.path.join(root, file))

        # 1. 上个版本没有，这个版本出现

    @sync_test_result
    def compile_statistics(self):
        # 按照数据库数据单元的方式保存数据
        # 格式为json，在文件夹内平铺
        statistics_folder = os.path.join(self.output_result_folder, 'statistics')
        create_folder(statistics_folder)

        # 统计目标个数
        total_scenario_analysis_path = os.path.join(project_path, 'Docs', 'Resources', 'scenario_info', 'scenario_analysis.json')
        with open(total_scenario_analysis_path, 'r', encoding='utf-8') as f:
            total_scenario_analysis = json.load(f)

        scenario_analysis = {}
        for tag_key in self.test_result['TagCombination'].keys():
            scenario_list = self.test_result['TagCombination'][tag_key]['scenario_id']
            scenario_analysis[tag_key] = {}

            for scenario_id in scenario_list:
                if f'{scenario_id}-{self.truth_source}' not in total_scenario_analysis.keys():
                    continue

                scenario_analysis[tag_key][scenario_id] = {}
                for characteristic in self.test_config['test_action']['output_characteristic']:
                    scenario_analysis[tag_key][scenario_id][characteristic] = (
                        total_scenario_analysis)[f'{scenario_id}-{self.truth_source}']['obstacles'][characteristic]

        scenario_analysis_path = os.path.join(statistics_folder, 'scenario_analysis.json')
        with open(scenario_analysis_path, 'w', encoding='utf-8') as f:
            json.dump(scenario_analysis, f, ensure_ascii=False, indent=4)

        self.test_result['OutputResult']['scenario_analysis'] = self.get_relpath(scenario_analysis_path)
        # ====================

        # 生成用于上传数据库的json文件，以及汇总结果OutputResult
        json_count = 0
        json_rows = []
        for tag_key in self.test_result['TagCombination'].keys():
            tag = self.test_result['TagCombination'][tag_key]['tag']
            scenario_list = self.test_result['TagCombination'][tag_key]['scenario_id']

            metric_statistics = eval(f'MetricStatistics.{self.test_topic}MetricStatistics()')

            for topic in self.test_result['TagCombination'][tag_key][self.test_topic].keys():
                topic_tag = topic.replace('/', '')

                for characteristic in self.test_config['target_characteristic']:
                    if characteristic not in self.test_result['TagCombination'][tag_key][self.test_topic][topic]:
                        continue

                    if characteristic not in self.test_config['test_action']['output_characteristic']:
                        continue

                    test_result = self.test_result['TagCombination'][tag_key][self.test_topic][topic][
                        characteristic]

                    info_json_data = {
                        'Product': self.product,
                        'Version': self.version,
                        'StartTime': str(self.test_date),
                        **{tag_type: tag_value for tag_type, tag_value in tag.items()},
                        'ScenarioGroup': scenario_list,
                        'TopicName': topic}

                    data = {
                        metric: pd.read_csv(self.get_abspath(data_path), index_col=False)
                        for metric, data_path in test_result.items()
                    }

                    input_parameter_container = {
                        'region_division': self.test_config['region_division'],
                        'characteristic': characteristic,
                    }

                    json_datas = metric_statistics.run(data, input_parameter_container)

                    for json_data in json_datas:
                        json_data = {**info_json_data, **json_data}

                        # 在生成测试结果表格的时候，做一下筛选:
                        # 1.速度仅在后处理存在
                        # 2.行人不存在航向角
                        # 3.快速和高速场景不存在行人和两轮车
                        # 4.行人和两轮车没有那么远的测试距离

                        if '速度' in json_data['MetricTypeName'] and json_data['TopicName'] != '/VA/Obstacles':
                            continue

                        if json_data['MetricTypeName'] == '航向角误差' and json_data['ObstacleName'] == '行人':
                            continue

                        if (('快速' in json_data['RoadTypeCondition'] or '高速' in json_data['RoadTypeCondition'])
                                and (json_data['ObstacleName'] in ['行人', '两轮车'])):
                            continue

                        if ('100~150' in json_data['RangeDetails'] and json_data['ObstacleName'] in ['行人', '两轮车']) \
                                or ('-100~-50' in json_data['RangeDetails'] and json_data['ObstacleName'] == '行人'):
                            continue

                        if json_data['TopicName'] == '/VA/PedResult' and json_data['ObstacleName'] in ['小车', '大车']:
                            continue

                        if json_data['TopicName'] == '/VA/VehicleResult' and json_data['ObstacleName'] in ['两轮车', '行人']:
                            continue

                        # 保存单个json文件
                        json_count += 1
                        json_name = (f'{json_count:06d}--{self.version}--{tag_key}--{topic_tag}--{json_data["ObstacleName"]}'
                                     f'--{json_data["RangeDetails"]}--{json_data["FeatureDetail"]}--{json_data["MetricTypeName"]}.json')
                        json_path = os.path.join(statistics_folder, json_name)

                        frequency_threshold = self.test_result['TagCombination'][tag_key][self.test_topic][topic][
                                                  'frequency'] * 2
                        if json_data['Output']['sample_count'] < frequency_threshold:
                            send_log(self, f'{json_count} {json_name} 样本少于{frequency_threshold}，不保存')
                            continue

                        json_data['ScenarioGroup'] = '&'.join(json_data['ScenarioGroup'])
                        json_data['uuid'] = str(uuid.uuid4())
                        send_log(self, f'{json_count} {json_name} 已保存')
                        with open(json_path, 'w', encoding='utf-8') as f:
                            json.dump(json_data, f, ensure_ascii=False, indent=4)

                        # 保存json文件的目录
                        json_rows.append(
                            [
                                json_count, self.version, tag_key, topic,
                                json_data['ObstacleName'], json_data['RangeDetails'],
                                json_data['FeatureDetail'], json_data['MetricTypeName'],
                                json_data['Output']['sample_count'], json.dumps(json_data['Output'])
                            ]
                        )

        path = os.path.join(statistics_folder, 'output_result.csv')
        output_result = pd.DataFrame(json_rows, columns=[
            'index', 'version', 'scenario_tag', 'topic', 'obstacle_type', 'region',
            'characteristic', 'metric', 'sample_count', 'result'])
        output_result.to_csv(path, index=False, encoding='utf_8_sig')
        self.test_result['OutputResult']['statistics'] = self.get_relpath(path)
        # ====================

        # 生成以场景类型作为sheet的xlsx
        characteristic_key = [test_encyclopaedia['Information'][self.test_topic]['characteristic'][c]['name']
                              for c in self.test_config['test_action']['output_characteristic']]
        topic_key = output_result['topic'].unique()
        scenario_tag_key = output_result['scenario_tag'].unique()
        obstacle_type_key = output_result['obstacle_type'].unique()
        region_key = output_result['region'].unique()
        metric_key = [v['name'] for v in test_encyclopaedia['Information'][self.test_topic]['metrics'].values()
                      if v['name'] in output_result['metric'].unique()]
        # 首先生成columns
        columns = [
            ('基本信息', '基本信息', 'OD类型'),
            ('基本信息', '基本信息', '目标物范围'),
            ('基本信息', '基本信息', 'start_distance'),
            ('基本信息', '基本信息', '场景类型'),
        ]
        for metric in metric_key:
            temp_data = output_result[output_result['metric'] == metric]
            for res_key in json.loads(temp_data.iloc[0]['result']).keys():
                if (
                        (metric == '准召信息') or
                        (metric in ['长度误差', '宽度误差', '高度误差'] and 'abs_95' in res_key) or
                        ('距离误差' in metric and 'abs_95' in res_key) or
                        ('abs_95' in res_key and '%' not in res_key)
                ):
                    if res_key not in ['TP', 'FP', 'FN', 'CTP', 'sample_count']:
                        columns.append(
                            (metric, res_key, 'kpi_target')
                        )
                    for topic in topic_key:
                        columns.append(
                            (metric, res_key, topic)
                        )

        self.test_result['OutputResult']['report_table'] = {}
        for i, characteristic in enumerate(characteristic_key):
            path = os.path.join(statistics_folder, f'{characteristic}-指标汇总.xlsx')
            self.test_result['OutputResult']['report_table'][characteristic] = self.get_relpath(path)

            # 使用ExcelWriter将DataFrame保存到不同的sheet中
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                for scenario_tag in scenario_tag_key:
                    rows = []
                    for obstacle_type in obstacle_type_key:
                        if '快速' in scenario_tag and obstacle_type in ['行人', '两轮车']:
                            continue

                        for region in region_key:
                            if ('100~150' in region and obstacle_type in ['行人', '两轮车']) \
                                or ('-100~-50' in region and obstacle_type == '行人'):
                                continue

                            start_distance = int(region.split('~')[0].split('(')[1])
                            row = [obstacle_type, region, start_distance, scenario_tag]
                            filter_data = output_result[
                                (output_result['characteristic'] == characteristic)
                                & (output_result['scenario_tag'] == scenario_tag)
                                & (output_result['obstacle_type'] == obstacle_type)
                                & (output_result['region'] == region)
                            ]

                            for metric, res_key, topic in columns[4:]:
                                if topic == 'kpi_target':
                                    kpi_threshold = get_obstacles_kpi_threshold(metric, res_key, obstacle_type, region=region)
                                    if kpi_threshold is None:
                                        row.append(np.nan)
                                    else:
                                        kpi_ratio = get_obstacles_kpi_ratio(metric, res_key, obstacle_type, self.kpi_date_label)
                                        row.append(kpi_threshold * kpi_ratio)

                                else:
                                    filter2_data = filter_data[
                                        (filter_data['metric'] == metric)
                                        & (filter_data['topic'] == topic)
                                    ]

                                    if len(filter2_data):
                                        c = json.loads(filter2_data.iloc[0]['result'])
                                        row.append(c[res_key])
                                    else:
                                        row.append(np.nan)

                            rows.append(row)

                    df = pd.DataFrame(rows, columns=pd.MultiIndex.from_tuples(columns))
                    df.sort_values(by=[('基本信息', '基本信息', 'OD类型'), ('基本信息', '基本信息', 'start_distance')], inplace=True)
                    df.drop(('基本信息', '基本信息', 'start_distance'), axis=1, inplace=True)
                    df.to_excel(writer, sheet_name=scenario_tag, merge_cells=True)

            workbook = Workbook()
            workbook.LoadFromFile(path)
            for sheet in workbook.Worksheets:
                sheet.DeleteRow(4)
                sheet.DeleteColumn(1)
            workbook.SaveToFile(path)
            workbook.Dispose()
            sheet_name1 = 'Evaluation Warning'
            workbook = openpyxl.load_workbook(path)
            worksheet = workbook[sheet_name1]
            workbook.remove(worksheet)
            workbook.save(path)
        # ====================

    @sync_test_result
    def visualize_output(self):

        def plot_table(axes, columns, columns_width, values, index=None, index_width=0):
            if index is None:
                index = []
            axes.patch.set_facecolor('white')
            axes.patch.set_alpha(0.3)
            for pos in ['right', 'top', 'left', 'bottom']:
                axes.spines[pos].set_visible(False)

            # 画index
            if index_width and len(index):
                axes.add_patch(pc.Rectangle(
                    xy=(-index_width, 0),
                    width=index_width, height=1,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=True,
                    edgecolor='grey',
                    facecolor='bisque',
                    linewidth=1))
                axes.text(-index_width / 2, 0.5, 'Test Case',
                          va='center', ha='center',
                          fontdict={
                              'family': 'Ubuntu',
                              'style': 'normal',
                              'weight': 'bold',
                              'color': 'black',
                              'size': font_size * 1.5,
                          })

                for i, idx in enumerate(index):
                    axes.add_patch(pc.Rectangle(
                        xy=(-index_width, -1 - i),
                        width=index_width, height=1,
                        angle=0, alpha=1,
                        rotation_point='xy',
                        fill=True,
                        edgecolor='grey',
                        facecolor='linen',
                        linewidth=1))

                    axes.text(-index_width * 0.95, -1 - i + 0.5, idx,
                              va='center', ha='left',
                              fontdict={
                                  'family': 'sans-serif',
                                  'style': 'normal',
                                  'weight': 'normal',
                                  'color': 'black',
                                  'size': font_size * 1.5,
                              })

            # 画columns
            current_x = 0
            for i, (col, width) in enumerate(zip(columns, columns_width)):
                axes.add_patch(pc.Rectangle(
                    xy=(current_x, 0),
                    width=width, height=1,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=True,
                    edgecolor='grey',
                    facecolor='lightsteelblue',
                    linewidth=1))

                axes.text(current_x + width * 0.5, 0.5, col,
                          va='center', ha='center',
                          fontdict={
                              'family': 'Ubuntu',
                              'style': 'normal',
                              'weight': 'bold',
                              'color': 'black',
                              'size': font_size * 1.5,
                          })

                current_x += width

            # 画values
            ii, jj = values.shape
            for i in range(ii):
                current_x = 0
                for j in range(jj):
                    width = columns_width[j]
                    color = 'white'
                    value = values[i, j]

                    if '%' in columns[j]:
                        value = '{:.2%}'.format(value)

                    elif isinstance(value, float) and (not np.isnan(value)):
                        if int(value) == value:
                            value = int(value)
                        else:
                            value = round(value, 4)

                    if contains_chinese(str(value)):
                        fontdict = {
                            'family': 'sans-serif',
                            'style': 'normal',
                            'weight': 'normal',
                            'color': 'black',
                            'size': font_size * 1.5,
                        }
                    else:
                        fontdict = {
                            'family': 'Ubuntu',
                            'style': 'normal',
                            'weight': 'normal',
                            'color': 'black',
                            'size': font_size * 1.5,
                        }

                    axes.add_patch(pc.Rectangle(
                        xy=(current_x, -i - 1),
                        width=width, height=1,
                        angle=0, alpha=1,
                        rotation_point='xy',
                        fill=True,
                        edgecolor='grey',
                        facecolor=color,
                        linewidth=1))
                    axes.text(current_x + width * 0.5, -i - 1 + 0.5, value,
                              va='center', ha='center',
                              fontdict=fontdict)

                    current_x += width

            axes.set_xlim(-index_width - 0.1, np.sum(columns_width) + 0.1)
            axes.set_ylim(-len(df) - 0.1, 2.1)
            axes.set_xticks([])
            axes.set_yticks([])

        def wrap_text(text, max_length):
            wrapped_lines = []
            for i in range(0, len(text), max_length):
                wrapped_lines.append(text[i:i + max_length])
            return '\n'.join(wrapped_lines)

        visualization_folder = os.path.join(self.output_result_folder, 'visualization')
        create_folder(visualization_folder)

        # 生成测试信息总览
        def plot_test_info(test_info, plot_path):

            def plot_rect(left_bottom_pt, width, height,
                          facecolor, text, text_angle,
                          font_size, fill=True):
                ax.add_patch(pc.Rectangle(
                    xy=left_bottom_pt,
                    width=width, height=height,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=fill,
                    edgecolor='grey',
                    facecolor=facecolor,
                    linewidth=1))
                va, ha = 'center', 'center'
                ax.text(left_bottom_pt[0] + width / 2, left_bottom_pt[1] + height / 2, text,
                        va=va, ha=ha, rotation=text_angle,
                        fontsize=font_size)

            cell_w = 2
            cell_h = 0.6
            font_size = 13

            fig_height = 0
            for i, (key, info) in enumerate(test_info.items()):
                if not isinstance(info, dict):
                    fig_height += cell_h
                else:
                    fig_height += cell_h * len(info.keys()) * 0.9
            fig_width = cell_w * 6

            fig_size = (fig_width + 0.1, fig_height + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            ax.axis('off')

            left_bottom_y = 0
            for i, (key, info) in enumerate(test_info.items()):
                if key == '测试版本':
                    left_bottom_y -= cell_h
                    plot_rect(left_bottom_pt=[0, left_bottom_y],
                              width=fig_width, height=cell_h, facecolor='lightsteelblue',
                              text=f'<{info}>测试报告信息总览', text_angle=0, font_size=font_size * 1.2)
                elif not isinstance(info, dict):
                    left_bottom_y -= cell_h
                    plot_rect(left_bottom_pt=[cell_w, left_bottom_y],
                              width=cell_w * 5, height=cell_h, facecolor='white',
                              text=info, text_angle=0, font_size=font_size)
                    plot_rect(left_bottom_pt=[0, left_bottom_y],
                              width=cell_w, height=cell_h, facecolor='lightsteelblue',
                              text=key, text_angle=0, font_size=font_size)
                else:
                    left_bottom_y -= cell_h * len(info.keys()) * 0.9
                    text = '\n\n'.join([f'{k}:{v}' for k, v in info.items()])
                    plot_rect(left_bottom_pt=[cell_w, left_bottom_y],
                              width=cell_w * 5, height=cell_h * len(info.keys()) * 0.9, facecolor='white',
                              text=text, text_angle=0, font_size=font_size)
                    plot_rect(left_bottom_pt=[0, left_bottom_y],
                              width=cell_w, height=cell_h * len(info.keys()) * 0.9, facecolor='lightsteelblue',
                              text=key, text_angle=0, font_size=font_size)

            ax.set_xlim(-0.1, fig_width + 0.1)
            ax.set_ylim(-fig_height - 0.1, 0.1)
            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)

        def plot_title_table(title_dict, plot_path):
            fig = plt.figure(figsize=(4 * len(title_dict), 3.5))
            fig.tight_layout()
            plt.subplots_adjust(left=0.05, right=0.95, top=0.95, bottom=0.05)
            grid = plt.GridSpec(1, 1, wspace=0.05, hspace=0.05)
            ax_title_page = fig.add_subplot(grid[0, 0])
            for pos in ['right', 'top', 'left', 'bottom']:
                ax_title_page.spines[pos].set_visible(False)

            center_x = 2
            for k, v in title_dict.items():
                ax_title_page.text(center_x, 0.875, k,
                                   va='center', ha='center',
                                   fontdict={
                                       'style': 'normal',
                                       'weight': 'normal',
                                       'color': 'black',
                                       'size': font_size * 2.5,
                                   })
                ax_title_page.text(center_x, -0.875, str(v),
                                   va='center', ha='center',
                                   fontdict={
                                       'family': 'Ubuntu',
                                       'style': 'normal',
                                       'weight': 'bold',
                                       'color': 'orange',
                                       'size': font_size * 6,
                                   })
                center_x += 4

            ax_title_page.set_xlim(0, len(title_dict) * 4)
            ax_title_page.set_ylim(-1.75, 1.75)
            ax_title_page.set_xticks([])
            ax_title_page.set_yticks([])

            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)
            fig.clf()
            plt.close()

        current_version = self.test_config['version']
        test_topic = self.test_config['test_topic']
        version_comparison_folder = self.test_config['version_comparison']
        old_version = ''
        if version_comparison_folder is not None and os.path.exists(version_comparison_folder):
            for root, dirs, files in os.walk(version_comparison_folder):
                for file in files:
                    if 'TestConfig.yaml' in file:
                        file_path = os.path.join(root, file)
                        with open(file_path, 'r', encoding='utf-8') as file:
                            old_test_config = yaml.safe_load(file)
                        if 'version_comparison' in old_test_config and old_test_config['test_topic'] == test_topic:
                            old_version = old_test_config['version']

        output_result = pd.read_csv(self.get_abspath(self.test_result['OutputResult']['statistics']), index_col=False)
        region_key = ';'.join(output_result['region'].unique())
        metric_key = ';'.join([v['name'] for v in test_encyclopaedia['Information'][self.test_topic]['metrics'].values()
                      if v['name'] in output_result['metric'].unique()])
        characteristic_key = {test_encyclopaedia['Information'][self.test_topic]['characteristic'][c]['name']:
                                  test_encyclopaedia['Information'][self.test_topic]['characteristic'][c]['description']
                              for c in self.test_config['test_action']['output_characteristic']}

        with open(self.get_abspath(self.test_result['OutputResult']['scenario_analysis']), 'r', encoding='utf-8') as file:
            scenario_analysis = json.load(file)

        target_num, ru_num = {}, {}
        for scenario_tag in scenario_analysis:
            for scenario_id in scenario_analysis[scenario_tag]:
                for target_type in scenario_analysis[scenario_tag][scenario_id]['is_coverageValid']:
                    if target_type in self.test_information['type']:
                        if target_type in ['cyclist', 'pedestrian']:
                            ru = 'VRU'
                        else:
                            ru = 'DRU'
                        if ru not in ru_num:
                            ru_num[ru] = 0
                        ru_num[ru] += scenario_analysis[scenario_tag][scenario_id]['is_coverageValid'][target_type]['all_region']

                        target_type_text = self.test_information['type'][target_type]['name']
                        if target_type_text not in target_num:
                            target_num[target_type_text] = 0
                        target_num[target_type_text] += scenario_analysis[scenario_tag][scenario_id]['is_coverageValid'][target_type]['all_region']

        scenario_list = {}
        total_frame = 0
        total_distance = 0
        for i, scenario_tag in enumerate(self.test_result['TagCombination']):
            scenario_list[scenario_tag] = {}
            scenario_ids = self.test_result['TagCombination'][scenario_tag]['scenario_id']

            frame, distance = 0, 0
            for scenario_id in scenario_ids:
                video_info_path = os.path.join(self.scenario_unit_folder, scenario_id, '00_ScenarioInfo', 'video_info.yaml')
                with open(video_info_path, 'r', encoding='utf-8') as file:
                    video_info = yaml.safe_load(file)
                frame += round(video_info['duration'] * video_info['fps'] * 291 / 300)
                ego_vx_path = os.path.join(self.scenario_unit_folder, scenario_id, '01_Data', 'General', 'pred_ego.csv')
                ego_vx = pd.read_csv(ego_vx_path, index_col=False)
                distance += np.ceil((ego_vx['time_stamp'].max() - ego_vx['time_stamp'].min()) * ego_vx['ego_vx'].mean() / 1000)

            total_frame += frame
            total_distance += distance
            scenario_list[scenario_tag] = f'帧数-{frame:.0f} frame, 里程-{distance:.0f} km     ------- Page {3 + 12 * i}'

        test_info = {
            '测试版本': current_version,
            '版本对比': f'{current_version} <--> {old_version}',
            # 'test_date': self.test_config['test_date'],
            '测试对象': ';'.join(self.test_config['test_item'].keys()),
            '目标类型': ';'.join([f'{k}-{v}' for k, v in target_num.items()]),
            '指标类型': metric_key,
            '区域划分': region_key,
            '特征类型': characteristic_key,
            '测试场景类型': scenario_list,
        }
        plot_path = os.path.join(visualization_folder, '测试信息汇总.png')
        send_log(self, f'{plot_path} 图片已保存')
        self.test_result['OutputResult']['test_info'] = self.get_relpath(plot_path)
        plot_test_info(test_info, plot_path)

        title_summary = {
            '回灌帧数\n(千帧)': f'{(total_frame / 1000):.0f}',
            '场景里程\n(千米)': f'{total_distance:.0f}',
            '场景类型\n(类)': len(self.test_result['TagCombination'].keys()),
        }
        for ru, num in ru_num.items():
            title_summary[f'{ru}\n(千)'] = f'{(num / 1000):.0f}'
        plot_path = os.path.join(visualization_folder, '测试报告首页.png')
        plot_title_table(title_summary, plot_path)
        self.test_result['OutputResult']['test_title'] = self.get_relpath(plot_path)
        # ====================

        # 将速度图和地图都移动过来
        for scenario_tag in self.test_result['TagCombination']:
            scenario_ids = self.test_result['TagCombination'][scenario_tag]['scenario_id']
            for scenario_id in scenario_ids:
                for root, dirs, files in os.walk(os.path.join(self.scenario_unit_folder, scenario_id)):
                    for file in files:
                        if '_Info.jpg' in file:
                            file_path = os.path.join(root, file)
                            os.makedirs(os.path.join(visualization_folder, scenario_tag), exist_ok=True)
                            new_file_path = os.path.join(visualization_folder, scenario_tag, f'{scenario_id}.jpg')
                            shutil.copy(file_path, new_file_path)
        # ====================

        # 按照场景和目标特征，生成各个指标的场景
        statistics = pd.read_csv(self.get_abspath(self.test_result['OutputResult']['statistics']), index_col=False)
        group_columns = ['scenario_tag', 'topic', 'obstacle_type', 'characteristic', 'metric']
        stat_group = statistics.groupby(group_columns)
        self.test_result['OutputResult']['visualization'] = {}

        df_tp_error = {}
        for df_name, df in stat_group:
            scenario_tag, topic, obstacle_type, characteristic, metric = df_name
            if characteristic not in self.test_result['OutputResult']['report_table'].keys():
                continue

            if characteristic == '静止目标' and '速度' in metric:
                continue

            characteristic_folder = os.path.join(visualization_folder, scenario_tag, characteristic)
            if not os.path.exists(characteristic_folder):
                os.makedirs(characteristic_folder)

            for index, row in df.iterrows():
                result_dict = json.loads(row['result'])
                for key, value in result_dict.items():
                    df.at[index, key] = value
            df['type_cate'] = pd.Categorical(df['obstacle_type'],
                                             categories=['小车', '大车', '两轮车', '行人'], ordered=True)
            df.sort_values(by=['type_cate', 'region'], inplace=True)
            drop_columns = group_columns + ['index', 'type_cate', 'result']
            df.drop(drop_columns, axis=1, inplace=True)
            df.insert(0, 'obstacle_type', obstacle_type)
            df.rename(columns={'region': 'Test Area[m]'}, inplace=True)

            columns = list(df.columns)
            columns_width = [max(get_string_display_length(c) / 5.2, 2) for c in columns]
            values = df.values

            fig_size = (np.sum(columns_width) + 0.1, (len(df) + 2) + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.02, right=0.98, top=0.98, bottom=0.02)
            plt.axis('off')
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            plot_table(ax, columns, columns_width, values)
            ax.text(np.sum(columns_width) / 2, 1.5, '-'.join(
                [scenario_tag, topic, characteristic, obstacle_type, metric]),
                    va='center', ha='center', fontsize=font_size * 2)

            pic_name = '--'.join(
                [scenario_tag, topic, characteristic, obstacle_type, metric]).replace('/', '')
            pic_path = os.path.join(characteristic_folder, f'{pic_name}.jpg')
            send_log(self, f'{pic_name} 已保存')
            canvas = FigureCanvas(fig)
            canvas.print_figure(pic_path, facecolor='white', dpi=100)
            fig.clf()
            plt.close()

            if scenario_tag not in df_tp_error.keys():
                df_tp_error[scenario_tag] = {}
            if topic not in df_tp_error[scenario_tag].keys():
                df_tp_error[scenario_tag][topic] = {}
            if characteristic not in df_tp_error[scenario_tag][topic].keys():
                df_tp_error[scenario_tag][topic][characteristic] = {}
            if obstacle_type not in df_tp_error[scenario_tag][topic][characteristic].keys():
                df_tp_error[scenario_tag][topic][characteristic][obstacle_type] = []

            # 汇总准召率
            if '准召信息' in metric:
                df.rename(columns={
                    'recall%': '召回率',
                    'precision%': '准确率',
                    'type_accuracy%': '类型准确率'}, inplace=True)
                df_tp_error[scenario_tag][topic][characteristic][obstacle_type].append(
                    df[['Test Area[m]', '召回率', '准确率', '类型准确率']])

            # 汇总tp_error
            else:
                df.rename(columns={'pass_ratio%': metric}, inplace=True)
                df_tp_error[scenario_tag][topic][characteristic][obstacle_type].append(
                    df[['Test Area[m]', metric]])
        # ====================

        # 合并tp_error的metric数据，以及可视化为饼图
        for scenario_tag in df_tp_error.keys():
            if scenario_tag not in self.test_result['OutputResult']['visualization'].keys():
                self.test_result['OutputResult']['visualization'][scenario_tag] = {}

            for topic in df_tp_error[scenario_tag].keys():
                if topic not in self.test_result['OutputResult']['visualization'][scenario_tag].keys():
                    self.test_result['OutputResult']['visualization'][scenario_tag][topic] = {}

                for characteristic in df_tp_error[scenario_tag][topic].keys():
                    if characteristic not in self.test_result['OutputResult']['visualization'][scenario_tag][
                        topic].keys():
                        self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic] = {}

                    for obstacle_type in df_tp_error[scenario_tag][topic][characteristic].keys():
                        if obstacle_type not in self.test_result['OutputResult']['visualization'][scenario_tag][topic][
                            characteristic].keys():
                            self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic][
                                obstacle_type] = {}

                        # 合并表格
                        merged_df = pd.DataFrame(columns=['Test Area[m]'])
                        for one_df in df_tp_error[scenario_tag][topic][characteristic][obstacle_type]:
                            merged_df = merged_df.merge(one_df, on='Test Area[m]', how='outer')

                        # 重新排列结果的顺序
                        merged_df = merged_df.set_index('Test Area[m]', drop=True)
                        result_column = test_encyclopaedia['Information'][self.test_topic]['result_column']

                        merged_df = merged_df[[c for c in result_column if c in merged_df.columns]]

                        csv_name = '--'.join([scenario_tag, topic, characteristic, obstacle_type]).replace('/', '')
                        csv_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{csv_name}.csv')
                        (merged_df.sort_values(by=['Test Area[m]'])
                         .to_csv(csv_path, index=True, encoding='utf_8_sig'))
                        send_log(self, f'汇总百分比 {csv_name} 数据已保存')

                        self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic][
                            obstacle_type]['data'] = self.get_relpath(csv_path)

                        # 画结果大饼图
                        height_ratios = [1] + [3] * len(merged_df)  # 第一行高度是1/3，其余行高度是1
                        width_ratios = [2] + [3] * len(merged_df.columns)  # 第一列宽度是1/3，其余列宽度是1
                        fig = plt.figure(figsize=(sum(width_ratios), sum(height_ratios)))
                        fig.tight_layout()
                        plt.subplots_adjust(left=0.02, right=0.98, top=0.98, bottom=0.02)

                        # 创建GridSpec对象
                        grid = plt.GridSpec(len(height_ratios), len(width_ratios),
                                            height_ratios=height_ratios, width_ratios=width_ratios,
                                            wspace=0, hspace=0)

                        # 先画index和columns
                        for i, col in enumerate(merged_df.columns):
                            ax = fig.add_subplot(grid[0, i + 1])
                            ax.xaxis.set_visible(False)
                            ax.yaxis.set_visible(False)
                            for spine in ax.spines.values():
                                spine.set_edgecolor('black')
                                spine.set_linewidth(2)

                            ax.text(0.5, 0.5, col, va='center', ha='center',
                                    fontsize=font_size * 2.5)

                        for i, idx in enumerate(merged_df.index):
                            ax = fig.add_subplot(grid[i + 1, 0])
                            ax.xaxis.set_visible(False)
                            ax.yaxis.set_visible(False)
                            for spine in ax.spines.values():
                                spine.set_edgecolor('black')
                                spine.set_linewidth(0)

                            ax.text(0.5, 0.5, idx.replace(',', '\n'), va='center', ha='center',
                                    rotation=45, fontsize=font_size * 2.5)

                        # 画单独的饼图
                        cmap = LinearSegmentedColormap.from_list('red_to_green', ['darkred', 'darkgreen'])
                        value_shape = merged_df.values.shape
                        for i in range(value_shape[0]):
                            for j in range(value_shape[1]):
                                value = merged_df.values[i, j]
                                if not np.isnan(value):
                                    ax = fig.add_subplot(grid[i + 1, j + 1])
                                    ax.axis('off')
                                    sizes = [value, 1 - value]
                                    colors = ['limegreen', 'lightcoral']
                                    wedgeprops = {'width': 0.4}  # 设置扇区宽度为0.3，得到一个空心的效果
                                    ax.pie(sizes, colors=colors, startangle=90, wedgeprops=wedgeprops)
                                    ax.text(0, 0, f'{value:.2%}', va='center', ha='center',
                                            fontsize=font_size * 2, color=cmap(value))

                        pic_name = '--'.join(
                            [scenario_tag, topic, characteristic, obstacle_type]).replace('/', '')
                        pic_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{pic_name}.jpg')
                        send_log(self, f'{pic_name} 已保存')
                        canvas = FigureCanvas(fig)
                        canvas.print_figure(pic_path, facecolor='white', dpi=100)
                        fig.clf()
                        plt.close()
                        send_log(self, f'汇总百分比 {pic_name} 图片已保存')

                        self.test_result['OutputResult']['visualization'][scenario_tag][topic][characteristic][
                            obstacle_type]['plot'] = self.get_relpath(pic_path)
        # ====================

        # 生成用于汇报的图
        def plot_single_metric(sub_data, old_data, plot_path):

            def plot_rect(left_bottom_pt, width, height,
                          facecolor, text, text_angle,
                          font_size, alpha=1, comparison=None):
                ratio = 1
                if (isinstance(comparison, float) or isinstance(comparison, int)) and len(text):
                    ratio = 0.75
                    if comparison < 0:
                        comparison_color = ['darkred', 'darkgreen'][metric_name != '准召信息']
                    else:
                        comparison_color = ['darkgreen', 'darkred'][metric_name != '准召信息']

                    if '%' in text:
                        comparison_text = f'{comparison:.2%}'
                    elif round(comparison, 0) == comparison:
                        comparison_text = f'{comparison:.0f}'
                    else:
                        comparison_text = f'{comparison:.2f}'

                    ax.add_patch(pc.Rectangle(
                        xy=(left_bottom_pt[0] + ratio * width, left_bottom_pt[1]),
                        width=width - ratio * width, height=height,
                        angle=0, alpha=alpha,
                        rotation_point='xy',
                        fill=True,
                        edgecolor='grey',
                        facecolor='white',
                        linewidth=1))
                    ax.text(left_bottom_pt[0] + (1 + ratio) * width / 2, left_bottom_pt[1] + height / 2,
                            comparison_text,
                            va='center', ha='center', rotation=270,
                            fontdict={
                                'family': 'sans-serif',  # 字体家族
                                'color': comparison_color,  # 字体颜色
                                'weight': 'bold',  # 字体粗细
                                'size': font_size * 0.7,  # 字体大小
                                'style': 'normal'  # 字体样式
                            })

                ax.add_patch(pc.Rectangle(
                    xy=left_bottom_pt,
                    width=width * ratio, height=height,
                    angle=0, alpha=alpha,
                    rotation_point='xy',
                    fill=True,
                    edgecolor='grey',
                    facecolor=facecolor,
                    linewidth=1))
                va, ha = 'center', 'center'
                ax.text(left_bottom_pt[0] + ratio * width / 2, left_bottom_pt[1] + height / 2, text,
                        va=va, ha=ha, rotation=text_angle,
                        fontsize=font_size)

            color_for_sheet = []
            cell_w = 1
            cell_h = 0.5
            font_size = 11
            fig_size = ((len(sub_data.columns) + 2) * cell_w + 0.1, (len(sub_data) + 3) * cell_h + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            ax.axis('off')

            metric_name = sub_data.columns[0][0]
            tag_name = sub_data.index[0][2]
            ax.text((len(sub_data.columns) - 2) / 2, cell_h * 2.5, f'<{tag_name}> {metric_name}',
                    va='center', ha='center', rotation=0,
                    fontsize=font_size * 2, color='peru')

            # 原点放置在数据的左上角
            # 画index
            plot_rect(left_bottom_pt=(-2 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='目标类型', text_angle=0, font_size=font_size)
            plot_rect(left_bottom_pt=(-1 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='测试范围[m]', text_angle=0, font_size=font_size)

            last_index_0 = sub_data.index[0][0]
            index_0_x, index_0_y = -2 * cell_w, 0
            index_width, index_height = cell_w, 0
            for i, index in enumerate(sub_data.index):
                if last_index_0 != index[0]:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)
                    index_height = 0
                index_height += cell_h
                index_0_y -= cell_h
                last_index_0 = index[0]
                if i == len(sub_data.index) - 1:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)

                plot_rect(left_bottom_pt=(-1 * cell_w, - (i + 1) * cell_h),
                          width=cell_w, height=cell_h, facecolor='bisque',
                          text=index[1].replace(',', '\n'), text_angle=0, font_size=font_size)

            # 画columns和data
            last_col_0 = sub_data.columns[0][1]
            col_0_x, col_0_y = 0, cell_h
            col_width, col_height = 0, cell_h
            for i, col in enumerate(sub_data.columns):
                if last_col_0 != col[1]:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)
                    col_0_x += col_width
                    col_width = 0
                col_width += cell_w
                last_col_0 = col[1]
                if i == len(sub_data.columns) - 1:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)

                if col[2] != 'kpi_target':
                    plot_rect(left_bottom_pt=(i * cell_w, 0),
                              width=cell_w, height=cell_h, facecolor='lightsteelblue',
                              text=wrap_text(col[2], 10), text_angle=0, font_size=font_size)
                    for j in range(len(sub_data.index)):
                        value = sub_data[col].tolist()[j]
                        if (col[0], col[1], 'kpi_target') in sub_data.columns:
                            kpi_target = sub_data[(col[0], col[1], 'kpi_target')].tolist()[j]
                            if np.isnan(value):
                                text = ''
                                color = 'lightgrey'
                            else:
                                if '%' in col[1]:
                                    text = f'{value:.2%}'
                                elif round(value, 0) == value:
                                    text = f'{value:.0f}'
                                else:
                                    text = f'{value:.2f}'
                                if np.isnan(kpi_target):
                                    color = 'white'
                                else:
                                    if value < kpi_target:
                                        color = ['lightcoral', 'limegreen'][metric_name != '准召信息']
                                    else:
                                        color = ['limegreen', 'lightcoral'][metric_name != '准召信息']
                        else:
                            if np.isnan(value):
                                text = ''
                                color = 'lightgrey'
                            else:
                                color = 'white'
                                if '%' in col[1]:
                                    text = f'{value:.2%}'
                                elif round(value, 0) == value:
                                    text = f'{value:.0f}'
                                else:
                                    text = f'{value:.2f}'

                        # 用于为excel表格涂色
                        color_for_sheet.append([list(col), j, color])

                        #取出old_data中相应格子的值
                        if old_data is None:
                            comparison = None
                        else:
                            if col not in old_data.columns:
                                comparison = None
                            else:
                                old_value = old_data[col].tolist()[j]
                                if np.isnan(old_value) or np.isnan(value):
                                    comparison = None
                                else:
                                    comparison = value - old_value

                        plot_rect(left_bottom_pt=(i * cell_w, - j * cell_h - cell_h),
                                  width=cell_w, height=cell_h, facecolor=color,
                                  text=text, text_angle=0, font_size=font_size * 1.2, comparison=comparison)

                else:
                    plot_rect(left_bottom_pt=(i * cell_w, 0),
                              width=cell_w, height=cell_h, facecolor='lightsteelblue',
                              text='指标期望', text_angle=0, font_size=font_size * 1.3)
                    for j in range(len(sub_data.index)):
                        value = sub_data.values[j, i]
                        color = 'lightcyan'
                        if np.isnan(value):
                            text = ''
                        elif '%' in col[1]:
                            text = f'{value:.2%}'
                        elif round(value, 0) == value:
                            text = f'{value:.0f}'
                        else:
                            text = f'{value:.2f}'
                        plot_rect(left_bottom_pt=(i * cell_w, - j * cell_h - cell_h),
                                  width=cell_w, height=cell_h, facecolor=color,
                                  text=text, text_angle=0, font_size=font_size * 1.2)

                        # 用于为excel表格涂色
                        color_for_sheet.append([list(col), j, color])

            ax.set_xlim(-2 * cell_w - 0.1, len(sub_data.columns) * cell_w + 0.1)
            ax.set_ylim(- len(sub_data) * cell_h - 0.1, 3 * cell_h + 0.1)
            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)

            return color_for_sheet

        def plot_summary_metric(summary_data, plot_path):

            def plot_rect(left_bottom_pt, width, height,
                          facecolor, text, text_angle,
                          font_size, fill=True):

                ax.add_patch(pc.Rectangle(
                    xy=left_bottom_pt,
                    width=width, height=height,
                    angle=0, alpha=1,
                    rotation_point='xy',
                    fill=fill,
                    edgecolor='grey',
                    facecolor=facecolor,
                    linewidth=1))
                va, ha = 'center', 'center'
                ax.text(left_bottom_pt[0] + width / 2, left_bottom_pt[1] + height / 2, text,
                        va=va, ha=ha, rotation=text_angle,
                        fontsize=font_size)

            cell_w = 1
            cell_h = 0.5
            font_size = 11
            fig_size = ((len(summary_data.columns) + 2) * cell_w + 0.1, (len(summary_data) + 3) * cell_h + 0.1)
            fig = plt.figure(figsize=fig_size)
            plt.tight_layout()
            plt.subplots_adjust(left=0.01, right=0.99, top=0.99, bottom=0.01)
            grid = plt.GridSpec(1, 1, wspace=0, hspace=0.02)
            ax = fig.add_subplot(grid[0, 0])
            ax.axis('off')

            tag_name = summary_data.index[0][2]
            ax.text((len(summary_data.columns) - 2) / 2, cell_h * 2.5, tag_name,
                    va='center', ha='center', rotation=0,
                    fontsize=font_size * 2, color='peru')

            # 原点放置在数据的左上角
            # 画index
            plot_rect(left_bottom_pt=(-2 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='目标类型', text_angle=0, font_size=font_size)
            plot_rect(left_bottom_pt=(-1 * cell_w, 0),
                      width=cell_w, height=cell_h, facecolor='lightsteelblue',
                      text='测试范围[m]', text_angle=0, font_size=font_size)

            last_index_0 = summary_data.index[0][0]
            index_0_x, index_0_y = -2 * cell_w, 0
            index_width, index_height = cell_w, 0
            for i, index in enumerate(summary_data.index):
                if last_index_0 != index[0]:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)
                    index_height = 0
                index_height += cell_h
                index_0_y -= cell_h
                last_index_0 = index[0]
                if i == len(summary_data.index) - 1:
                    plot_rect(left_bottom_pt=(index_0_x, index_0_y),
                              width=index_width, height=index_height, facecolor='bisque',
                              text=last_index_0, text_angle=0, font_size=font_size * 1.2)

                plot_rect(left_bottom_pt=(-1 * cell_w, - (i + 1) * cell_h),
                          width=cell_w, height=cell_h, facecolor='bisque',
                          text=index[1].replace(',', '\n'), text_angle=0, font_size=font_size)

            # 画columns和data
            last_col_0 = summary_data.columns[0][1]
            col_0_x, col_0_y = 0, cell_h
            col_width, col_height = 0, cell_h
            for i, col in enumerate(summary_data.columns):
                if last_col_0 != col[1]:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)
                    col_0_x += col_width
                    col_width = 0
                col_width += cell_w
                last_col_0 = col[1]
                if i == len(summary_data.columns) - 1:
                    plot_rect(left_bottom_pt=(col_0_x, col_0_y),
                              width=col_width, height=col_height, facecolor='lightsteelblue',
                              text=last_col_0, text_angle=0, font_size=font_size * 1.2)

                plot_rect(left_bottom_pt=(i * cell_w, 0),
                          width=cell_w, height=cell_h, facecolor='lightsteelblue',
                          text=wrap_text(col[2], 10), text_angle=0, font_size=font_size)

                good_face_image = Image.open(os.path.join(project_path, 'Docs', 'Resources', 'Icon', 'good_face.png'))
                bad_face_image = Image.open(os.path.join(project_path, 'Docs', 'Resources', 'Icon', 'bad_face.png'))

                for j in range(len(summary_data.index)):
                    plot_rect(left_bottom_pt=(i * cell_w, -j * cell_h - cell_h),
                              width=col_width, height=col_height, facecolor='white',
                              text='', text_angle=0, fill=False, font_size=font_size * 1.2)

                    value = summary_data.values[j, i]
                    extend = (i * cell_w + (cell_w - cell_h) / 2, i * cell_w + cell_w - (cell_w - cell_h) / 2,
                              -j * cell_h - cell_h, -j * cell_h)
                    if value == 1:
                        ax.imshow(good_face_image, extent=extend)
                    elif value == 0:
                        ax.imshow(bad_face_image, extent=extend)

            ax.set_xlim(-2 * cell_w - 0.1, len(summary_data.columns) * cell_w + 0.1)
            ax.set_ylim(- len(summary_data) * cell_h - 0.1, 3 * cell_h + 0.1)
            canvas = FigureCanvas(fig)
            canvas.print_figure(plot_path, facecolor='white', dpi=100)

        # 获取版本对比的文件夹
        version_comparison_folder = self.test_config['version_comparison']
        if version_comparison_folder is not None and os.path.exists(version_comparison_folder):
            comparison_valid = 1
        else:
            comparison_valid = 0

        metric_pass_ratio_by_characteristic_scenario = {} # 用于统计各特征和场景下的指标合格率
        color_for_sheets = {} # excel的涂色字典
        self.test_result['OutputResult']['report_plot'] = {}
        for characteristic, excel_path in self.test_result['OutputResult']['report_table'].items():
            excel_path = self.get_abspath(excel_path)
            self.test_result['OutputResult']['report_plot'][characteristic] = {}
            color_for_sheets[characteristic] = {}
            metric_pass_ratio_by_characteristic_scenario[characteristic] = {}

            # 寻找其他版本相同的文件
            old_excel_path = None
            if comparison_valid:
                for root, dirs, files in os.walk(version_comparison_folder):
                    for file in files:
                        if os.path.basename(excel_path) in file and self.test_topic in root:
                            old_excel_path = os.path.join(root, file)

            for _, scenario_tag in enumerate(pd.ExcelFile(excel_path).sheet_names):
                data = pd.read_excel(excel_path, sheet_name=scenario_tag, header=[0, 1, 2], index_col=[0, 1, 2])
                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag] = {}
                color_for_sheets[characteristic][scenario_tag] = []

                old_data = None
                if old_excel_path is not None and scenario_tag in pd.ExcelFile(excel_path).sheet_names:
                    old_data = pd.read_excel(old_excel_path, sheet_name=scenario_tag, header=[0, 1, 2], index_col=[0, 1, 2])

                sub_data_col = {}
                for col in data.columns:
                    if col[0] not in sub_data_col:
                        sub_data_col[col[0]] = []
                    sub_data_col[col[0]].append(col)

                # 获取汇总笑脸图，以及分指标kpi图
                pass_or_fail_summary = []
                for metric, cols in sub_data_col.items():

                    if metric == '准召信息':
                        new_cols = []
                        for col in cols:
                            if col[1] not in ['TP', 'CTP', 'FP', 'FN', 'sample_count']:
                                new_cols.append(col)
                        sub_data = data[new_cols]

                        # 将sample数量也单独绘制
                        new_cols = []
                        for col in cols:
                            if col[1] in ['TP', 'FP', 'FN', 'CTP']:
                                new_cols.append(col)
                        sample_data = data[new_cols]
                        sample_plot_path = os.path.join(visualization_folder, scenario_tag, characteristic,
                                                 f'{scenario_tag}--{characteristic}--样本数量.png')
                        plot_single_metric(sample_data, old_data, sample_plot_path)
                        self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag][
                            '样本数量'] = self.get_relpath(sample_plot_path)
                        send_log(self, f'{sample_plot_path} 图片已保存')

                        __metrics = list(dict.fromkeys(sub_data.columns.get_level_values(1)))
                        for __metric in __metrics:
                            temp_data = sub_data.loc[:,
                                        [(col_0, col_1, col_2) for col_0, col_1, col_2 in sub_data.columns if col_1 == __metric]]
                            pass_or_fails = []
                            for idx, row in temp_data.iterrows():
                                pass_or_fail = []
                                for t0, t1, topic in row.index:
                                    if topic == 'kpi_target':
                                        continue
                                    if np.isnan(row[(t0, t1, topic)]) or np.isnan(row[(t0, t1, 'kpi_target')]):
                                        pass_or_fail.append(-1)
                                    elif row[(t0, t1, topic)] > row[(t0, t1, 'kpi_target')]:
                                        pass_or_fail.append(1)
                                    else:
                                        pass_or_fail.append(0)
                                pass_or_fails.append(pass_or_fail)

                            new_data = sub_data.loc[:, [(col_0, col_1, col_2) for col_0, col_1, col_2 in temp_data.columns if
                                                        col_2 != 'kpi_target']]
                            pass_or_fail_data = pd.DataFrame(pass_or_fails, columns=new_data.columns, index=new_data.index)
                            pass_or_fail_summary.append(pass_or_fail_data)

                    else:
                        sub_data = data[cols]
                        __metrics = list(dict.fromkeys(sub_data.columns.get_level_values(1)))
                        pass_or_fails_compare = []
                        for __metric in __metrics:
                            temp_data = sub_data.loc[:,
                                        [(col_0, col_1, col_2) for col_0, col_1, col_2 in sub_data.columns if col_1 == __metric]]
                            pass_or_fails = []
                            for idx, row in temp_data.iterrows():
                                pass_or_fail = []
                                for t0, t1, topic in row.index:
                                    if topic == 'kpi_target':
                                        continue
                                    if np.isnan(row[(t0, t1, topic)]) or np.isnan(row[(t0, t1, 'kpi_target')]):
                                        pass_or_fail.append(-1)
                                    elif row[(t0, t1, topic)] > row[(t0, t1, 'kpi_target')]:
                                        pass_or_fail.append(0)
                                    else:
                                        pass_or_fail.append(1)
                                pass_or_fails.append(pass_or_fail)
                            pass_or_fails_compare.append(np.array(pass_or_fails))

                        # 与或合并测试结果
                        compare_res = np.zeros_like(pass_or_fails_compare[0])
                        for ii in range(compare_res.shape[0]):
                            for jj in range(compare_res.shape[1]):
                                res = [pass_or_fails[ii, jj] for pass_or_fails in pass_or_fails_compare]
                                if 1 in res:
                                    compare_res[ii, jj] = 1
                                elif 0 in res:
                                    compare_res[ii, jj] = 0
                                else:
                                    compare_res[ii, jj] = -1

                        pass_or_fail_columns = []
                        for col_0, col_1, col_2 in sub_data.columns:
                            if col_2 == 'kpi_target':
                                continue
                            if (col_0, col_0, col_2) not in pass_or_fail_columns:
                                pass_or_fail_columns.append((col_0, col_0, col_2))
                        pass_or_fail_data = pd.DataFrame(compare_res, columns=pass_or_fail_columns, index=sub_data.index)
                        pass_or_fail_summary.append(pass_or_fail_data)

                    plot_path = os.path.join(visualization_folder, scenario_tag, characteristic,
                                             f'{scenario_tag}--{characteristic}--{metric}.png')
                    self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag][
                        metric] = self.get_relpath(plot_path)
                    color_for_sheets[characteristic][scenario_tag].extend(plot_single_metric(sub_data, old_data, plot_path))
                    send_log(self, f'{plot_path} 图片已保存')

                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary'] = []
                plot_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{scenario_tag}--{characteristic}_summary_1.png')
                metric_summary = pd.concat(pass_or_fail_summary[:5], axis=1)
                plot_summary_metric(metric_summary, plot_path)
                send_log(self, f'{plot_path} 图片已保存')
                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary'].append(self.get_relpath(plot_path))

                plot_path = os.path.join(visualization_folder, scenario_tag, characteristic, f'{scenario_tag}--{characteristic}_summary_2.png')
                metric_summary = pd.concat(pass_or_fail_summary[5:], axis=1)
                plot_summary_metric(metric_summary, plot_path)
                send_log(self, f'{plot_path} 图片已保存')
                self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary'].append(self.get_relpath(plot_path))

                metric_pass_count = np.sum(pd.concat(pass_or_fail_summary, axis=1).values == 1)
                metric_fail_count = np.sum(pd.concat(pass_or_fail_summary, axis=1).values == 0)
                metric_pass_ratio_by_characteristic_scenario[characteristic][scenario_tag] = {
                    'pass_count': int(metric_pass_count), 'fail_count': int(metric_fail_count),
                    'ratio': float(metric_pass_count / (metric_pass_count + metric_fail_count))}

        self.test_result['OutputResult']['metric_pass_ratio'] = metric_pass_ratio_by_characteristic_scenario
        # ====================

        # 根据比对结果，对excel源文件涂色
        with open(os.path.join(visualization_folder, 'color_for_sheets.yaml'), 'w', encoding='utf-8') as f:
            yaml.dump(color_for_sheets,
                      f, encoding='utf-8', allow_unicode=True, sort_keys=False)
        fill_type = {
            'lightcoral': PatternFill(fill_type="solid", start_color="FFB6C1"),
            'limegreen': PatternFill(fill_type="solid", start_color="32CD32"),
            'lightcyan': PatternFill(fill_type="solid", start_color="E0FFFF"),
            'lightgrey': PatternFill(fill_type="solid", start_color="D3D3D3"),
            'white': PatternFill(fill_type="solid", start_color="FFFFFF")
        }
        color_list = [
                         'CC4422', '33FF57', '2244CC', 'CC2288', '8822CC',
                         'CCAA00', '8B4513', 'D45A4A', '6600A3', 'A68C67',
                     ] * 20

        for characteristic, file in self.test_result['OutputResult']['report_table'].items():
            workbook = openpyxl.load_workbook(self.get_abspath(file))
            for scenario_tag in workbook.sheetnames:
                sheet = workbook[scenario_tag]
                sheet_colors = color_for_sheets[characteristic][scenario_tag]
                max_column = sheet.max_column

                column, i = 1, 0
                while column <= max_column:
                    if sheet.cell(row=1, column=column).value is not None:
                        i += 1
                    sheet.cell(row=1, column=column).fill = PatternFill(fill_type="solid", start_color=color_list[i])
                    column += 1

                for color_item in sheet_colors:
                    column = 1
                    while sheet.cell(row=1, column=column).value != color_item[0][0] and column <= max_column:
                        column += 1
                    while sheet.cell(row=2, column=column).value != color_item[0][1] and column <= max_column:
                        column += 1
                    while sheet.cell(row=3, column=column).value != color_item[0][2] and column <= max_column:
                        column += 1
                    if column <= max_column:
                        cell = sheet.cell(row=color_item[1] + 4, column=column)
                        cell.fill = fill_type[color_item[2]]
                        cell.border = Border(
                            left=Side(style='thin', color="000000"),
                            right=Side(style='thin', color="000000"),
                            top=Side(style='thin', color="000000"),
                            bottom=Side(style='thin', color="000000")
                        )
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            workbook.save(os.path.join(visualization_folder, f'{characteristic}.xlsx'))

    @sync_test_result
    def gen_report(self):

        report_title = f'{self.product}_{self.test_topic}_数据回灌感知测试报告'
        send_log(self, f'开始生成报告 {report_title}')

        title_background = os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'TitlePage.png')
        logo = os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'ZoneLogo.png')
        report_generator = PDFReportTemplate(report_title=report_title,
                                             test_time=self.test_date.split(' ')[0],
                                             tester='Hil_DataReplay_TestStand',
                                             version=self.version,
                                             title_page=title_background,
                                             title_summary_img=self.get_abspath(self.test_result['OutputResult']['test_title']),
                                             logo=logo)

        heading = '报告信息总览'
        send_log(self, f'生成页面 {heading}')
        report_generator.addOnePage(
            heading=heading,
            text_list=None,
            img_list=[[self.get_abspath(self.test_result['OutputResult']['test_info'])]],
        )

        for characteristic in self.test_result['OutputResult']['report_plot'].keys():

            if characteristic != '关键目标':
                continue

            for scenario_tag in self.test_result['OutputResult']['report_plot'][characteristic]:

                send_log(self, f'生成页面 {characteristic} {scenario_tag}')
                report_generator.addTitlePage(
                    title=f'{scenario_tag}<{characteristic}>',
                    page_type='sequence',
                    stamp=None,
                    sub_title_list=None,
                )

                for img in self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag]['summary']:

                    send_log(self, f'生成页面 {characteristic} {scenario_tag} Overview')
                    report_generator.addOnePage(
                        heading=f'{scenario_tag}<{characteristic}> Overview',
                        text_list=['绿色笑脸 = Pass, 红色哭脸 = Fail'],
                        img_list=[[self.get_abspath(img)]],
                    )

                for metric, img in self.test_result['OutputResult']['report_plot'][characteristic][scenario_tag].items():
                    if metric == 'summary':
                        continue

                    send_log(self, f'生成页面 {characteristic} {scenario_tag} {metric}')
                    report_generator.addOnePage(
                        heading=f'{scenario_tag}<{characteristic}> {metric}',
                        text_list=['绿色代表“Pass”和“好于”，红色代表“Fail”和“差于”', '侧边竖写数字为与对比版本的差值'],
                        img_list=[[self.get_abspath(img)]],
                    )

        report_generator.addTitlePage(title='附 录',
                                      page_type='sequence',
                                      stamp=None,
                                      sub_title_list=None)

        for scenario_tag in self.test_result['OutputResult']['visualization'].keys():

            # 测试场景展示页面
            img_count = 0
            img_list = []
            for i, scenario_id in enumerate(self.test_result['TagCombination'][scenario_tag]['scenario_id']):
                scenario_info_img = os.path.join(self.output_result_folder, 'visualization', scenario_tag, f'{scenario_id}.jpg')
                img_list.append([scenario_info_img])
                img_count += 1
                if img_count == 3 or i == len(self.test_result['TagCombination'][scenario_tag]['scenario_id'])-1:
                    report_generator.addOnePage(heading=f'{scenario_tag} 测试场景',
                                                text_list=None,
                                                img_list=img_list)
                    img_list = []
                    img_count = 0

        # 测试环境
        heading = '测试环境——硬件在环 | 数采回灌'
        text = ['将真实控制器部署在仿真测试台架中，将实车路采的传感器数据, '
                '通过回灌测试设备回放及注入智驾控制器, 模拟运行场景和控制对象:',
                'A. 实车路采数据: 实车运行时采集的视频图片, 总线Can, Lidar点云等实时数据.',
                'B. 数据回灌设备: 数采数据后处理, 信号同步, 编码解码等, 模拟传感器输出, 将数据注入控制器.',
                'C. 智驾控制器: 接收模拟传感器信号, 模拟实车运行场景, 输出算法感知结果.',
                'D. ECU-Test，实现测试环境自动化运行及数据自动化分析，用于测试用例的执行和测试结果的生成.', ]
        img = [
            [os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'DataReplayTestEnv.png')]
        ]
        report_generator.addOnePage(heading=heading,
                                    text_list=text,
                                    img_list=img)

        # 测试原理
        heading = '测试原理——硬件在环 | 数采回灌'
        text = ['对比感知算法结果与参考真值, 计算性能指标, 输出评价结果:',
                'A. 以标注平台或具备更高感知能力的工具基于数采数据提供标注真值结果作为GroundTruth(GT).',
                'B. 控制器感知节点订阅信息提供场景各要素特征测量值，MeasuredValue(MV).',
                'C. 对比并评价感知信息是否符合预期，包括准召率，定位精度, 几何精度, 类型准确度等.',
                '--------------------',
                '测试结果局限性:',
                'A. 测试结果的置信度受限于真值结果对于真实世界的准确程度.',
                'B. 测试结果的全面性依赖于数采数据的样本规模, 需要时间的积累.']
        img = [
            [os.path.join(get_project_path(), 'Docs', 'Resources', 'report_figure', 'DataReplayTestPrinciple.png')]
        ]
        report_generator.addOnePage(heading=heading,
                                    text_list=text,
                                    img_list=img)

        report_generator.page_count += 0
        self.report_path = os.path.join(self.task_folder, f'{self.product}_{self.test_topic}_数据回灌测试报告({self.test_config["version"]}).pdf')
        report_generator.genReport(report_path=self.report_path, compress=1)

    def start(self):
        if (self.test_action['scenario_unit']['preprocess']
                or self.test_action['scenario_unit']['match']
                or self.test_action['scenario_unit']['metric']
                or self.test_action['scenario_unit']['bug']
                or self.test_action['scenario_unit']['render']):
            self.analyze_scenario_unit()

        if self.test_action['tag_combination']:
            self.combine_scenario_tag()

        if self.test_action['statistics']:
            self.summary_bug_items()
            # self.compile_statistics()

        if self.test_action['visualization']:
            self.visualize_output()

        if self.test_action['gen_report']:
            self.gen_report()